from main_file import decrypt_ds2_sl2, encrypt_modified_files
from main_file_import import decrypt_ds2_sl2_import
import json, shutil, os , struct
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from pathlib import Path
# openpyxl is lazy-loaded in export/import functions to speed up startup
from typing import Optional, Any
import sys
import threading
import re
import logging
# project modules
from log_config import setup_logging
from basic_class import Item
import globals
from globals import (ITEM_TYPE_RELIC, ITEM_TYPE_GOODS,
                     WORKING_DIR, COLOR_MAP)

from relic_checker import RelicChecker, InvalidReason, is_curse_invalid
from source_data_handler import SourceDataHandler, get_system_language
from vessel_handler import LoadoutHandler, is_vessel_unlocked


def get_base_dir():
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent
    else:
        return Path(__file__).resolve().parent


# Setup Logger
setup_logging(get_base_dir().joinpath("logs").as_posix())
logger = logging.getLogger(__name__)


# Global variables
os.chdir(WORKING_DIR)

# Data storage - SourceDataHandler and LoadoutHandler are lazy-initialized to speed up startup
data_source: Optional[SourceDataHandler] = None
relic_checker: Optional[RelicChecker] = None
loadout_handler: Optional[LoadoutHandler] = None
items_json = {}
effects_json = {}
userdata_path = None


def _ensure_data_source():
    """Lazy initialize data_source on first use"""
    global data_source, loadout_handler
    if data_source is None:
        logger.info("Initializing SourceDataHandler")
        data_source = SourceDataHandler(language=get_system_language())


# Config file path for remembering last opened file
CONFIG_FILE = os.path.join(get_base_dir(), "editor_config.json")


def load_config():
    """Load saved configuration"""
    try:
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, 'r') as f:
                return json.load(f)
    except Exception as e:
        logger.warning("Could not load config file: %s", e)
    return {}

def save_config(config):
    """Save configuration to file"""
    try:
        with open(CONFIG_FILE, 'w') as f:
            json.dump(config, f, indent=2)
    except Exception as e:
        logger.warning("Could not save config file: %s", e)

def autosize_treeview_columns(tree, padding=20, min_width=50):
    """Auto-size treeview columns based on content width, prevent stretching"""
    import tkinter.font as tkfont

    for col in tree['columns']:
        # Get header text width
        header_text = tree.heading(col)['text']
        header_font = tkfont.nametofont('TkHeadingFont')
        max_content_width = header_font.measure(header_text)

        # Check all items for max content width (sample first 100 for performance)
        default_font = tkfont.nametofont('TkDefaultFont')
        children = tree.get_children()[:100]
        for item in children:
            values = tree.item(item, 'values')
            col_idx = tree['columns'].index(col)
            if col_idx < len(values):
                text = str(values[col_idx])
                text_width = default_font.measure(text)
                if text_width > max_content_width:
                    max_content_width = text_width

        # Apply width with padding - size to content, no max limit
        # Set stretch=False to prevent columns from growing beyond content width
        final_width = max(min_width, max_content_width + padding)
        tree.column(col, width=final_width, minwidth=final_width, stretch=False)

imported_data=None
MODE = None
IMPORT_MODE=None
char_name_list = []
char_name_list_import = []
ga_relic = []
ga_items = []
current_murks = 0
current_sigs = 0
# AOB_search='00 00 00 00 ?? 00 00 00 ?? ?? 00 00 00 00 00 00 ??'
AOB_search='00 00 00 00 0A 00 00 00 ?? ?? 00 00 00 00 00 00 06'
from_aob_steam= 44
steam_id=None


# Vessel slot color meanings
# Red = Burning Scene relics only
# Green = Tranquil Scene relics only
# Blue = Drizzly Scene relics only
# Yellow = Luminous Scene relics only
# White = Universal (any relic)

def get_vessel_info(char_name, vessel_slot):
    """Get vessel information for a character's vessel slot.

    Uses AntiqueStandParam.csv to get:
    - Vessel name from GoodsName
    - Relic slot color requirements (normal and deep)
    - Character association

    Args:
        char_name: Character name (e.g., 'Wylder')
        vessel_slot: Vessel slot number (0-10)

    Returns:
        Dictionary with 'name' and optionally 'colors' (6 colors for normal/deep slots)
    """
    global data_source

    # Calculate vessel ID from character and slot
    # Characters 0-9 have vessels at IDs 1000-1006, 2000-2006, etc.
    # Shared vessels (slots 7-9) use IDs 19000-19002
    # Shared vessel slot 10 uses ID 19010
    char_id = data_source.character_names.index(char_name) if char_name in data_source.character_names else -1

    if char_id < 0 or char_id >= 10:
        return {'name': f"Vessel {vessel_slot}"}

    # Determine vessel ID
    if vessel_slot <= 6:
        # Character-specific vessels: 1000-1006, 2000-2006, etc.
        vessel_id = (char_id + 1) * 1000 + vessel_slot
    elif vessel_slot <= 9:
        # Shared vessels: 19000-19002 map to slots 7-9
        vessel_id = 19000 + (vessel_slot - 7)
    else:
        # Shared vessel 10: 19010
        vessel_id = 19010

    # Try to get vessel data from source_data_handler
    if data_source is not None:
        try:
            vessel_data_info = data_source.get_vessel_data(vessel_id)
            if vessel_data_info:
                return {
                    'name': vessel_data_info.get('Name', f'Vessel {vessel_slot}'),
                    'colors': vessel_data_info.get('Colors', None),
                    'character': vessel_data_info.get('Character', char_name),
                    'unlockFlag': vessel_data_info.get('unlockFlag', 0)
                }
        except Exception as e:
            logger.error("Error retrieving vessel data for ID %d: %s", vessel_id, e)

    return {'name': f"Vessel {vessel_slot}", 'unlockFlag': 0}


def load_json_data():
    global items_json, effects_json
    _ensure_data_source()  # Lazy init data_source
    try:
        items_json = data_source.get_relic_origin_structure()
        effects_json = data_source.get_effect_origin_structure()

        return True

    except FileNotFoundError as e:
        messagebox.showerror(
            "Error",
            f"JSON files not found: {str(e)}\nManual editing only available."
        )
        return False


def reload_language(language_code):
    global items_json, effects_json, data_source
    _ensure_data_source()  # Lazy init data_source
    result = data_source.reload_text(language_code)
    items_json = data_source.get_relic_origin_structure()
    effects_json = data_source.get_effect_origin_structure()
    return result


def parse_items(data_type, start_offset, slot_count=5120):
    items = []
    offset = start_offset
    logger.info("Parsing %d items starting at offset 0x%X", slot_count, start_offset)
    for _ in range(slot_count):
        item = Item.from_bytes(data_type, offset)
        items.append(item)
        offset += item.size
    return items, offset


# Global to store acquisition order mapping (GA handle -> acquisition ID from inventory entry)
ga_acquisition_order = {}

def parse_inventory_acquisition_order(data_type, items_end_offset):
    """
    Parse the inventory section to get the acquisition ID of relics.
    Each inventory entry has a 2-byte acquisition ID at offset +12 within the 14-byte entry.
    Lower acquisition ID = acquired earlier (oldest).
    Returns a dict mapping GA handle -> acquisition ID.
    """
    global ga_acquisition_order
    ga_acquisition_order = {}

    # Inventory section starts at items_end_offset + 0x650
    inventory_start = items_end_offset + 0x650
    logger.debug("Parsing inventory acquisition order from 0x%X", inventory_start)

    # Build set of known relic GA handles for quick lookup
    relic_ga_set = set()
    for relic in ga_relic:
        relic_ga_set.add(relic[0])  # relic[0] is ga_handle

    if not relic_ga_set:
        return ga_acquisition_order

    # Scan inventory section for relic GA handles
    # Inventory entries are 14 bytes: [4 bytes prefix] [4 bytes GA handle] [4 bytes unknown] [2 bytes acq_id]
    inventory_data = data_type[inventory_start:]

    # Scan for GA handles and extract acquisition ID
    for offset in range(0, min(len(inventory_data) - 14, 0x3000), 2):
        potential_ga = struct.unpack_from('<I', inventory_data, offset + 4)[0]
        if potential_ga in relic_ga_set and potential_ga not in ga_acquisition_order:
            # Extract acquisition ID from last 2 bytes of the 14-byte entry
            acq_id = struct.unpack_from('<H', inventory_data, offset + 12)[0]
            ga_acquisition_order[potential_ga] = acq_id

    return ga_acquisition_order


def parse_inventory_section(data_type, name_offset):
    globals.ga_goods = []
    globals.goods_id_list = []
    globals.ga_relics_list = []
    start_offset = name_offset + 0x5B8  # same section with parse acq order function
    base_size = 12 + 2  # 2 bytes unknown(maybe some kind of flag, sellable or favorite)
    # First 4 bytes: Item last index
    # Followed by Item structures (Base size 14 bytes):
    # - 4 bytes: ga_handle, Composite LE (Byte 0: Type 0xB0=Goods, Bytes 1-3: ID)
    # - 4 bytes: item_quantity
    # - 4 bytes: acquisition ID
    # - 1 byte: bool -> is favorite
    # - 1 byte: bool -> is sellable
    counts = struct.unpack_from("<I", data_type, start_offset)[0]
    MAX_INVENTORY_COUNT = 3071
    cursor = start_offset + 4
    i = 0
    for _ in range(MAX_INVENTORY_COUNT):
        if cursor >= len(data_type) or i >= counts:
            break
        ga_handle = struct.unpack_from("<I", data_type, cursor)[0]
        if ga_handle == 0:
            cursor += base_size
            continue
        
        iten_amount = struct.unpack_from("<I", data_type, cursor + 4)[0]
        acquisition = struct.unpack_from("<I", data_type, cursor + 8)[0]
        is_favorite, is_sellable = struct.unpack_from("<BB", data_type, cursor + 12)
        logger.debug("Item %d at 0x%X: GA Handle=0x%X, Amount=%d, Acquisition=0x%X, is_favorite=%d, is_sellable=%d",
                     i, cursor, ga_handle, iten_amount, acquisition, is_favorite, is_sellable)
        type_bits = ga_handle & 0xFF000000
        if type_bits == ITEM_TYPE_GOODS:
            goods_id = ga_handle & 0x00FFFFFF
            globals.ga_goods.append((ga_handle, goods_id))
            globals.goods_id_list.append(goods_id)
        elif type_bits == ITEM_TYPE_RELIC:
            globals.ga_relics_list.append(ga_handle)
        cursor += base_size
        i += 1
    logger.debug("counts of goods items: %d", counts)
    

def debug_ga_relic_check():
    global ga_relic
    _ga_handles = [item[0] for item in ga_relic]
    for ga_handle in globals.ga_relics_list:
        if ga_handle not in _ga_handles:
            logger.debug("Relic GA handle 0x%X found in inventory but not in ga_relic parsed items", ga_handle)


def gaprint(data_type):
    global ga_relic, ga_items
    ga_items = []
    ga_relic = []
    start_offset = 0x14
    slot_count = 5120
    items, end_offset = parse_items(data_type, start_offset, slot_count)

    for item in items:
        type_bits = item.gaitem_handle & 0xF0000000
        parsed_item = (
                item.gaitem_handle,
                item.item_id,
                item.effect_1,
                item.effect_2,
                item.effect_3,
                item.sec_effect1,
                item.sec_effect2,
                item.sec_effect3,
                item.offset,
                item.size,
            )
        ga_items.append(parsed_item)

        if type_bits == ITEM_TYPE_RELIC:
            ga_relic.append(parsed_item)

    # Parse inventory section to get acquisition order
    parse_inventory_acquisition_order(data_type, end_offset)
    
    # Parse goods for check if player had vessel unlocked
    parse_inventory_section(data_type, end_offset+0x94)  # end_offset+0x94 = name_offset
    # debug_ga_relic_check()
    return end_offset


def read_char_name(data):
    name_offset = gaprint(data) + 0x94
    max_chars = 16
    for cur in range(name_offset, name_offset + max_chars * 2, 2):
        if data[cur:cur + 2] == b'\x00\x00':
            max_chars = (cur - name_offset) // 2
            break
    raw_name = data[name_offset:name_offset + max_chars * 2]
    name = raw_name.decode("utf-16-le", errors="ignore").rstrip("\x00")
    return name if name else None


def debug_dump_complete_relic_analysis(file_data):
    """Complete deep dive analysis of relics in save file and data files."""

    output_lines = []
    def log(msg=""):
        output_lines.append(msg)
        print(msg)

    log("\n" + "="*100)
    log("COMPLETE RELIC DEEP DIVE ANALYSIS")
    log("="*100)

    # =========================================================================
    # PART 1: SAVE FILE ANALYSIS
    # =========================================================================
    log(f"\n{'='*50}")
    log("PART 1: SAVE FILE STRUCTURE")
    log(f"{'='*50}")
    log(f"File size: {len(file_data)} bytes (0x{len(file_data):X})")

    # Find all relic GA handles in the file
    relic_handles = []
    for offset in range(0, len(file_data) - 4, 4):
        val = struct.unpack_from('<I', file_data, offset)[0]
        if (val & 0xF0000000) == ITEM_TYPE_RELIC and val != 0:
            relic_handles.append((offset, val))

    log(f"\nTotal relic GA handles found: {len(relic_handles)}")

    # Analyze the relic item data structure (from gaprint parsing)
    log(f"\n--- Relic Item Data (from ga_relic) ---")
    log(f"Total relics parsed: {len(ga_relic)}")
    if ga_relic:
        log("\nRelic structure: (ga_handle, item_id, effect1, effect2, effect3, curse1, curse2, curse3, offset, size)")
        log("\nFirst 10 relics with full data:")
        for i, relic in enumerate(ga_relic[:10]):
            ga_handle, item_id, eff1, eff2, eff3, curse1, curse2, curse3, offset, size = relic
            real_id = item_id - 2147483648
            item_info = items_json.get(str(real_id), {})
            name = item_info.get('name', 'Unknown')
            color = item_info.get('color', 'Unknown')
            log(f"\n  Relic {i+1}:")
            log(f"    GA Handle: 0x{ga_handle:08X}")
            log(f"    Item ID: {item_id} (real: {real_id})")
            log(f"    Name: {name}")
            log(f"    Color: {color}")
            log(f"    Effects: {eff1}, {eff2}, {eff3}")
            log(f"    Curses: {curse1}, {curse2}, {curse3}")
            log(f"    Offset: 0x{offset:06X}, Size: {size} bytes")

            # Show raw bytes around this relic
            if offset > 0 and offset + size < len(file_data):
                log(f"    Raw data at offset (first 80 bytes):")
                for j in range(0, min(80, size), 16):
                    hex_str = ' '.join(f'{file_data[offset+j+k]:02X}' for k in range(min(16, size-j)))
                    log(f"      0x{offset+j:06X}: {hex_str}")

    # =========================================================================
    # PART 2: ANALYZE RELIC STRUCTURE IN DETAIL
    # =========================================================================
    log(f"\n{'='*50}")
    log("PART 2: DETAILED RELIC BYTE STRUCTURE")
    log(f"{'='*50}")

    if ga_relic:
        # Pick one relic and show complete structure
        sample_relic = ga_relic[0]
        ga_handle, item_id, eff1, eff2, eff3, curse1, curse2, curse3, offset, size = sample_relic

        log(f"\nSample relic at offset 0x{offset:06X} (size {size} bytes):")
        log(f"Interpreting structure field by field:")

        cursor = offset
        fields = [
            ("GA Handle", 4),
            ("Item ID", 4),
            ("Durability?", 4),
            ("Unknown 1", 4),
            ("Effect 1", 4),
            ("Effect 2", 4),
            ("Effect 3", 4),
        ]

        # Read and display each potential field
        for field_name, field_size in fields:
            if cursor + field_size <= len(file_data):
                val = struct.unpack_from('<I', file_data, cursor)[0]
                log(f"  0x{cursor:06X} [{field_name:12}]: {val} (0x{val:08X})")
                cursor += field_size

        # Show remaining bytes
        remaining = size - (cursor - offset)
        if remaining > 0:
            log(f"\n  Remaining {remaining} bytes:")
            for i in range(0, remaining, 16):
                hex_str = ' '.join(f'{file_data[cursor+i+k]:02X}' for k in range(min(16, remaining-i)))
                vals = struct.unpack_from(f'<{min(4, (remaining-i)//4)}I', file_data, cursor+i) if remaining-i >= 4 else []
                log(f"    0x{cursor+i:06X}: {hex_str}")
                if vals:
                    log(f"             As ints: {', '.join(str(v) for v in vals)}")

    # =========================================================================
    # PART 3: DATA FILES ANALYSIS
    # =========================================================================
    log(f"\n{'='*50}")
    log("PART 3: DATA FILES ANALYSIS")
    log(f"{'='*50}")

    # Analyze items.json
    log(f"\n--- items.json Analysis ---")
    log(f"Total items: {len(items_json)}")

    colors_found = {}
    for item_id, item_data in items_json.items():
        color = item_data.get('color')
        if color not in colors_found:
            colors_found[color] = []
        colors_found[color].append((item_id, item_data.get('name', 'Unknown')))

    log(f"\nColors found in items.json:")
    for color, items in sorted(colors_found.items(), key=lambda x: str(x[0])):
        log(f"  {color}: {len(items)} items")
        # Show first 3 examples
        for item_id, name in items[:3]:
            log(f"    - ID {item_id}: {name}")

    # Analyze effects.json
    log(f"\n--- effects.json Analysis ---")
    log(f"Total effects: {len(effects_json)}")

    # Sample some effects
    log("\nSample effects:")
    for i, (eff_id, eff_data) in enumerate(list(effects_json.items())[:10]):
        log(f"  {eff_id}: {eff_data}")

    # =========================================================================
    # PART 4: VESSEL STRUCTURE ANALYSIS
    # =========================================================================
    log(f"\n{'='*50}")
    log("PART 4: VESSEL STRUCTURE DEEP DIVE")
    log(f"{'='*50}")

    # Find vessel structures and analyze surrounding data
    vessel_findings = []
    for offset in range(0x10000, min(0x50000, len(file_data) - 28), 4):
        val = struct.unpack_from('<I', file_data, offset)[0]
        if 1000 <= val <= 10010 or 19000 <= val <= 19020:
            char_id = (val - 1000) // 1000
            vessel_slot = val % 1000
            if vessel_slot <= 10:
                ga_handles = list(struct.unpack_from('<6I', file_data, offset + 4))
                has_relics = any((h & 0xF0000000) == ITEM_TYPE_RELIC for h in ga_handles if h != 0)
                vessel_findings.append({
                    'offset': offset,
                    'vessel_id': val,
                    'char_id': char_id,
                    'vessel_slot': vessel_slot,
                    'ga_handles': ga_handles,
                    'has_relics': has_relics
                })

    log(f"\nFound {len(vessel_findings)} vessel entries")

    # Show vessels with relics and analyze surrounding data
    log("\n--- Vessels WITH relics (detailed) ---")
    for v in vessel_findings:
        if v['has_relics']:
            log(f"\nVessel ID {v['vessel_id']} (Char {v['char_id']}, Slot {v['vessel_slot']}) at 0x{v['offset']:06X}:")
            log(f"  GA Handles: {[f'0x{h:08X}' for h in v['ga_handles']]}")

            # Show 32 bytes BEFORE the vessel ID
            if v['offset'] >= 32:
                log(f"  32 bytes BEFORE vessel ID:")
                for i in range(8):
                    check_offset = v['offset'] - 32 + (i * 4)
                    check_val = struct.unpack_from('<I', file_data, check_offset)[0]
                    log(f"    0x{check_offset:06X}: {check_val:10} (0x{check_val:08X})")

            # Show 32 bytes AFTER the GA handles
            after_offset = v['offset'] + 4 + 24  # vessel_id + 6 handles
            if after_offset + 32 < len(file_data):
                log(f"  32 bytes AFTER GA handles:")
                for i in range(8):
                    check_offset = after_offset + (i * 4)
                    check_val = struct.unpack_from('<I', file_data, check_offset)[0]
                    log(f"    0x{check_offset:06X}: {check_val:10} (0x{check_val:08X})")

            # Resolve relic names for this vessel
            log(f"  Relic details:")
            for idx, ga in enumerate(v['ga_handles']):
                if ga != 0 and (ga & 0xF0000000) == ITEM_TYPE_RELIC:
                    # Find this relic in ga_relic
                    for relic in ga_relic:
                        if relic[0] == ga:
                            real_id = relic[1] - 2147483648
                            item_info = items_json.get(str(real_id), {})
                            name = item_info.get('name', 'Unknown')
                            color = item_info.get('color', 'Unknown')
                            log(f"    Slot {idx}: {name} ({color}) - Effects: {relic[2]}, {relic[3]}, {relic[4]}")
                            break
                else:
                    log(f"    Slot {idx}: (empty)")

    # =========================================================================
    # PART 5: LOOK FOR ADDITIONAL PATTERNS
    # =========================================================================
    log(f"\n{'='*50}")
    log("PART 5: SEARCHING FOR ADDITIONAL PATTERNS")
    log(f"{'='*50}")

    # Look for color-related values near relics
    log("\n--- Searching for color values (0-3) near relic data ---")
    if ga_relic:
        for relic in ga_relic[:5]:
            offset = relic[8]
            real_id = relic[1] - 2147483648
            item_info = items_json.get(str(real_id), {})
            color = item_info.get('color', 'Unknown')

            log(f"\nRelic '{item_info.get('name', 'Unknown')}' (color={color}) at 0x{offset:06X}:")
            # Search in 20 bytes before and after for small values 0-3
            search_start = max(0, offset - 20)
            search_end = min(len(file_data), offset + 100)

            small_vals = []
            for i in range(search_start, search_end):
                if file_data[i] <= 3:
                    small_vals.append((i, file_data[i]))

            if small_vals:
                log(f"  Small values (0-3) found in range 0x{search_start:06X}-0x{search_end:06X}:")
                for pos, val in small_vals[:20]:
                    log(f"    0x{pos:06X}: {val}")

    # =========================================================================
    # PART 6: CSV FILE ANALYSIS
    # =========================================================================
    log(f"\n{'='*50}")
    log("PART 6: CSV PARAM FILE ANALYSIS")
    log(f"{'='*50}")

    # Read and analyze EquipParamAntique.csv
    try:
        import csv
        import os
        csv_path = os.path.join(os.path.dirname(__file__), "Resources/Param/EquipParamAntique.csv")
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        log(f"\nEquipParamAntique.csv: {len(rows)} rows")
        log(f"Columns: {list(rows[0].keys()) if rows else 'N/A'}")

        # Analyze relicColor distribution
        color_dist = {}
        deep_relic_count = 0
        for row in rows:
            color = row.get('relicColor', 'N/A')
            is_deep = row.get('isDeepRelic', '0')
            if color not in color_dist:
                color_dist[color] = 0
            color_dist[color] += 1
            if is_deep == '1':
                deep_relic_count += 1

        log(f"\nrelicColor distribution:")
        for color, count in sorted(color_dist.items()):
            log(f"  {color}: {count} items")

        log(f"\nDeep relics (isDeepRelic=1): {deep_relic_count}")

        # Show sample rows with different colors
        log(f"\nSample rows by color:")
        shown_colors = set()
        for row in rows:
            color = row.get('relicColor', 'N/A')
            if color not in shown_colors and len(shown_colors) < 5:
                shown_colors.add(color)
                log(f"\n  Color {color} example (ID={row.get('ID')}):")
                for key, val in list(row.items())[:15]:
                    log(f"    {key}: {val}")
    except Exception as e:
        log(f"\nError reading CSV: {e}")

    # =========================================================================
    # SUMMARY
    # =========================================================================
    log(f"\n{'='*100}")
    log("SUMMARY")
    log(f"{'='*100}")
    log(f"- Save file has {len(ga_relic)} relics parsed")
    log(f"- items.json has {len(items_json)} items with colors: {list(colors_found.keys())}")
    log(f"- {len(vessel_findings)} vessel entries found")
    log(f"- Vessels with relics: {sum(1 for v in vessel_findings if v['has_relics'])}")

    # Write to file
    with open("debug_relic_analysis.txt", "w", encoding="utf-8") as f:
        f.write("\n".join(output_lines))
    log(f"\n[Output also saved to debug_relic_analysis.txt]")

    log("\n" + "="*100)


def debug_dump_save_analysis(file_data):
    """Debug function to analyze save file structure and find vessel-related data."""
    # Call the complete analysis
    debug_dump_complete_relic_analysis(file_data)


def get_character_loadout(char_name):
    """Get the current relic loadout for a character.

    Returns:
        Dict with vessel_slot -> {'relics': list of (ga_handle, relic_info), 'unlocked': bool}
    """
    
    char_id = data_source.character_names.index(char_name) if char_name in data_source.character_names else -1
    if char_id < 0:
        return {}
    hero_type = char_id + 1
    loadout = {}
    for idx, vessel in enumerate(loadout_handler.heroes[hero_type].vessels):
        relics = []
        has_any_relic = False
        for ga in vessel['relics']:
            if ga != 0 and (ga & 0xF0000000) == ITEM_TYPE_RELIC:
                has_any_relic = True
                # Find relic info
                relic_info = None
                for relic in ga_relic:
                    if len(relic) >= 8 and relic[0] == ga:  # ga_handle match
                        real_id = relic[1] - 2147483648
                        item_data = items_json.get(str(real_id), {})
                        relic_info = {
                            'ga': ga,
                            'real_id': real_id,
                            'name': item_data.get('name', f'Unknown ({real_id})'),
                            'color': item_data.get('color', 'Unknown'),
                            'effects': [relic[2], relic[3], relic[4]],
                            'curses': [relic[5], relic[6], relic[7]]
                        }
                        break
                relics.append((ga, relic_info))
            else:
                relics.append((0, None))

        # A vessel is considered "unlocked" if:
        # 1. It's vessel slot 0 (always unlocked - the default vessel)
        # 2. It has any relic assigned (player must have unlocked it to assign relics)
        # Note: We can't reliably detect unlock status from save data alone,
        # so we show "Unknown" for empty vessels beyond slot 0
        is_unlocked = idx == 0 or has_any_relic

        loadout[idx] = {
            'relics': relics,
            'unlocked': is_unlocked,
            'has_relics': has_any_relic
        }

    return loadout


# Color mapping for UI display
RELIC_COLOR_HEX = {
    'Red': '#FF4444',
    'Blue': '#4488FF',
    'Yellow': '#B8860B',  # Dark goldenrod - readable on light backgrounds
    'Green': '#44BB44',
    'Purple': '#AA44FF',
    'Orange': '#FF8C00',
    'White': '#AAAAAA',  # Gray for universal slot - readable on light backgrounds
    'Unknown': '#888888',
    None: '#888888'
}


def read_murks_and_sigs(data):
    global current_murks, current_sigs
    offset = gaprint(data)
    name_offset = offset + 0x94
    murks_offset = name_offset + 52
    sigs_offset = name_offset - 64
    
    current_murks = struct.unpack_from('<I', data, murks_offset)[0]
    current_sigs = struct.unpack_from('<I', data, sigs_offset)[0]
    
    return current_murks, current_sigs


def write_murks_and_sigs(murks_value, sigs_value):
    global loadout_handler
    offset = gaprint(globals.data)
    name_offset = offset + 0x94
    murks_offset = name_offset + 52
    sigs_offset = name_offset - 64
    
    # Write murks
    murks_bytes = murks_value.to_bytes(4, 'little')
    globals.data = globals.data[:murks_offset] + murks_bytes + globals.data[murks_offset+4:]
    
    # Write sigs
    sigs_bytes = sigs_value.to_bytes(4, 'little')
    globals.data = globals.data[:sigs_offset] + sigs_bytes + globals.data[sigs_offset+4:]
    
    save_current_data()


def split_files(file_path, folder_name):
    file_name = os.path.basename(file_path)
    split_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), folder_name)
    #clean current dir
    if os.path.exists(split_dir):
        shutil.rmtree(split_dir)  # delete folder and everything inside
    os.makedirs(split_dir, exist_ok=True)

    if file_name.lower() == 'memory.dat':
        with open(file_path, "rb") as f:
            header = f.read(0x80)
            with open(os.path.join(split_dir, "header"), "wb") as out:
                out.write(header)
            
            chunk_size = 0x100000
            for i in range(10):
                data = f.read(chunk_size)
                if not data:
                    break
                with open(os.path.join(split_dir, f"userdata{i}"), "wb") as out:
                    data=bytearray(data)
                    data=(0x00100010).to_bytes(4, "little")+ data
                    out.write(data)
            
            regulation = f.read()
            if regulation:
                with open(os.path.join(split_dir, "regulation"), "wb") as out:
                    out.write(regulation)

    elif file_path.lower().endswith('.sl2'):
        # Accept any .sl2 file (supports custom save names from ModEngine 3, etc.)
        decrypt_ds2_sl2(file_path)

def save_file():
    save_current_data()

    if MODE=='PC':

        output_sl2_file=filedialog.asksaveasfilename( initialfile="NR0000.sl2", title="Save PC SL2 save as")
        if not output_sl2_file:
            return
        
        encrypt_modified_files(output_sl2_file)

    if MODE == 'PS4':  ### HERE
        print('data length', len(globals.data))
        
        # Validate data length before proceeding
        expected_length = 0x100004
        if len(globals.data) != expected_length:
            messagebox.showerror('Error', 
                            f'Modified userdata size is invalid. '
                            f'Expected {hex(expected_length)}, got {hex(len(globals.data))}. Cannot save.')
            return
        
        try:
            split_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'decrypted_output')
            
            # Validate split directory exists
            if not os.path.exists(split_dir):
                messagebox.showerror("Error", f"Directory not found: {split_dir}")
                return
            
            output_file = filedialog.asksaveasfilename(
                initialfile="memory.dat",
                title="Save PS4 save as",
                defaultextension=".dat",
                filetypes=[("DAT files", "*.dat"), ("All files", "*.*")]
            )
            
            if not output_file:
                return
            
            # Track total bytes written for validation
            total_bytes_written = 0
            
            with open(output_file, "wb") as out:
                # 1. HEADER
                header_path = os.path.join(split_dir, "header")
                if not os.path.exists(header_path):
                    messagebox.showerror("Error", f"Header file not found: {header_path}")
                    return
                
                with open(header_path, "rb") as f:
                    header_data = f.read()
                    if len(header_data) != 0x80:
                        messagebox.showerror("Error", 
                                        f"Invalid header size: {hex(len(header_data))}. "
                                        f"Expected {hex(0x80)} bytes.")
                        return
                    out.write(header_data)
                    total_bytes_written += len(header_data)
                
                print(f"Written header: {hex(total_bytes_written)} bytes")
                
                # 2. USERDATA 0–9
                check_padding = (0x00100010).to_bytes(4, "little")
                userdata_chunks_found = 0
                
                for i in range(10):
                    userdata_path = os.path.join(split_dir, f"userdata{i}")
                    
                    if not os.path.exists(userdata_path):
                        # Check if this is expected (some saves may have fewer chunks)
                        if i == 0:
                            messagebox.showerror("Error", 
                                            f"Required file not found: {userdata_path}")
                            return
                        else:
                            print(f"Warning: userdata{i} not found, stopping at {i} chunks")
                            break
                    
                    # Read original
                    with open(userdata_path, "rb") as f:
                        block = f.read()
                    
                    # Validate block has data
                    if len(block) < 4:
                        messagebox.showerror("Error", 
                                        f"userdata{i} is too small ({len(block)} bytes)")
                        return
                    
                    # PS4 USERDATA should start with 4 bytes padding
                    if block[:4] == check_padding:
                        # Strip the padding
                        block = block[4:]
                    else:
                        # Padding missing - this is suspicious but warn and continue
                        print(f"Warning: userdata{i} does not start with expected padding {check_padding.hex()}")
                        print(f"         Got: {block[:4].hex()}")
                        # Don't add padding, just use as-is
                    
                    # Validate chunk size (should be 0x100000 for full chunks)
                    expected_chunk_size = 0x100000
                    if len(block) != expected_chunk_size and i < 9:  # Last chunk might be smaller
                        print(f"Warning: userdata{i} has unexpected size {hex(len(block))}, "
                            f"expected {hex(expected_chunk_size)}")
                    
                    # Write block to output
                    out.write(block)
                    total_bytes_written += len(block)
                    userdata_chunks_found += 1
                
                print(f"Written {userdata_chunks_found} userdata chunks: {hex(total_bytes_written)} bytes total")
                
                # 3. REGULATION
                regulation_path = os.path.join(split_dir, "regulation")
                if os.path.exists(regulation_path):
                    with open(regulation_path, "rb") as f:
                        regulation_data = f.read()
                        if regulation_data:
                            out.write(regulation_data)
                            total_bytes_written += len(regulation_data)
                            print(f"Written regulation: {len(regulation_data)} bytes")
                        else:
                            print("Warning: regulation file is empty")
                else:
                    print("Warning: regulation file not found, skipping")
            
            # 4. SIZE VALIDATION
            final_size = os.path.getsize(output_file)
            expected_final_size = 0x12A00A0
            
            print(f"Final file size: {hex(final_size)} (expected: {hex(expected_final_size)})")
            
            if final_size != expected_final_size:
                messagebox.showerror('ERROR',
                                f'Invalid output file size!\n'
                                f'Expected: {hex(expected_final_size)} ({expected_final_size:,} bytes)\n'
                                f'Got: {hex(final_size)} ({final_size:,} bytes)\n'
                                f'Difference: {final_size - expected_final_size:+,} bytes\n\n'
                                f'File may be corrupt. Check the source files in {split_dir}')
                return
            
            messagebox.showinfo('Success', f'File saved successfully to:\n{output_file}')
            print(f"Successfully saved to: {output_file}")
            
        except PermissionError as e:
            messagebox.showerror("Permission Error", 
                            f"Cannot write to file. Check permissions.\n{str(e)}")
        except IOError as e:
            messagebox.showerror("I/O Error", 
                            f"Error reading/writing files.\n{str(e)}")
        except Exception as e:
            messagebox.showerror("Exception", 
                            f"Unexpected error occurred:\n{str(e)}\n\n"
                            f"Check console for details.")
            import traceback
            traceback.print_exc()

            

def name_to_path():
    global char_name_list, MODE
    char_name_list = []
    unpacked_folder = WORKING_DIR / 'decrypted_output'

    prefix = "userdata" if MODE == 'PS4' else "USERDATA_0"

    for i in range(10):
        file_path = os.path.join(unpacked_folder, f"{prefix}{i}")
        if not os.path.exists(file_path):
            continue

        try:
            with open(file_path, "rb") as f:
                file_data = f.read()
                # Check minimum file size before parsing
                if len(file_data) < 0x1000:  # Minimum expected size
                    print(f"Warning: {file_path} is too small ({len(file_data)} bytes), skipping")
                    continue
                name = read_char_name(file_data)
                if name:
                    char_name_list.append((name, file_path))
        except struct.error as e:
            print(f"Error parsing save file {file_path}: Data structure error - {e}")
            print(f"  This may indicate a corrupted save file or incompatible format")
        except IndexError as e:
            import traceback
            traceback.print_exc()
            print(f"Error parsing save file {file_path}: Index error - {e}")
            print(f"  This may indicate a corrupted save file")
        except Exception as e:
            import traceback
            traceback.print_exc()
            print(f"Error reading {file_path}: {e}")

def name_to_path_import():
    global char_name_list_import, IMPORT_MODE
    char_name_list_import = []
    unpacked_folder = WORKING_DIR / 'decrypted_output_import'
    
    prefix = "userdata" if IMPORT_MODE == 'PS4' else "USERDATA_0"
    
    for i in range(10):
        file_path = os.path.join(unpacked_folder, f"{prefix}{i}")
        if not os.path.exists(file_path):
            continue
            
        try:
            with open(file_path, "rb") as f:
                file_data = f.read()
                name = read_char_name(file_data)
                if name:
                    char_name_list_import.append((name, file_path))
        except Exception as e:
            print(f"Error reading {file_path}: {e}")


def delete_relic(ga_index, item_id):
    
    last_offset = gaprint(globals.data)
    inventory_start = last_offset + 0x650
    inventory_end = inventory_start + 0xA7AB
    inventory_data = globals.data[inventory_start:inventory_start + inventory_end]
    
    ga_bytes = ga_index.to_bytes(4, byteorder='little')
    replacement = bytes.fromhex('00000000FFFFFFFF')
    
    for ga, id, e1, e2, e3, e4, e5, e6, offset, size in ga_relic:
        real_id = id - 2147483648
        
        if ga_index == ga and real_id == item_id:
            inventory_offset = inventory_data.find(ga_bytes)
            match = inventory_offset + inventory_start
            
            globals.data = globals.data[:match] + b"\x00" * 14 + globals.data[match+14:]

            globals.data = globals.data[:offset] + globals.data[offset+80:]
            globals.data = globals.data[:offset] + replacement + globals.data[offset:]

            globals.data = globals.data[:-0x1C] + b'\x00' * 72 + globals.data[-0x1C:]
            
            save_current_data()
            if relic_checker:
                relic_checker.remove_illegal(ga_index)
            return True
    return False


def add_relic() -> tuple[bool, Any]:
    """
    Adds a new relic item to the player's save file.

    This function performs the following steps:
    1. Parses the save data to find all existing items and relics
    2. Generates a unique GA (Game Asset) handle for the new relic
    3. Finds the last relic in the inventory section to determine insertion point
    4. Locates an empty item slot that can be converted into a relic slot
    5. Updates both the inventory reference and the item slot data
    6. Adjusts file padding to maintain constant file size

    Returns:
        bool: True if the relic was successfully added, False otherwise
        ga: New GA handle of the added relic, or None on failure
    """

    last_offset = gaprint(globals.data)

    inventory_start = last_offset + 0x650
    inventory_end = inventory_start + 0xA7AB
    inventory_data = globals.data[inventory_start : inventory_start + inventory_end]

    # Find the first existing GA handle among all relics.
    gas = sorted([ga for ga, id, e1, e2, e3, e4, e5, e6, offset, size in ga_relic])
    first_available_ga = -1
    for i in range(0, len(gas) - 1):
        if gas[i] + 1 != gas[i+1]:
            first_available_ga = gas[i] + 1
            break

    # Find the new unique ga for the relic
    if len(gas) == 0:
        # If no relics exist, use a default starting value (ITEM_TYPE_RELIC | 0x85).
        new_ga = (ITEM_TYPE_RELIC | 0x00000085)
    if first_available_ga > 0:
        # Use first available ga
        new_ga = first_available_ga
    else:
        # Generate a new unique GA handle by incrementing the max.
        new_ga = max(gas) + 1
    if new_ga > (ITEM_TYPE_RELIC | 0xfffffff):
        # Inventory full
        return False, None
    new_ga_bytes = new_ga.to_bytes(4, byteorder="little")

    # Build the inventory entry (14 bytes) for the new relic.
    id_to_write = bytearray(new_ga_bytes + b"\x01\x00\x00\x00'|\x00\x00\x00\x00")

    # Build the full relic item data (80 bytes) for the item slot section.
    replacement = bytearray(
        new_ga_bytes
        + b"u\x00\x00\x80u\x00\x00\x80\xff\xff\xff\xff\xe4xk\x00 \xb4l\x00\x9e\xd5j\x00\xff\xff\xff\xff\xff\xff\xff\xff\x00\x00\x00\xff\x00\x00\x00\x00\x00\x00\x00\x00\xff\xff\xff\xff\xff\xff\xff\xff\xff\xff\xff\xff\xff\xff\xff\xff\xff\xff\xff\xff\xff\xff\xff\xff\x00\x00\x00\x00\x00\x00\x00\x00"
    )

    # Find the position of the last relic in the inventory section.
    last_inventory = -1
    for ga, id, e1, e2, e3, e4, e5, e6, offset, size in ga_relic:
        gab = ga.to_bytes(4, byteorder="little")
        last_inventory = max(last_inventory, inventory_data.find(gab))
    if last_inventory == -1:
        # No existing relics found in inventory - cannot determine insertion point
        return False, None

    # Find an empty slot in the item data section to convert into a relic slot.
    empty_slot_offset = -1
    for ga, id, e1, e2, e3, e4, e5, e6, offset, size in ga_items:
        if ga == 0 and size == 8:
            empty_slot_offset = offset
            break

    if empty_slot_offset == -1:
        # No empty slot available to convert - inventory is full
        return False, None

    # Write the inventory entry at the calculated position.
    match = last_inventory + inventory_start
    globals.data = globals.data[:match] + id_to_write + globals.data[match + 14 :]

    # Convert the empty 8-byte slot into an 80-byte relic slot.
    globals.data = (
        globals.data[:empty_slot_offset] + globals.data[empty_slot_offset + 8 :]
    )
    # Then, insert the 80-byte relic data at the same position.
    globals.data = (
        globals.data[:empty_slot_offset]
        + replacement
        + globals.data[empty_slot_offset:]
    )

    # Remove 72 bytes from the padding area at the end of the file.
    # The save file must maintain a constant size for the game to load it.
    globals.data = globals.data[: -0x1C - 72] + globals.data[-0x1C:]

    save_current_data()
    return True, new_ga


def modify_relic(ga_index, item_id, new_effects, new_item_id=None):

    for ga, id, e1, e2, e3, e4, e5, e6, offset, size in ga_relic:
        real_id = id - 2147483648

        if ga_index == ga and real_id == item_id:
            # Update item ID if provided
            if new_item_id is not None and new_item_id != real_id:
                # Convert to internal format
                new_id_internal = new_item_id + 2147483648
                item_id_offset = offset + 4  # Skip GA handle (4 bytes)
                item_id_bytes = new_id_internal.to_bytes(4, byteorder='little')
                globals.data = globals.data[:item_id_offset] + item_id_bytes + globals.data[item_id_offset+4:]

            # Modify effects in the relic data structure
            effect_offset = offset + 16  # Skip handle, id, durability, unk_1

            # Write primary effects
            for i, eff in enumerate(new_effects[:3]):
                eff_bytes = eff.to_bytes(4, byteorder='little')
                pos = effect_offset + (i * 4)
                globals.data = globals.data[:pos] + eff_bytes + globals.data[pos+4:]

            # Write secondary effects
            sec_effect_offset = effect_offset + 12 + 0x1C  # Skip padding
            for i, eff in enumerate(new_effects[3:6]):
                eff_bytes = eff.to_bytes(4, byteorder='little')
                pos = sec_effect_offset + (i * 4)
                globals.data = globals.data[:pos] + eff_bytes + globals.data[pos+4:]

            save_current_data()
            if relic_checker:
                relic_checker.update_illegal(ga_index,
                                             item_id if new_item_id is None else new_item_id,
                                             new_effects)
            return True
    return False


def modify_relic_by_ga(ga_index, new_effects, new_item_id, sort_effects=True):
    """Modify a relic by GA handle only (doesn't require matching item_id)"""

    for ga, id, e1, e2, e3, e4, e5, e6, offset, size in ga_relic:
        if ga_index == ga:
            # Sort effects if requested (fixes effect ordering issues)
            if sort_effects and relic_checker:
                new_effects = relic_checker.sort_effects(new_effects)

            # Update item ID
            new_id_internal = new_item_id + 2147483648
            item_id_offset = offset + 4  # Skip GA handle (4 bytes)
            item_id_bytes = new_id_internal.to_bytes(4, byteorder='little')
            globals.data = globals.data[:item_id_offset] + item_id_bytes + globals.data[item_id_offset+4:]

            # Modify effects in the relic data structure
            effect_offset = offset + 16  # Skip handle, id, durability, unk_1

            # Write primary effects
            for i, eff in enumerate(new_effects[:3]):
                eff_bytes = eff.to_bytes(4, byteorder='little')
                pos = effect_offset + (i * 4)
                globals.data = globals.data[:pos] + eff_bytes + globals.data[pos+4:]

            # Write secondary effects
            sec_effect_offset = effect_offset + 12 + 0x1C  # Skip padding
            for i, eff in enumerate(new_effects[3:6]):
                eff_bytes = eff.to_bytes(4, byteorder='little')
                pos = sec_effect_offset + (i * 4)
                globals.data = globals.data[:pos] + eff_bytes + globals.data[pos+4:]

            save_current_data()
            if relic_checker:
                relic_checker.update_illegal(ga_index,
                                             new_item_id,
                                             new_effects)
            return True
    return False


def check_illegal_relics():
    global relic_checker
    illegal_relics = relic_checker.illegal_gas
    # for ga, relic_id, e1, e2, e3, e4, e5, e6, offset, size in ga_relic:
    #     # Skip relic entirely if its ID is invalid
    #     if relic_id in (0, -1, 4294967295):
    #         continue

    #     effects = [e1, e2, e3, e4, e5, e6]
    #     used_ids = set()
    #     used_base_names = {}
    #     is_illegal = False

    #     for idx, eff in enumerate(effects, start=1):
    #         # Treat unknown effects as empty
    #         if eff in (0, -1, 4294967295):
    #             continue

    #         eff_key = str(eff)

    #         # Rule 1 — in illegal JSON
    #         if eff_key in ill_effects_json:
    #             is_illegal = True
    #             break

    #         # Lookup in main effects DB
    #         eff_name = effects_json.get(eff_key, {}).get("name", f"Unknown({eff})")

    #         # Rule 2 — duplicate ID
    #         if eff in used_ids:
    #             is_illegal = True
    #             break
    #         used_ids.add(eff)

    #         # Rule 3 — conflicting tiers
    #         base_name = eff_name.rsplit(" +", 1)[0] if " +" in eff_name else eff_name
    #         if base_name in used_base_names:
    #             is_illegal = True
    #             break

    #         used_base_names[base_name] = eff_name

    #     if is_illegal:
    #         illegal_relics.append(ga)

    return illegal_relics


def get_forbidden_relics():
    forbidden_relic_ids = RelicChecker.UNIQUENESS_IDS
    # forbidden_relic_ids = {
    #     1000, 1010, 1020, 1030, 1040, 1050, 1060, 1070, 1080, 1090,
    #     1100, 1110, 1120, 1130, 1140, 1150, 1160, 1170, 1180, 1190,
    #     1200, 1210, 1220, 1230, 1240, 1250, 1260, 1270, 11004, 10001,
    #     1400, 1410, 1420, 1430, 1440, 1450, 1460, 1470, 1480, 1490,
    #     1500, 1510, 1520
    # }
    return forbidden_relic_ids


def split_files_import(file_path, folder_name):
    global IMPORT_MODE
    file_name = os.path.basename(file_path)
    split_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), folder_name)
    # clean current dir
    if os.path.exists(split_dir):
        shutil.rmtree(split_dir)  # delete folder and everything inside
    os.makedirs(split_dir, exist_ok=True)

    if file_name.lower() == 'memory.dat':
        IMPORT_MODE = 'PS4'
        with open(file_path, "rb") as f:
            header = f.read(0x80)
            with open(os.path.join(split_dir, "header"), "wb") as out:
                out.write(header)
            
            chunk_size = 0x100000
            for i in range(10):
                data = f.read(chunk_size)
                if not data:
                    break
                with open(os.path.join(split_dir, f"userdata{i}"), "wb") as out:
                    data=bytearray(data)
                    data=(0x00100010).to_bytes(4, "little")+ data
                    out.write(data)
            
            regulation = f.read()
            if regulation:
                with open(os.path.join(split_dir, "regulation"), "wb") as out:
                    out.write(regulation)

    elif file_path.lower().endswith('.sl2'):
        # Accept any .sl2 file (supports custom save names from ModEngine 3, etc.)
        IMPORT_MODE='PC'
        decrypt_ds2_sl2_import(file_path)

def import_save():
    global imported_data
    global char_name_list_import

    if globals.data==None:
        messagebox.showerror('Error', 'Please select a character to replace first')
        return

    import_path = filedialog.askopenfilename()
    if not import_path:
        return

    # Split and generate list
    split_files_import(import_path, "decrypted_output_import")
    name_to_path_import()  # generates char_name_list_import = [(name, path), ...]

    # Show popup window with buttons
    show_import_popup()
    

def show_import_popup():
    popup = tk.Toplevel()
    popup.title("Select Character to Import")

    label = tk.Label(popup, text="Choose a character:", font=("Arial", 12, "bold"))
    label.pack(pady=10)

    # Create buttons for each character
    for name, path in char_name_list_import:
        btn = tk.Button(
            popup, 
            text=name, 
            width=30, 
            command=lambda p=path: load_imported_data_and_close(p, popup)
        )
        btn.pack(pady=3)

def load_imported_data_and_close(path, popup):
    load_imported_data(path)
    popup.destroy()

def load_imported_data(path):
    global imported_data

    # Check if steam_id was found - required for import
    if steam_id is None:
        messagebox.showerror("Error",
            "Cannot import save: Steam ID not found in current save file.\n\n"
            "The Steam ID pattern could not be detected in your save.\n"
            "This may indicate a corrupted or incompatible save file.")
        return

    with open (path, "rb") as f:
        imported_data=f.read()

    offsets = aob_search(imported_data, AOB_search)
    if not offsets:
        messagebox.showerror("Error",
            "Cannot import save: Steam ID pattern not found in the imported save file.\n\n"
            "The imported save may be corrupted or incompatible.")
        return

    offset = offsets[0] + 44
    imported_data = imported_data[:offset] + bytes.fromhex(steam_id) + imported_data[offset + 8:]


    if len(imported_data) <= len(globals.data):
        globals.data = imported_data + globals.data[len(imported_data):]

    else:
        globals.data = imported_data[:len(globals.data)]

    for name, file in char_name_list_import:
        if path == file:
            char_name=name
    save_current_data()
    messagebox.showinfo("Success", f"Character '{char_name}' imported successfully. Save the file and open it again to see changes.")
    

    





    



def export_relics_to_excel(filepath="relics.xlsx"):
    # Lazy import openpyxl only when needed
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    if not ga_relic:
        return False, "No relics found in ga_relic."

    wb = Workbook()
    ws = wb.active
    ws.title = "Relics"

    # Column headers
    headers = [
        "Item ID",
        "Relic Name",
        "Relic Color",
        "Effect 1 (ID)", "Effect 1 (Name)",
        "Effect 2 (ID)", "Effect 2 (Name)",
        "Effect 3 (ID)", "Effect 3 (Name)",
        "Sec Effect 1 (ID)", "Sec Effect 1 (Name)",
        "Sec Effect 2 (ID)", "Sec Effect 2 (Name)",
        "Sec Effect 3 (ID)", "Sec Effect 3 (Name)",
    ]

    ws.append(headers)

    # Helper to fetch effect name
    def get_eff_name(eid):
        if eid in (0, -1, 4294967295):
            return "None"
        key = str(eid)
        return effects_json.get(key, {}).get("name", f"UnknownEffect({eid})")

    # Fill sheet
    for (_, item_id, e1, e2, e3, se1, se2, se3, offset, size) in ga_relic:
        # Skip invalid relics
        if item_id in (0, -1, 0xFFFFFFFF):
            continue

        real_id = item_id - 2147483648
        
        # Get relic name
        item_key = str(real_id)
        relic_name = items_json.get(item_key, {}).get("name", f"UnknownRelic({real_id})")
        relic_color = items_json.get(item_key, {}).get("color", "Unknown")

        row = [
            real_id,
            relic_name,
            relic_color,
            e1, get_eff_name(e1),
            e2, get_eff_name(e2),
            e3, get_eff_name(e3),
            se1, get_eff_name(se1),
            se2, get_eff_name(se2),
            se3, get_eff_name(se3),
        ]

        ws.append(row)

    # Auto-size columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max_len + 2

    try:
        wb.save(filepath)
        return True, f"Excel file saved: {filepath}"
    except Exception as e:
        return False, f"Failed to save Excel: {str(e)}"


def import_relics_from_excel(filepath):
    """
    Imports relics from an Excel file and modifies current relics to match.
    If imported list is longer than current relic list, extras are ignored.
    """
    # Lazy import openpyxl only when needed
    from openpyxl import load_workbook

    global ga_relic

    if not ga_relic:
        return False, "No relics loaded in current save"

    try:
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Skip header row
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        modifications_made = 0
        
        # Iterate through current relics and imported data simultaneously
        for idx, (ga, item_id, e1, e2, e3, se1, se2, se3, offset, size) in enumerate(ga_relic):
            # If we've exhausted the imported list, stop
            if idx >= len(rows):
                break
            
            row = rows[idx]
            
            # Extract data from Excel row
            # Format: Item ID, Relic Name, Relic Color, E1(ID), E1(Name), E2(ID), E2(Name)...
            new_item_id = row[0] if row[0] is not None else 0
            new_e1 = row[3] if row[3] is not None else 0
            new_e2 = row[5] if row[5] is not None else 0
            new_e3 = row[7] if row[7] is not None else 0
            new_se1 = row[9] if row[9] is not None else 0
            new_se2 = row[11] if row[11] is not None else 0
            new_se3 = row[13] if row[13] is not None else 0
            
            # Convert item ID back to internal format
            new_item_id_internal = new_item_id + 2147483648
            
            # Modify the relic
            real_id = item_id - 2147483648
            
            # Update item ID if different
            if new_item_id != real_id:
                # Write new item ID
                item_id_offset = offset + 4  # Skip GA handle
                item_id_bytes = new_item_id_internal.to_bytes(4, byteorder='little')
                globals.data = globals.data[:item_id_offset] + item_id_bytes + globals.data[item_id_offset+4:]
            
            # Update effects
            new_effects = [new_e1, new_e2, new_e3, new_se1, new_se2, new_se3]
            if modify_relic(ga, real_id, new_effects):
                modifications_made += 1
        
        save_current_data()
        return True, f"Successfully imported and modified {modifications_made} relic(s)"
        
    except Exception as e:
        return False, f"Failed to import Excel: {str(e)}"


def delete_all_illegal_relics():
    """Delete all relics with illegal effects"""
    
    illegal_gas = check_illegal_relics()
    
    if not illegal_gas:
        return 0, "No illegal relics found"
    
    deleted_count = 0
    failed_deletions = []
    
    for ga in illegal_gas:
        # Find the relic with this GA
        for ga_handle, id, e1, e2, e3, e4, e5, e6, offset, size in ga_relic:
            if ga_handle == ga:
                real_id = id - 2147483648
                if delete_relic(ga, real_id):
                    deleted_count += 1
                else:
                    failed_deletions.append(ga)
                break
    
    if failed_deletions:
        return deleted_count, f"Deleted {deleted_count} relics, but {len(failed_deletions)} failed"
    else:
        return deleted_count, f"Successfully deleted {deleted_count} illegal relics"


def save_current_data():
    global userdata_path, loadout_handler
    if globals.data and userdata_path:
        with open(userdata_path, 'wb') as f:

            f.write(globals.data)

def aob_to_pattern(aob: str):

    parts = aob.split()
    pattern = bytearray()
    mask = bytearray()
    for p in parts:
        if p == "??":
            pattern.append(0x00)   # placeholder
            mask.append(0)         # 0 = wildcard (must NOT be 0x00)
        else:
            pattern.append(int(p, 16))
            mask.append(1)         # 1 = must match exactly
    return bytes(pattern), bytes(mask)


def aob_search(data: bytes, aob: str):
    pattern, mask = aob_to_pattern(aob)
    L = len(pattern)
    mv = memoryview(data)

    start = 0x58524  # skip below this offset
    end = len(data) - L + 1

    for i in range(start, end):

        # Check bytes
        for j in range(L):

            b = mv[i + j]

            # Non-wildcard: must match exactly
            if mask[j]:
                if b != pattern[j]:
                    break

            # Wildcard: 
            # 2025-12-28: Allow 0x00 to resolve Steam ID detection issues.
            # Narrowed down AOB_str (bytes 5 & 17 fixed) to prevent false positives.
            else:
                # if b == 0:  # Removed this restriction
                #     break
                continue

        else:
            # Inner loop did not break → MATCH FOUND
            return [i]

    return []


def find_steam_id(section_data):
    # # 假設你的 Steam ID 是 '76561198000000000' (17位數字)
    # # 先將它轉為 8 byte 的 little-endian 二進制格式 (這是 Steam ID 常見的儲存方式)
    # import struct
    # target_steam_id_hex = struct.pack('<Q', int(76561198013358313)).hex().upper()
    # # 或者直接用你已知的 16進位 字串搜尋

    # # 搜尋 section_data 中你 ID 出現的所有位置
    # target_bytes = struct.pack('<Q', int(76561198013358313))
    # index = section_data.find(target_bytes)
    # print(f"你的 Steam ID 出現在偏移量: {hex(index)}")
    # if index != -1:
    #     search_start = index - 44
    #     actual_aob = section_data[search_start : search_start + 17].hex(' ').upper()
    #     print(f"預期 AOB 位置的實際數據為: {actual_aob}")
    #     print(f"原本定義的 AOB 模式為: 00 00 00 00 ?? 00 00 00 ?? ?? 00 00 00 00 00 00 ??")

    offsets = aob_search(section_data, AOB_search)
    if not offsets:
        # AOB pattern not found - return None instead of crashing
        print("Warning: Steam ID AOB pattern not found in save data")
        print(f"  AOB pattern searched: {AOB_search}")
        print(f"  Save data size: {len(section_data)} bytes")
        print(f"  This may happen with PS4 saves or after a game update")
        return None

    offset = offsets[0] + 44
    steam_id = section_data[offset:offset+8]

    hex_str = steam_id.hex().upper()

    return hex_str

class SearchableCombobox(ttk.Frame):
    """A combobox with search functionality and manual entry"""
    def __init__(self, parent, values, **kwargs):
        super().__init__(parent)
        
        self.all_values = values
        self.var = tk.StringVar()
        
        # Entry widget for typing
        self.entry = ttk.Entry(self, textvariable=self.var, **kwargs)
        self.entry.pack(fill='x')
        
        # Listbox for suggestions
        self.listbox = tk.Listbox(self, height=6)
        self.listbox.pack(fill='both', expand=True)
        self.listbox.pack_forget()  # Hidden initially
        
        # Bind events
        self.entry.bind('<KeyRelease>', self.on_keyrelease)
        self.entry.bind('<FocusOut>', self.on_focusout)
        self.listbox.bind('<<ListboxSelect>>', self.on_select)
        self.listbox.bind('<FocusOut>', self.on_focusout)
        
        self.update_listbox(values)
    
    def on_keyrelease(self, event):
        # Filter values based on entry
        value = self.var.get().lower()
        
        if value == '':
            filtered = self.all_values
        else:
            filtered = [item for item in self.all_values if value in item.lower()]
        
        self.update_listbox(filtered)
        
        if filtered and event.keysym not in ('Up', 'Down', 'Return'):
            self.listbox.pack(fill='both', expand=True)
        elif not filtered:
            self.listbox.pack_forget()
    
    def update_listbox(self, values):
        self.listbox.delete(0, tk.END)
        for item in values:
            self.listbox.insert(tk.END, item)
    
    def on_select(self, event):
        if self.listbox.curselection():
            self.var.set(self.listbox.get(self.listbox.curselection()))
            self.listbox.pack_forget()
    
    def on_focusout(self, event):
        # Small delay to allow click on listbox
        self.after(100, lambda: self.listbox.pack_forget())
    
    def get(self):
        return self.var.get()
    
    def set(self, value):
        self.var.set(value)


class SaveEditorGUI:
    def __init__(self, root):
        global loadout_handler
        # ttk Style setting
        style = ttk.Style()
        style.configure('Add.TButton', foreground='green', font=('Arial', 10, 'bold'))
        style.configure('Danger.TButton', foreground='red', font=('Arial', 10))
        
        _ensure_data_source()
        self.root = root
        self.root.title("Elden Ring NightReign Save Editor")
        self.root.geometry("1000x700")

        # Modify dialog reference
        self.modify_dialog = None

        # Clipboard for copy/paste relic effects
        self.clipboard_effects = None  # Will store (effects, item_id, item_name)

        # Track last selected character index for config saving
        self.last_char_index = None

        # Create notebook for tabs
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=10)

        # Create tabs
        self.file_tab = ttk.Frame(self.notebook)
        self.inventory_tab = ttk.Frame(self.notebook)
        self.vessels_tab = ttk.Frame(self.notebook)

        self.notebook.add(self.file_tab, text="File Management")
        self.notebook.add(self.inventory_tab, text="Relics")
        self.notebook.add(self.vessels_tab, text="Vessels")

        self.setup_file_tab()
        self.setup_inventory_tab()
        self.setup_vessels_tab()

        # Try to load last opened file after UI is set up
        self.root.after(100, self.try_load_last_file)
        
    def run_task_async(self, task_func, args=(), title="Processing", label_text="Please wait...", callback=None):
        """
        A universal async wrapper to run heavy tasks without freezing the GUI.
        
        :param task_func: The function to execute in background.
        :param args: Tuple of arguments to pass to the task_func.
        :param title: Window title of the loading popup.
        :param label_text: Message displayed in the loading popup.
        :param callback: A function to run in the main thread after task_func finishes successfully.
        """
        # 1. Create a top-level loading window
        loading_win = tk.Toplevel(self.root)
        loading_win.title(title)
        loading_win.geometry("350x150")
        loading_win.resizable(False, False)
        loading_win.attributes("-topmost", True)
        loading_win.transient(self.root)
        
        # Center the loading window relative to main window
        loading_win.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (loading_win.winfo_width() // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (loading_win.winfo_height() // 2)
        loading_win.geometry(f"+{x}+{y}")

        # UI Elements
        ttk.Label(loading_win, text=label_text, font=('Arial', 10), wraplength=300).pack(pady=20)
        progress = ttk.Progressbar(loading_win, mode='indeterminate', length=250)
        progress.pack(pady=10)
        progress.start(10)

        # Critical: Lock the main GUI to prevent data corruption/multiple clicks
        loading_win.grab_set()

        def thread_wrapper():
            try:
                # 2. Execute the heavy task
                task_func(*args)
                
                # 3. Task finished: Close window and run callback in main thread
                self.root.after(0, lambda: finish_task(True))
            except Exception as e:
                # Task failed: Close window and show error in main thread
                err_msg = str(e)
                self.root.after(0, lambda: finish_task(False, err_msg))

        def finish_task(success, err_msg=None):
            # Release the GUI lock and close popup
            loading_win.destroy()
            
            if success:
                # Execute the provided callback if any
                if callback:
                    callback()
            else:
                messagebox.showerror("Error", f"Task failed:\n{err_msg}")

        # 4. Start the background thread
        bg_thread = threading.Thread(target=thread_wrapper, daemon=True)
        bg_thread.start()

    def try_load_last_file(self):
        """Try to load the last opened save file and character"""
        global MODE

        config = load_config()
        last_file = config.get('last_file')
        last_mode = config.get('last_mode')
        last_char_index = config.get('last_char_index', 0)

        if not last_file or not os.path.exists(last_file):
            return

        try:
            # Set mode
            MODE = last_mode

            # Split files
            split_files(last_file, 'decrypted_output')

            # Load JSON data
            if not load_json_data():
                return
            self.update_inventory_comboboxes()
            self.update_vessel_tab_comboboxes()

            # Get character names
            name_to_path()

            # Display character buttons
            self.display_character_buttons()

            # Auto-select the last used character if valid
            if char_name_list and 0 <= last_char_index < len(char_name_list):
                name, path = char_name_list[last_char_index]
                self.on_character_click(last_char_index, path, name)

        except Exception as e:
            import traceback
            traceback.print_exc()
            print(f"Could not auto-load last file: {e}")

    def setup_file_tab(self):
        # Main container
        container = ttk.Frame(self.file_tab)
        container.pack(fill='both', expand=True, padx=15, pady=15)
        
        # File loading section
        file_frame = ttk.LabelFrame(container, text="Save File Management", padding=15)
        file_frame.pack(fill='x', pady=(0, 15))
        
        info_label = ttk.Label(file_frame, text="Load your Elden Ring NightReign save file to begin editing", 
                              foreground='gray')
        info_label.pack(pady=(0, 10))
        
        ttk.Button(file_frame, text="📁 Open Save File", command=self.open_file, width=20).pack(pady=5)
        ttk.Button(file_frame, text="💾 Save Modified File", command=self.save_changes).pack(padx=5)
        ttk.Button(file_frame, text="💾 Import save (PC/PS4)", command=self.import_save_tk).pack(padx=5)
        
        # Stats section
        stats_frame = ttk.LabelFrame(container, text="Character Statistics", padding=15)
        stats_frame.pack(fill='x', pady=(0, 15))
        
        # Murks row
        murks_row = ttk.Frame(stats_frame)
        murks_row.pack(fill='x', pady=5)
        
        ttk.Label(murks_row, text="Murks:", font=('Arial', 11, 'bold'), width=15).pack(side='left')
        self.murks_display = ttk.Label(murks_row, text="N/A", font=('Arial', 11), foreground='blue')
        self.murks_display.pack(side='left', padx=10)
        ttk.Button(murks_row, text="Edit", command=self.modify_murks, width=10).pack(side='right', padx=5)
        
        # Sigs row
        sigs_row = ttk.Frame(stats_frame)
        sigs_row.pack(fill='x', pady=5)
        
        ttk.Label(sigs_row, text="Sigs:", font=('Arial', 11, 'bold'), width=15).pack(side='left')
        self.sigs_display = ttk.Label(sigs_row, text="N/A", font=('Arial', 11), foreground='blue')
        self.sigs_display.pack(side='left', padx=10)
        ttk.Button(sigs_row, text="Edit", command=self.modify_sigs, width=10).pack(side='right', padx=5)
        
        ttk.Button(stats_frame, text="🔄 Refresh Stats", command=self.refresh_stats).pack(pady=(10, 0))
        
        # Character selection section
        char_frame = ttk.LabelFrame(container, text="Select Character", padding=15)
        char_frame.pack(fill='both', expand=True)
        
        ttk.Label(char_frame, text="Choose a character to load:", foreground='gray').pack(anchor='w', pady=(0, 10))
        
        # Scrollable character list
        canvas = tk.Canvas(char_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(char_frame, orient="vertical", command=canvas.yview)
        self.char_button_frame = ttk.Frame(canvas)
        
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side='right', fill='y')
        canvas.pack(side='left', fill='both', expand=True)
        canvas.create_window((0, 0), window=self.char_button_frame, anchor='nw')

        self.char_button_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    def setup_vessels_tab(self):
        """Setup the Vessels tab for viewing and managing character loadouts"""
        # Top controls
        controls_frame = ttk.Frame(self.vessels_tab)
        controls_frame.pack(fill='x', padx=10, pady=5)

        ttk.Button(controls_frame, text="🔄 Refresh", command=self.refresh_vessels).pack(side='left', padx=5)
        ttk.Button(controls_frame, text="💾 Save Loadout", command=self.save_loadout).pack(side='left', padx=5)
        ttk.Button(controls_frame, text="📂 Load Loadout", command=self.load_loadout).pack(side='left', padx=5)

        # Character selector
        ttk.Label(controls_frame, text="Character:").pack(side='left', padx=(20, 5))
        # Only show the 10 real playable characters in the dropdown (not internal parsing names)
        playable_characters = data_source.character_names[:10]
        self.vessel_char_var = tk.StringVar(value=playable_characters[0])
        self.vessel_char_combo = ttk.Combobox(controls_frame, textvariable=self.vessel_char_var,
                                               values=playable_characters, state="readonly", width=12)
        self.vessel_char_combo.pack(side='left', padx=5)
        self.vessel_char_combo.bind("<<ComboboxSelected>>", lambda e: self.refresh_vessels())

        # Main content - split into vessel slots
        self.vessels_content = ttk.Frame(self.vessels_tab)
        self.vessels_content.pack(fill='both', expand=True, padx=10, pady=5)

        # Create vessel slot frames (7 vessels per character)
        self.vessel_frames = []
        self.vessel_trees = []

        # Use a canvas with scrollbar for the vessel display
        canvas_frame = ttk.Frame(self.vessels_content)
        canvas_frame.pack(fill='both', expand=True)

        self.vessels_canvas = tk.Canvas(canvas_frame)
        scrollbar = ttk.Scrollbar(canvas_frame, orient='vertical', command=self.vessels_canvas.yview)
        self.vessels_inner_frame = ttk.Frame(self.vessels_canvas)

        self.vessels_canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side='right', fill='y')
        self.vessels_canvas.pack(side='left', fill='both', expand=True)
        self.vessels_window_id = self.vessels_canvas.create_window((0, 0), window=self.vessels_inner_frame, anchor='nw')

        def on_inner_frame_configure(e):
            self.vessels_canvas.configure(scrollregion=self.vessels_canvas.bbox("all"))

        def on_canvas_configure(e):
            # Make inner frame fill the canvas width
            self.vessels_canvas.itemconfig(self.vessels_window_id, width=e.width)

        self.vessels_inner_frame.bind("<Configure>", on_inner_frame_configure)
        self.vessels_canvas.bind("<Configure>", on_canvas_configure)

        # Enable mouse wheel scrolling anywhere in the canvas
        def _on_vessels_mousewheel(event):
            self.vessels_canvas.yview_scroll(int(-1*(event.delta/120)), "units")

        # Bind mousewheel to canvas and inner frame
        self.vessels_canvas.bind("<MouseWheel>", _on_vessels_mousewheel)
        self.vessels_inner_frame.bind("<MouseWheel>", _on_vessels_mousewheel)

        # Also bind to canvas_frame for better coverage
        canvas_frame.bind("<MouseWheel>", _on_vessels_mousewheel)

        # Configure inner frame to expand column 0
        self.vessels_inner_frame.columnconfigure(0, weight=1)

        # Create 11 vessel slot displays (1 column layout)
        # Vessel names will be updated dynamically based on selected character
        def open_new_preset_name_dialog():
            # Dialog to enter new preset name
            def check_regex(P):
                if re.fullmatch(r"[a-zA-Z0-9 ]{0,18}", P):
                    return True
                return False
            vcmd = self.root.register(check_regex)
            dialog = tk.Toplevel(self.root)
            dialog.title("New Preset Name")
            dialog.transient(self.root)
            dialog.grab_set()
            
            label = ttk.Label(dialog, text="Enter name for new preset:")
            label.pack(pady=10, padx=10)
            
            name_entry = ttk.Entry(dialog, validate='key', validatecommand=(vcmd, '%P'), width=30)
            name_entry.pack(pady=5, padx=10)
            name_entry.focus_set()
            
            result = {"name": None}  # Use a dict to pass value back
            
            def on_ok():
                result["name"] = name_entry.get().strip()
                if result["name"] is None or result["name"].strip() == "":
                    messagebox.showerror("Error", "Preset name cannot be empty")
                    return
                dialog.destroy()
            
            def on_cancel():
                dialog.destroy()

            ok_button = ttk.Button(dialog, text="OK", command=on_ok)
            ok_button.pack(side=tk.LEFT, padx=5, pady=10)
            
            cancel_button = ttk.Button(dialog, text="Cancel", command=on_cancel)
            cancel_button.pack(side=tk.RIGHT, padx=5, pady=10)
            
            self.root.wait_window(dialog)
            return result["name"]
        
        def on_add_to_preset(vessel_slot):
            hero_type = self.vessel_char_combo.current()+1
            preset_name = open_new_preset_name_dialog()
            vessel_id = loadout_handler.heroes[hero_type].vessels[vessel_slot]['vessel_id']
            relics = loadout_handler.heroes[hero_type].vessels[vessel_slot]['relics']
            if preset_name is None:
                return
            try:
                loadout_handler.push_preset(hero_type, vessel_id, relics, preset_name)
                save_current_data()
            except Exception as e:
                messagebox.showerror("Error", str(e))
                return
            self.refresh_inventory_and_vessels()
            
        # Store status labels for updating
        self.vessel_status_labels = []
            
        for i in range(11):
            vessel_frame = ttk.LabelFrame(self.vessels_inner_frame, text=f"Vessel Slot {i}")
            vessel_frame.grid(row=i, column=0, padx=5, pady=5, sticky='nsew')

            # Bind mousewheel to vessel frame
            vessel_frame.bind("<MouseWheel>", _on_vessels_mousewheel)

            # Status label for unlock state
            label_line_frame = ttk.Frame(vessel_frame)
            label_line_frame.pack(padx=5, pady=2, fill='x')
            status_label = ttk.Label(label_line_frame, text="")
            status_label.pack(side='left', padx=5)
            add_to_preset_button = ttk.Button(label_line_frame, text="➕ Add to Preset", command=lambda v=i: on_add_to_preset(v))
            add_to_preset_button.pack(side='right', padx=5)

            # Treeview for relics in this vessel - 6 slots (3 normal + 3 deep relics)
            columns = ('Slot', 'Type', 'Color', 'Relic Name', 'ID', 'Effect 1', 'Effect 2', 'Effect 3')
            tree = ttk.Treeview(vessel_frame, columns=columns, show='headings', height=6)

            tree.heading('Slot', text='#')
            tree.heading('Type', text='Type')
            tree.heading('Color', text='Color')
            tree.heading('Relic Name', text='Name')
            tree.heading('ID', text='ID')
            tree.heading('Effect 1', text='Effect 1')
            tree.heading('Effect 2', text='Effect 2')
            tree.heading('Effect 3', text='Effect 3')

            tree.column('Slot', width=25)
            tree.column('Type', width=50)
            tree.column('Color', width=55)
            tree.column('Relic Name', width=150)
            tree.column('ID', width=70)
            tree.column('Effect 1', width=120)
            tree.column('Effect 2', width=120)
            tree.column('Effect 3', width=120)

            tree.pack(fill='both', expand=True, padx=5, pady=5)

            # Bind mousewheel to tree
            tree.bind("<MouseWheel>", _on_vessels_mousewheel)

            # Configure color tags for the treeview
            for color_name, color_hex in RELIC_COLOR_HEX.items():
                if color_name:
                    tree.tag_configure(color_name, foreground=color_hex)
            # Add special tag for empty slots
            tree.tag_configure('empty', foreground='#666666')
            # Add tag for color mismatch (wrong color relic in slot)
            tree.tag_configure('mismatch', foreground='#FF6B6B', background='#331111')
            # Add tag for deep relic slots (slots 4-6) - just use a subtle indicator
            tree.tag_configure('deep_slot', foreground='#9999BB')

            # Bind right-click for context menu
            tree.bind('<Button-3>', lambda e, v=i: self.show_vessel_context_menu(e, v))
            # Bind double-click to open replace dialog
            tree.bind('<Double-1>', lambda e, v=i: self.on_vessel_relic_double_click(e, v))

            self.vessel_frames.append(vessel_frame)
            self.vessel_trees.append(tree)
            self.vessel_status_labels.append(status_label)

        # ============================================
        # PRESETS SECTION
        # ============================================
        self.presets_section = ttk.LabelFrame(self.vessels_inner_frame, text="📁 Saved Presets")
        self.presets_section.grid(row=11, column=0, padx=5, pady=10, sticky='nsew')

        # Modern card-based layout - no internal scroll, expands with content
        # Cards flow naturally and main vessels canvas handles all scrolling
        self.presets_cards_frame = tk.Frame(self.presets_section, bg='#2a2a3d')
        self.presets_cards_frame.pack(fill='both', expand=True, padx=2, pady=2)

        # Bind mousewheel to scroll the main vessels canvas
        self.presets_cards_frame.bind("<MouseWheel>", _on_vessels_mousewheel)
        self.presets_section.bind("<MouseWheel>", _on_vessels_mousewheel)

        # Store preset data for lookup (widget -> preset info)
        self.preset_data_map = {}

        # Configure grid weights (single column)
        for i in range(12):  # 11 vessels + 1 presets section
            self.vessels_inner_frame.grid_rowconfigure(i, weight=1)
        self.vessels_inner_frame.grid_columnconfigure(0, weight=1)

    def refresh_vessels(self):
        """Refresh the vessels display for the selected character"""
        if globals.data is None:
            return
        
        self.update_vessel_tab_comboboxes()

        char_name = self.vessel_char_var.get()
        loadout = get_character_loadout(char_name)

        # Clear and populate each vessel tree
        for vessel_slot, tree in enumerate(self.vessel_trees):
            # Clear existing items
            for item in tree.get_children():
                tree.delete(item)

            # Get vessel info for this character
            vessel_data_info = get_vessel_info(char_name, vessel_slot)
            vessel_name = vessel_data_info['name']
            hero_type = self.vessel_char_combo.current()+1
            vessel_id = loadout_handler.heroes[hero_type].vessels[vessel_slot]['vessel_id']

            # Update vessel frame title with actual vessel name
            if vessel_slot < len(self.vessel_frames):
                self.vessel_frames[vessel_slot].config(text=vessel_name)

                vessel_info = loadout.get(vessel_slot, {})
                has_relics = vessel_info.get('has_relics', False)
                is_unlocked = is_vessel_unlocked(vessel_id, data_source)

                # Update status label
                if vessel_slot < len(self.vessel_status_labels):
                    unlock_flag = vessel_data_info.get('unlockFlag', 0)
                    if is_unlocked:
                        # Has relics = definitely unlocked
                        relic_count = sum(1 for _, r in vessel_info.get('relics', []) if r is not None)
                        self.vessel_status_labels[vessel_slot].config(
                            text=f"✅ Unlocked ({relic_count}/6 relics)",
                            foreground='green')
                    elif vessel_slot == 0 or unlock_flag == 0:
                        # Slot 0 is always unlocked (default vessel)
                        # unlock_flag == 0 means no flag required (always unlocked)
                        self.vessel_status_labels[vessel_slot].config(
                            text="✅ Unlocked (Empty)",
                            foreground='green')
                    else:
                        # Empty non-zero slots - show unlock flag ID for reference
                        self.vessel_status_labels[vessel_slot].config(
                            text=f"❓ Empty (Unlock Flag: {unlock_flag})",
                            foreground='#888888')

            # Get relics for this vessel
            vessel_info = loadout.get(vessel_slot, {'relics': [], 'unlocked': False})
            relics = vessel_info.get('relics', [])

            # Ensure we always have 6 slots displayed (3 normal + 3 deep relic slots)
            while len(relics) < 6:
                relics.append((0, None))
            # Only show first 6 slots
            relics = relics[:6]

            # Get vessel colors for empty slot display
            vessel_colors = vessel_data_info.get('colors', None)

            for idx, (ga, relic_info) in enumerate(relics):
                # Determine if this is a deep slot (slots 4-6, index 3-5)
                is_deep_slot = idx >= 3
                slot_type = "🔮 Deep" if is_deep_slot else "Normal"

                # Get expected slot color from vessel data
                # Colors tuple: (normal1, normal2, normal3, deep1, deep2, deep3)
                expected_color = None
                if vessel_colors:
                    expected_color = vessel_colors[idx] if idx < len(vessel_colors) else None

                if relic_info:
                    # Get effect names (full text, no truncation)
                    effect_names = []
                    for eff in relic_info['effects']:
                        if eff == 0:
                            effect_names.append("None")
                        elif eff == 4294967295:
                            effect_names.append("Empty")
                        elif str(eff) in effects_json:
                            eff_name = effects_json[str(eff)]["name"].replace('\n', ' ').strip()
                            effect_names.append(eff_name)
                        else:
                            effect_names.append(f"Unknown")

                    # Get relic color
                    relic_color = relic_info.get('color', 'Unknown')
                    relic_color_display = relic_color if relic_color else "?"

                    # Build tags - use relic color for coloring, deep_slot for background
                    tags = [relic_color]
                    if is_deep_slot:
                        tags.append('deep_slot')

                    tree.insert('', 'end', values=(
                        idx + 1,
                        slot_type,
                        relic_color_display,
                        relic_info['name'],
                        relic_info['real_id'],
                        effect_names[0],
                        effect_names[1],
                        effect_names[2]
                    ), tags=tuple(tags))
                else:
                    # Empty slot - show expected color from vessel data
                    slot_color_display = expected_color if expected_color else "-"
                    tags = ['empty']
                    if expected_color:
                        tags.append(expected_color)  # Add color tag for styling
                    if is_deep_slot:
                        tags.append('deep_slot')

                    tree.insert('', 'end', values=(
                        idx + 1,
                        slot_type,
                        slot_color_display,
                        "(Empty)",
                        "-",
                        "-",
                        "-",
                        "-"
                    ), tags=tuple(tags))

            # Auto-size columns after populating
            autosize_treeview_columns(tree)

        # Refresh presets section
        self.refresh_presets()

    def refresh_presets(self):
        """Refresh the presets display with modern card layout"""
        # Clear existing cards
        for widget in self.presets_cards_frame.winfo_children():
            widget.destroy()
        self.preset_data_map = {}

        char_name = self.vessel_char_var.get()
        char_id = data_source.character_names.index(char_name) if char_name in data_source.character_names else -1
        hero_type = char_id + 1
        if char_id == -1:
            return

        # Helper to bind mousewheel scrolling to all widgets in a tree
        def bind_scroll_recursive(widget):
            """Bind mousewheel to widget and all descendants for main vessels canvas scrolling"""
            def scroll_vessels(event):
                self.vessels_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            widget.bind("<MouseWheel>", scroll_vessels)
            for child in widget.winfo_children():
                bind_scroll_recursive(child)

        # Build a lookup dict from ga_relic list for quick access (with effects)
        ga_to_relic_info = {}
        for relic in ga_relic:
            if len(relic) >= 5:
                ga_handle = relic[0]
                item_id = relic[1]
                real_id = item_id - 2147483648 if item_id > 2147483648 else item_id
                item_info = items_json.get(str(real_id), {})
                is_deep = 2000000 <= real_id <= 2019999
                effects = [relic[2], relic[3], relic[4]] if len(relic) >= 5 else []
                curses = [relic[5], relic[6], relic[7]] if len(relic) >= 8 else []
                ga_to_relic_info[ga_handle] = {
                    'name': item_info.get('name', f'ID:{real_id}'),
                    'color': item_info.get('color', 'Unknown'),
                    'real_id': real_id,
                    'is_deep': is_deep,
                    'effects': effects,
                    'curses': curses
                }

        # Colors for UI
        BG_CARD = '#363650'
        BG_HEADER = '#454570'
        FG_TEXT = '#e8e8e8'
        FG_DIM = '#9999aa'
        FG_EFFECT = '#88cc88'
        FG_CURSE = '#cc8888'

        color_hex_map = {
            'Red': '#ff6666', 'Blue': '#6699ff', 'Yellow': '#ddbb44',
            'Green': '#66cc66', 'Purple': '#bb77ff', 'Orange': '#ff9944',
            'White': '#cccccc', 'Unknown': '#888888'
        }

        # Find presets for this character
        card_row = 0
        
        presets = loadout_handler.heroes[hero_type].presets

        for preset in presets:
            vessel_id = preset.get('vessel_id', 0)
            vessel_slot = loadout_handler.get_vessel_index_in_hero(hero_type, vessel_id)
            preset_name = preset.get('name', 'Unknown')
            ga_handles = preset.get('relics', [])

            # Collect relic info with names and effects grouped per relic
            relic_data_list = []  # For color indicators
            relics_with_effects = []  # List of (relic_name, color, is_deep, effects, curses)

            for ga in ga_handles:
                if ga != 0 and (ga & 0xF0000000) == ITEM_TYPE_RELIC:
                    relic_info = ga_to_relic_info.get(ga)
                    if relic_info:
                        color = relic_info.get('color', 'Unknown')
                        is_deep = relic_info.get('is_deep', False)
                        relic_name = relic_info.get('name', 'Unknown Relic')
                        relic_data_list.append({'color': color, 'is_deep': is_deep})

                        # Collect effects for this relic
                        effects = []
                        for eff in relic_info.get('effects', []):
                            if eff and eff not in [0, -1, 4294967295]:
                                eff_name = effects_json.get(str(eff), {}).get('name', f'Effect {eff}')
                                effects.append(eff_name)

                        # Collect curses for this relic
                        curses = []
                        for curse in relic_info.get('curses', []):
                            if curse and curse not in [0, -1, 4294967295]:
                                curse_name = effects_json.get(str(curse), {}).get('name', f'Curse {curse}')
                                curses.append(curse_name)

                        relics_with_effects.append((relic_name, color, is_deep, effects, curses))

            # Get vessel name
            vessel_info = get_vessel_info(char_name, vessel_slot)
            vessel_name = vessel_info.get('name', f'Vessel {vessel_slot}')

            # Create card frame
            card = tk.Frame(self.presets_cards_frame, bg=BG_CARD, relief='flat', bd=0)
            card.pack(fill='x', padx=8, pady=6)

            # Header row with name, vessel, and color indicators
            header = tk.Frame(card, bg=BG_HEADER, cursor='hand2')
            header.pack(fill='x', padx=2, pady=2)

            # Collapse/expand indicator
            collapse_var = tk.StringVar(value='▼')
            collapse_lbl = tk.Label(header, textvariable=collapse_var, font=('Segoe UI', 9),
                                    fg=FG_DIM, bg=BG_HEADER, cursor='hand2')
            collapse_lbl.pack(side='left', padx=(10, 0), pady=6)

            tk.Label(header, text=f"📋 {preset_name}", font=('Segoe UI', 11, 'bold'),
                        fg=FG_TEXT, bg=BG_HEADER).pack(side='left', padx=(5, 10), pady=6)
            tk.Label(header, text=f"({vessel_name})", font=('Segoe UI', 9),
                        fg=FG_DIM, bg=BG_HEADER).pack(side='left', padx=5)

            # Color indicators on the right
            colors_frame = tk.Frame(header, bg=BG_HEADER)
            colors_frame.pack(side='right', padx=10)

            for rd in relic_data_list:
                color = rd['color']
                is_deep = rd['is_deep']
                c_hex = color_hex_map.get(color, '#888888')
                indicator = tk.Label(colors_frame, text='●' if not is_deep else '◆',
                                    font=('Segoe UI', 12), fg=c_hex, bg=BG_HEADER)
                indicator.pack(side='left', padx=2)

            # Collapsible content frame
            content_frame = tk.Frame(card, bg=BG_CARD)
            content_frame.pack(fill='x')

            # Relics section - show each relic with its effects
            effects_frame = tk.Frame(content_frame, bg=BG_CARD)
            effects_frame.pack(fill='x', padx=12, pady=(8, 10))

            if relics_with_effects:
                for relic_name, color, is_deep, effects, curses in relics_with_effects:
                    c_hex = color_hex_map.get(color, FG_EFFECT)

                    # Relic name header with deep indicator
                    relic_prefix = "🔮 " if is_deep else "◆ "
                    tk.Label(effects_frame, text=f"{relic_prefix}{relic_name}",
                            font=('Segoe UI', 9, 'bold'), fg=c_hex, bg=BG_CARD,
                            anchor='w').pack(anchor='w', pady=(4, 0))

                    # Effects for this relic
                    for eff_name in effects:
                        eff_name = eff_name.replace('\n', ' ').replace('\r', ' ').strip()
                        tk.Label(effects_frame, text=f"    ✦ {eff_name}",
                                font=('Segoe UI', 9), fg=c_hex, bg=BG_CARD,
                                anchor='w').pack(anchor='w')

                    # Curses for this relic
                    for curse_name in curses:
                        curse_name = curse_name.replace('\n', ' ').replace('\r', ' ').strip()
                        tk.Label(effects_frame, text=f"    ⚠ {curse_name}",
                                font=('Segoe UI', 9, 'italic'), fg=FG_CURSE, bg=BG_CARD,
                                anchor='w').pack(anchor='w')
            else:
                tk.Label(effects_frame, text="No relics", font=('Segoe UI', 9, 'italic'),
                            fg=FG_DIM, bg=BG_CARD).pack(anchor='w')

            # Edit button (inside content_frame so it collapses too)
            btn_frame = tk.Frame(content_frame, bg=BG_CARD)
            btn_frame.pack(fill='x', padx=10, pady=(0, 8))

            preset_data = {
                'char_name': char_name,
                'char_id': char_id,
                'vessel_slot': vessel_slot,
                'preset': preset,
                'ga_to_relic_info': ga_to_relic_info
            }
            
            def on_equip_preset(target_preset, target_vessel_slot):
                cur_preset_name = target_preset.get('name', 'Unknown')
                _vessel_info = get_vessel_info(char_name, target_vessel_slot)
                cur_vessel_name = _vessel_info.get('name', f'Vessel {target_vessel_slot}')
                cur_preset = target_preset
                if messagebox.askyesno("Confirm Load",
                                       f"Are you sure you want to load preset '{cur_preset_name}' "
                                       f"into {char_name}'s {cur_vessel_name}?\n\n"
                                       f"This will overwrite the current relics in that vessel."):
                    try:
                        loadout_handler.equip_preset(cur_preset['hero_type'], cur_preset['index'])
                        save_current_data()
                        messagebox.showinfo("Success", f"Preset '{cur_preset_name}' loaded successfully!")
                        self.refresh_inventory_and_vessels()
                    except Exception as e:
                        messagebox.showerror("Error", f"Failed to load preset: {str(e)}")
                
            # Edit button
            equip_btn = ttk.Button(btn_frame, text="⚙️ Equip Preset",
                                    command=lambda pd=preset_data: on_equip_preset(pd['preset'], pd['vessel_slot']))
            equip_btn.pack(side='right', padx=5)
            edit_btn = ttk.Button(btn_frame, text="✏️ Edit Preset",
                                    command=lambda pd=preset_data: self.edit_preset_relics(pd))
            edit_btn.pack(side='right', padx=5)

            # Toggle function for collapse/expand
            def make_toggle(cf, cv):
                def toggle(event=None):
                    if cf.winfo_viewable():
                        cf.pack_forget()
                        cv.set('▶')
                    else:
                        cf.pack(fill='x')
                        cv.set('▼')
                return toggle

            toggle_fn = make_toggle(content_frame, collapse_var)
            header.bind('<Button-1>', toggle_fn)
            collapse_lbl.bind('<Button-1>', toggle_fn)
            # Make all children of header clickable
            for child in header.winfo_children():
                child.bind('<Button-1>', toggle_fn)

            # Bind scroll to all widgets in this card
            bind_scroll_recursive(card)

            # Store for reference
            self.preset_data_map[id(card)] = preset_data
            card_row += 1

        # Show message if no presets
        if card_row == 0:
            no_presets = tk.Label(self.presets_cards_frame,
                                  text="No saved presets for this character",
                                  font=('Segoe UI', 10, 'italic'), fg=FG_DIM, bg='#2a2a3d')
            no_presets.pack(pady=20)
            bind_scroll_recursive(no_presets)

    def edit_preset_relics(self, preset_info):
        """Open dialog to edit relics in a preset"""
        # Debug at start of function
        with open('preset_debug.txt', 'w') as f:
            f.write("edit_preset_relics called\n")
            f.write(f"preset_info keys: {preset_info.keys()}\n")

        char_name = preset_info['char_name']
        vessel_slot = preset_info['vessel_slot']
        preset = preset_info['preset']
        ga_to_relic_info = preset_info['ga_to_relic_info']

        preset_name = preset.get('name', 'Unknown')
        preset_offset = preset.get('offsets', 0)
        ga_handles = preset.get('relics', [])

        # Build extended relic info with effects from ga_relic
        ga_to_full_info = {}
        for relic in ga_relic:
            if len(relic) >= 8:
                ga_handle = relic[0]
                item_id = relic[1]
                real_id = item_id - 2147483648 if item_id > 2147483648 else item_id
                item_info = items_json.get(str(real_id), {})
                ga_to_full_info[ga_handle] = {
                    'name': item_info.get('name', f'ID:{real_id}'),
                    'color': item_info.get('color', 'Unknown'),
                    'real_id': real_id,
                    'effects': [relic[2], relic[3], relic[4]],
                    'curses': [relic[5], relic[6], relic[7]],
                    'ga_handle': ga_handle
                }

        # Create edit dialog - game-like layout
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Edit Preset: {preset_name}")
        dialog.geometry("1100x700")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg='#1a1a2e')

        # Color scheme matching game
        BG_DARK = '#1a1a2e'
        BG_PANEL = '#252545'
        BG_SLOT = '#2d2d50'
        BG_SLOT_SELECTED = '#4a4a80'
        FG_TEXT = '#e0e0e0'
        FG_DIM = '#888899'
        FG_CURSE = '#cc6666'

        # Store state
        slot_relic_data = {}
        selected_slot = tk.IntVar(value=0)
        slot_frames = []

        def get_effect_name(eff_id):
            if eff_id in [0, -1, 4294967295]:
                return None
            return effects_json.get(str(eff_id), {}).get('name', f'Unknown ({eff_id})')

        # ============ LEFT PANEL ============
        left_panel = tk.Frame(dialog, bg=BG_DARK, width=420)
        left_panel.pack(side='left', fill='y', padx=10, pady=10)
        left_panel.pack_propagate(False)

        # Header
        header = tk.Label(left_panel, text=f"{char_name} - {preset_name}",
                          font=('Segoe UI', 12, 'bold'), fg=FG_TEXT, bg=BG_DARK)
        header.pack(anchor='w', pady=(0, 10))

        vessel_info = get_vessel_info(char_name, vessel_slot)
        vessel_name = vessel_info.get('name', f'Vessel {vessel_slot}')
        tk.Label(left_panel, text=vessel_name, font=('Segoe UI', 10),
                 fg=FG_DIM, bg=BG_DARK).pack(anchor='w')

        # Slots row (horizontal like game UI)
        slots_frame = tk.Frame(left_panel, bg=BG_DARK)
        slots_frame.pack(fill='x', pady=15)

        def update_slot_display():
            """Refresh slot button appearances"""
            for idx, frame in enumerate(slot_frames):
                is_selected = selected_slot.get() == idx
                relic_data = slot_relic_data.get(idx)
                color = relic_data.get('color', '') if relic_data else ''

                # Background color
                bg = BG_SLOT_SELECTED if is_selected else BG_SLOT
                frame.configure(bg=bg)
                for child in frame.winfo_children():
                    child.configure(bg=bg)

                # Border color based on relic color
                border_color = RELIC_COLOR_HEX.get(color, '#555555')
                frame.configure(highlightbackground=border_color,
                               highlightcolor=border_color if is_selected else border_color)

        def select_slot(idx):
            selected_slot.set(idx)
            update_slot_display()
            update_details_panel()

        def on_slot_double_click(idx):
            # Just select the slot - the relic list on the right will auto-update
            select_slot(idx)

        # Create 6 slot buttons
        for idx in range(6):
            is_deep = idx >= 3
            ga = ga_handles[idx] if idx < len(ga_handles) else 0

            if ga != 0 and (ga & 0xF0000000) == ITEM_TYPE_RELIC:
                relic_info = ga_to_full_info.get(ga, ga_to_relic_info.get(ga, {}))
                slot_relic_data[idx] = relic_info
            else:
                slot_relic_data[idx] = None

            # Slot frame
            slot_frame = tk.Frame(slots_frame, bg=BG_SLOT, width=60, height=70,
                                  highlightthickness=2, highlightbackground='#555555')
            slot_frame.pack(side='left', padx=3)
            slot_frame.pack_propagate(False)
            slot_frames.append(slot_frame)

            # Slot number
            slot_num = tk.Label(slot_frame, text=f"{'🔮' if is_deep else ''}{idx+1}",
                               font=('Segoe UI', 9), fg=FG_DIM, bg=BG_SLOT)
            slot_num.pack(pady=(3, 0))

            # Relic indicator (color square)
            relic_data = slot_relic_data.get(idx)
            color = relic_data.get('color', '') if relic_data else ''
            indicator_color = RELIC_COLOR_HEX.get(color, '#333333')
            indicator = tk.Frame(slot_frame, bg=indicator_color, width=30, height=30)
            indicator.pack(pady=5)

            # Bind clicks
            for widget in [slot_frame, slot_num, indicator]:
                widget.bind('<Button-1>', lambda e, i=idx: select_slot(i))
                widget.bind('<Double-Button-1>', lambda e, i=idx: on_slot_double_click(i))

        # Deep slots label
        tk.Label(left_panel, text="Slots 1-3: Normal  |  Slots 4-6: 🔮 Deep",
                 font=('Segoe UI', 8), fg=FG_DIM, bg=BG_DARK).pack(anchor='w', pady=(0, 10))

        # Details panel (below slots)
        details_panel = tk.Frame(left_panel, bg=BG_PANEL, relief='flat')
        details_panel.pack(fill='both', expand=True, pady=10)

        # Relic name label
        relic_name_label = tk.Label(details_panel, text="", font=('Segoe UI', 11, 'bold'),
                                    fg=FG_TEXT, bg=BG_PANEL, anchor='w')
        relic_name_label.pack(fill='x', padx=10, pady=(10, 5))

        # Effects container
        effects_container = tk.Frame(details_panel, bg=BG_PANEL)
        effects_container.pack(fill='both', expand=True, padx=10, pady=5)

        def is_unique_relic(relic_id):
            """Check if a relic is unique (all effect pools contain only 1 effect each)"""
            try:
                pools = data_source.get_relic_pools_seq(relic_id)
                # Check each non-empty effect pool
                for pool_id in pools[:3]:  # Only check effect pools, not curse pools
                    if pool_id == -1:
                        continue
                    pool_effects = data_source.get_pool_effects(pool_id)
                    if len(pool_effects) > 1:
                        return False  # Has multiple effects, not unique
                return True  # All pools have 0 or 1 effect
            except (KeyError, Exception):
                return True  # If we can't determine, treat as unique (safer)

        def update_details_panel():
            """Update the details panel with selected slot's relic info"""
            # Clear effects container
            for widget in effects_container.winfo_children():
                widget.destroy()

            idx = selected_slot.get()
            relic_data = slot_relic_data.get(idx)

            if not relic_data:
                relic_name_label.config(text="(Empty Slot)")
                tk.Label(effects_container, text="Double-click slot or use picker to assign a relic",
                        fg=FG_DIM, bg=BG_PANEL, font=('Segoe UI', 9)).pack(anchor='w')
                # Disable edit button for empty slots
                if 'edit_btn' in dir():
                    edit_btn.config(state='disabled')
                return

            relic_name_label.config(text=f"📦 {relic_data['name']}")

            # Check if this is a unique relic (can't be edited)
            real_id = relic_data.get('real_id', 0)
            is_unique = is_unique_relic(real_id)

            # Show effects and curses paired together (like game UI)
            effects = relic_data.get('effects', [])
            curses = relic_data.get('curses', [])
            color = relic_data.get('color', 'Unknown')
            color_hex = RELIC_COLOR_HEX.get(color, '#888888')

            for i in range(3):
                eff = effects[i] if i < len(effects) else 0
                curse = curses[i] if i < len(curses) else 0

                eff_name = get_effect_name(eff)
                curse_name = get_effect_name(curse)

                if eff_name or curse_name:
                    # Effect row frame
                    row = tk.Frame(effects_container, bg=BG_PANEL)
                    row.pack(fill='x', pady=3)

                    # Color indicator square
                    indicator = tk.Frame(row, bg=color_hex, width=12, height=12)
                    indicator.pack(side='left', padx=(0, 8))

                    # Text container
                    text_frame = tk.Frame(row, bg=BG_PANEL)
                    text_frame.pack(side='left', fill='x')

                    # Blessing
                    if eff_name:
                        tk.Label(text_frame, text=eff_name, font=('Segoe UI', 9),
                                fg=FG_TEXT, bg=BG_PANEL, anchor='w').pack(anchor='w')

                    # Curse (in red, italicized)
                    if curse_name:
                        tk.Label(text_frame, text=curse_name, font=('Segoe UI', 9, 'italic'),
                                fg=FG_CURSE, bg=BG_PANEL, anchor='w').pack(anchor='w')

            # Update edit button state based on whether relic is unique
            # (nonlocal edit_btn defined after button creation)
            try:
                if is_unique:
                    edit_btn.config(state='disabled')
                else:
                    edit_btn.config(state='normal')
            except NameError:
                pass  # Button not created yet

        # Buttons below details
        btn_frame = tk.Frame(left_panel, bg=BG_DARK)
        btn_frame.pack(fill='x', pady=10)

        def clear_selected_slot():
            idx = selected_slot.get()
            if not slot_relic_data.get(idx):
                return
            if not messagebox.askyesno("Clear Slot", f"Clear relic from slot {idx + 1}?"):
                return
            ga_handles[idx] = 0
            try:
                loadout_handler.replace_preset_relic(preset['hero_type'], idx, 0, preset_index=preset['index'])
                slot_relic_data[idx] = None
                update_slot_display()
                update_details_panel()
                self.refresh_presets()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to clear relic from preset:\n{e}")
            # self.write_preset_relic(preset_offset, idx, 0)
            # slot_relic_data[idx] = None
            # update_slot_display()
            # update_details_panel()
            # self.refresh_presets()

        def edit_selected_relic():
            """Open the modify dialog for the relic in the selected slot"""
            idx = selected_slot.get()
            relic_data = slot_relic_data.get(idx)
            if not relic_data:
                messagebox.showinfo("No Relic", "Select a slot with a relic to edit.")
                return

            ga = ga_handles[idx]
            real_id = relic_data.get('real_id', 0)

            # Release grab so we can interact with the edit dialog
            dialog.grab_release()

            def refresh_after_edit():
                # Refresh main inventory (this updates ga_relic)
                self.refresh_inventory_lightly()
                # Rebuild relic info from updated ga_relic
                for relic in ga_relic:
                    if len(relic) >= 8:
                        ga_handle = relic[0]
                        item_id = relic[1]
                        real_id = item_id - 2147483648 if item_id > 2147483648 else item_id
                        item_info = items_json.get(str(real_id), {})
                        ga_to_full_info[ga_handle] = {
                            'name': item_info.get('name', f'ID:{real_id}'),
                            'color': item_info.get('color', 'Unknown'),
                            'real_id': real_id,
                            'effects': [relic[2], relic[3], relic[4]],
                            'curses': [relic[5], relic[6], relic[7]],
                        }
                # Reload relic data for this slot
                new_relic_data = ga_to_full_info.get(ga)
                if new_relic_data:
                    slot_relic_data[idx] = new_relic_data
                    update_slot_display()
                    update_details_panel()

            def on_edit_dialog_close():
                # Restore grab when edit dialog closes
                if dialog.winfo_exists():
                    dialog.grab_set()

            # Open or reuse modify dialog
            if not self.modify_dialog or not self.modify_dialog.dialog.winfo_exists():
                self.modify_dialog = ModifyRelicDialog(self.root, ga, real_id, refresh_after_edit)
            else:
                self.modify_dialog.load_relic(ga, real_id)
                self.modify_dialog.callback = refresh_after_edit

            # Set up close protocol and bring to front
            self.modify_dialog.dialog.protocol("WM_DELETE_WINDOW", lambda: [self.modify_dialog.dialog.destroy(), on_edit_dialog_close()])
            self.modify_dialog.dialog.lift()
            self.modify_dialog.dialog.focus_force()

        ttk.Button(btn_frame, text="Clear Slot", command=clear_selected_slot).pack(side='left', padx=5)
        edit_btn = ttk.Button(btn_frame, text="✏️ Edit Relic", command=edit_selected_relic)
        edit_btn.pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Close", command=dialog.destroy).pack(side='right', padx=5)

        # ============ RIGHT PANEL - Relic List Picker ============
        right_panel = tk.Frame(dialog, bg=BG_DARK)
        right_panel.pack(side='right', fill='both', expand=True, padx=10, pady=10)

        # Filter bar
        filter_frame = tk.Frame(right_panel, bg=BG_DARK)
        filter_frame.pack(fill='x', pady=(0, 10))

        relic_list_title = tk.Label(filter_frame, text="Available Relics", font=('Segoe UI', 11, 'bold'),
                                    fg=FG_TEXT, bg=BG_DARK)
        relic_list_title.pack(side='left')

        # Show all colors checkbox
        show_all_colors_var = tk.BooleanVar(value=False)
        show_all_cb = ttk.Checkbutton(filter_frame, text="Show all colors",
                                       variable=show_all_colors_var)
        show_all_cb.pack(side='right', padx=5)

        # Search entry
        search_var = tk.StringVar()
        search_entry = ttk.Entry(filter_frame, textvariable=search_var, width=20)
        search_entry.pack(side='right', padx=5)
        tk.Label(filter_frame, text="🔍", font=('Segoe UI', 10), fg=FG_DIM, bg=BG_DARK).pack(side='right')
        
        # Equippe by
        equipped_by_frame = tk.Frame(right_panel, bg=BG_DARK)
        equipped_by_frame.pack(fill='x', pady=(0, 10))
        equipped_by_var = tk.StringVar(value='None')
        equipped_by_cb = ttk.Combobox(equipped_by_frame, textvariable=equipped_by_var, state='readonly',
                                      values=['None', 'Other Characters', 'This Character', 'All'])
        equipped_by_cb.pack(side='right', padx=5)
        tk.Label(equipped_by_frame, text="Equipped by:", font=('Segoe UI', 12, 'bold'),
                 fg=FG_DIM, bg=BG_DARK).pack(side='right')

        # Relic list with columns
        list_container = tk.Frame(right_panel, bg=BG_DARK)
        list_container.pack(fill='both', expand=True)

        columns = ('Name', 'Color', 'Effect 1', 'Effect 2', 'Effect 3')
        relic_listbox = ttk.Treeview(list_container, columns=columns, show='headings', height=20)
        relic_listbox.heading('Name', text='Relic Name')
        relic_listbox.heading('Color', text='Color')
        relic_listbox.heading('Effect 1', text='Effect 1')
        relic_listbox.heading('Effect 2', text='Effect 2')
        relic_listbox.heading('Effect 3', text='Effect 3')

        relic_listbox.column('Name', width=180)
        relic_listbox.column('Color', width=60)
        relic_listbox.column('Effect 1', width=150)
        relic_listbox.column('Effect 2', width=150)
        relic_listbox.column('Effect 3', width=150)

        scrollbar = ttk.Scrollbar(list_container, orient='vertical', command=relic_listbox.yview)
        relic_listbox.configure(yscrollcommand=scrollbar.set)
        relic_listbox.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # Configure row colors by relic color
        for color, hex_color in RELIC_COLOR_HEX.items():
            if color:  # Skip None key
                relic_listbox.tag_configure(color, foreground=hex_color)

        # Build list of all relics with deep flag from inventory (ga_relic)
        all_relics = []
        with open('preset_debug.txt', 'a') as f:
            f.write(f"Building all_relics from ga_relic ({len(ga_relic)} entries)\n")
        for relic in ga_relic:
            if len(relic) >= 2:  # Just need ga_handle and item_id at minimum
                ga_handle = relic[0]
                item_id = relic[1]
                real_id = item_id - 2147483648 if item_id > 2147483648 else item_id
                item_info = items_json.get(str(real_id), {})
                # Deep relic check: ID range 2000000-2019999
                is_deep_relic = 2000000 <= real_id <= 2019999
                # Get effects/curses if available
                effects = [relic[2], relic[3], relic[4]] if len(relic) >= 5 else [0, 0, 0]
                curses = [relic[5], relic[6], relic[7]] if len(relic) >= 8 else [0, 0, 0]
                all_relics.append({
                    'ga_handle': ga_handle,
                    'name': item_info.get('name', f'ID:{real_id}'),
                    'color': item_info.get('color', 'Unknown'),
                    'effects': effects,
                    'curses': curses,
                    'real_id': real_id,
                    'is_deep': is_deep_relic
                })
        with open('preset_debug.txt', 'a') as f:
            f.write(f"Built all_relics with {len(all_relics)} entries\n")
            # Sample first 5 relics
            for i, r in enumerate(all_relics[:5]):
                f.write(f"  Sample {i}: {r['name']}, color={r['color']}, is_deep={r['is_deep']}\n")

        # Store relic data by treeview item ID
        item_to_relic = {}

        # Get vessel slot colors for proper filtering
        vessel_colors = vessel_info.get('colors', None)

        def populate_relic_list():
            """Populate the relic list based on selected slot and filters"""
            # Clear existing
            for item in relic_listbox.get_children():
                relic_listbox.delete(item)
            item_to_relic.clear()

            idx = selected_slot.get()
            is_deep_slot = idx >= 3

            # Get slot color from vessel data (proper way to handle White/Universal slots)
            slot_color = None
            if vessel_colors and idx < len(vessel_colors):
                slot_color = vessel_colors[idx]

            # For White/Universal slots, show all colors by default
            is_universal_slot = (slot_color == 'White')

            search_text = search_var.get().lower()
            show_all = show_all_colors_var.get() or is_universal_slot
            equipped_by_str = equipped_by_var.get()

            # Debug info - append to file
            with open('preset_debug.txt', 'a') as f:
                f.write(f"populate_relic_list: slot={idx}, is_deep={is_deep_slot}, "
                        f"slot_color={slot_color}, is_universal={is_universal_slot}, show_all={show_all}, all_relics={len(all_relics)}\n")

            # Update title to show filtering info
            slot_type = "Deep" if is_deep_slot else "Normal"
            if is_universal_slot:
                relic_list_title.config(text=f"Available {slot_type} Relics (Universal Slot)")
            elif slot_color and not show_all:
                relic_list_title.config(text=f"Available {slot_type} Relics ({slot_color})")
            else:
                relic_list_title.config(text=f"Available {slot_type} Relics (All Colors)")

            filtered = []
            ga_hero_map = loadout_handler.relic_ga_hero_map
            hero_type =self.vessel_char_combo.current()+1
            for relic in all_relics:
                # Filter by deep/normal based on selected slot
                if is_deep_slot and not relic['is_deep']:
                    continue  # Deep slots need deep relics
                if not is_deep_slot and relic['is_deep']:
                    continue  # Normal slots need normal relics
                
                if relic['ga_handle'] in ga_handles:
                    continue  # Ignore relic in this preset
                # Filter by Equipped state
                match equipped_by_str:
                    case 'None':
                        if len(ga_hero_map.get(relic['ga_handle'], [])) != 0:
                            continue
                    case 'Other Characters':
                        if hero_type in ga_hero_map.get(relic['ga_handle'], []) or not ga_hero_map.get(relic['ga_handle'], []):
                            continue
                    case 'This Character':
                        if hero_type not in ga_hero_map.get(relic['ga_handle'], []) or not ga_hero_map.get(relic['ga_handle'], []):
                            continue
                    case 'All':
                        pass
                    case _:
                        pass

                # Color filter (based on slot color from vessel data, unless "show all" or universal slot)
                if not show_all and slot_color and slot_color != 'White' and relic['color'] != slot_color:
                    continue

                # Search filter
                if search_text:
                    searchable = relic['name'].lower()
                    for eff in relic['effects']:
                        eff_name = get_effect_name(eff)
                        if eff_name:
                            searchable += ' ' + eff_name.lower()
                    if search_text not in searchable:
                        continue

                filtered.append(relic)

            with open('preset_debug.txt', 'a') as f:
                f.write(f"Filtered to {len(filtered)} relics\n")

            # Populate list
            for relic in filtered:
                eff_names = []
                for eff in relic['effects']:
                    name = get_effect_name(eff)
                    # Strip newlines from effect names
                    if name:
                        name = name.replace('\n', ' ').replace('\r', ' ').strip()
                    eff_names.append(name or '-')

                item_id = relic_listbox.insert('', 'end',
                    values=(relic['name'], relic['color'], eff_names[0], eff_names[1], eff_names[2]),
                    tags=(relic['color'],))
                item_to_relic[item_id] = relic

            # Auto-size columns based on content
            autosize_treeview_columns(relic_listbox)

        def on_slot_changed():
            """Called when selected slot changes - refresh relic list"""
            populate_relic_list()

        def assign_selected_relic():
            """Assign the selected relic to the current slot"""
            selection = relic_listbox.selection()
            if not selection:
                return

            relic_data = item_to_relic.get(selection[0])
            if not relic_data:
                return

            idx = selected_slot.get()
            new_ga = relic_data['ga_handle']

            # Update in memory
            ga_handles[idx] = new_ga

            # Write to file
            try:
                loadout_handler.replace_preset_relic(preset['hero_type'], idx, new_ga,
                                                     preset_index=preset['index'])
                slot_relic_data[idx] = relic_data
                update_slot_display()
                update_details_panel()
                populate_relic_list()  # Refresh list (color filter may change)
                self.refresh_presets()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to assign preset relic:\n{e}")
                
            # if self.write_preset_relic(preset_offset, idx, new_ga):
            #     slot_relic_data[idx] = relic_data
            #     update_slot_display()
            #     update_details_panel()
            #     populate_relic_list()  # Refresh list (color filter may change)
            #     self.refresh_presets()

        # Double-click to assign
        relic_listbox.bind('<Double-1>', lambda e: assign_selected_relic())

        # Bind filter events
        search_var.trace_add('write', lambda *args: populate_relic_list())
        show_all_colors_var.trace_add('write', lambda *args: populate_relic_list())
        equipped_by_var.trace_add('write', lambda *args: populate_relic_list())

        # Assign button
        assign_btn_frame = tk.Frame(right_panel, bg=BG_DARK)
        assign_btn_frame.pack(fill='x', pady=(10, 0))
        ttk.Button(assign_btn_frame, text="Assign to Selected Slot",
                   command=assign_selected_relic).pack(side='left')

        # Override select_slot to also refresh relic list
        original_select_slot = select_slot
        def select_slot_with_refresh(idx):
            original_select_slot(idx)
            populate_relic_list()

        # Re-bind slot frames to use the new function
        for idx, frame in enumerate(slot_frames):
            for widget in [frame] + list(frame.winfo_children()):
                widget.bind('<Button-1>', lambda e, i=idx: select_slot_with_refresh(i))
                widget.bind('<Double-Button-1>', lambda e, i=idx: select_slot_with_refresh(i))

        # Initial population
        update_slot_display()
        update_details_panel()
        populate_relic_list()

    def write_preset_relic(self, preset_offset, slot_idx, new_ga_handle):
        """Write a relic GA handle to a preset slot in the save file"""
        if globals.data is None:
            return False

        # Preset structure: [vessel_id (4 bytes)] [6 GA handles (4 bytes each)]
        # GA handles start at offset + 4
        relic_offset = preset_offset['relics'] + (slot_idx * 4)

        try:
            struct.pack_into('<I', globals.data, relic_offset, new_ga_handle)
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Failed to write preset relic: {e}")
            return False

    # def show_preset_relic_replacement_dialog(self, parent_dialog, relic_tree, slot_items,
    #                                          ga_handles, slot_idx, is_deep, current_color,
    #                                          preset_offset, char_name, vessel_slot, preset_name,
    #                                          slot_relic_data=None, update_details_func=None,
    #                                          ga_to_full_info=None):
    #     """Show dialog to select a replacement relic for a preset slot"""
    #     dialog = tk.Toplevel(parent_dialog)
    #     dialog.title(f"Select Relic for Slot {slot_idx + 1}")
    #     dialog.geometry("950x600")
    #     dialog.transient(parent_dialog)
    #     dialog.grab_set()

    #     # Options frame
    #     options_frame = ttk.Frame(dialog)
    #     options_frame.pack(fill='x', padx=10, pady=5)

    #     # Color filter option
    #     any_color_var = tk.BooleanVar(value=False)

    #     def refresh_list():
    #         self.refresh_preset_relic_list(
    #             replacement_tree, details_text, current_color,
    #             any_color_var.get(), is_deep, search_var.get()
    #         )

    #     ttk.Checkbutton(options_frame, text="Show all colors", variable=any_color_var,
    #                     command=refresh_list).pack(side='left', padx=5)

    #     # Search
    #     ttk.Label(options_frame, text="Search:").pack(side='left', padx=5)
    #     search_var = tk.StringVar()
    #     search_entry = ttk.Entry(options_frame, textvariable=search_var, width=30)
    #     search_entry.pack(side='left', padx=5)
    #     search_var.trace('w', lambda *args: refresh_list())

    #     # Main content - split into list and details
    #     content_frame = ttk.Frame(dialog)
    #     content_frame.pack(fill='both', expand=True, padx=10, pady=5)

    #     # Left side - Relic list
    #     list_frame = ttk.LabelFrame(content_frame, text="Available Relics")
    #     list_frame.pack(side='left', fill='both', expand=True, padx=(0, 5))

    #     columns = ('Name', 'Color')
    #     replacement_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=18)
    #     replacement_tree.heading('Name', text='Relic Name')
    #     replacement_tree.heading('Color', text='Color')

    #     replacement_tree.column('Name', width=280)
    #     replacement_tree.column('Color', width=80)

    #     scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=replacement_tree.yview)
    #     replacement_tree.configure(yscrollcommand=scrollbar.set)

    #     replacement_tree.pack(side='left', fill='both', expand=True)
    #     scrollbar.pack(side='right', fill='y')

    #     # Right side - Details panel
    #     details_outer = ttk.LabelFrame(content_frame, text="Relic Details")
    #     details_outer.pack(side='right', fill='both', expand=True, padx=(5, 0))

    #     details_text = tk.Text(details_outer, wrap='word', width=45, height=20, font=('Consolas', 9))
    #     details_text.pack(fill='both', expand=True, padx=5, pady=5)
    #     details_text.config(state='disabled')

    #     # Configure text tags
    #     details_text.tag_configure('title', font=('Segoe UI', 11, 'bold'))
    #     details_text.tag_configure('section', font=('Segoe UI', 10, 'bold'), foreground='#4a90d9')
    #     details_text.tag_configure('curse', foreground='#cc4444')

    #     # Populate initial list
    #     self.refresh_preset_relic_list(replacement_tree, details_text, current_color, False, is_deep, "")

    #     # Button frame
    #     btn_frame = ttk.Frame(dialog)
    #     btn_frame.pack(fill='x', padx=10, pady=10)

    #     def on_select():
    #         selection = replacement_tree.selection()
    #         if not selection:
    #             messagebox.showwarning("No Selection", "Please select a relic.")
    #             return

    #         # Get relic data from our stored map
    #         item_data = replacement_tree.item(selection[0])
    #         new_ga = item_data.get('tags', [None])[0]
    #         if new_ga is None:
    #             return
    #         new_ga = int(new_ga)

    #         values = item_data['values']
    #         new_name = values[0]
    #         new_color = values[1]

    #         # Update in memory
    #         ga_handles[slot_idx] = new_ga

    #         # Write to file
    #         if self.write_preset_relic(preset_offset, slot_idx, new_ga):
    #             # Update the parent tree display
    #             slot_label = f"{'🔮 ' if is_deep else ''}Slot {slot_idx + 1}"
    #             for item_id, idx in slot_items.items():
    #                 if idx == slot_idx:
    #                     relic_tree.item(item_id, values=(slot_label, new_name, new_color))
    #                     break

    #             # Update slot_relic_data if provided
    #             if slot_relic_data is not None and ga_to_full_info is not None:
    #                 slot_relic_data[slot_idx] = ga_to_full_info.get(new_ga)

    #             # Update details in parent dialog
    #             if update_details_func:
    #                 update_details_func()

    #             # Refresh the presets tree in the main window
    #             self.refresh_presets()
    #             dialog.destroy()

    #     ttk.Button(btn_frame, text="Select", command=on_select).pack(side='left', padx=5)
    #     ttk.Button(btn_frame, text="Cancel", command=dialog.destroy).pack(side='right', padx=5)

    #     # Double-click to select
    #     replacement_tree.bind("<Double-1>", lambda e: on_select())

    def refresh_preset_relic_list(self, tree, details_text, current_color, any_color, is_deep, search_text):
        """Refresh the list of available relics for preset replacement"""
        # Clear current items
        for item in tree.get_children():
            tree.delete(item)

        search_lower = search_text.lower() if search_text else ""

        # Store relic data for details lookup
        relic_data_map = {}

        def get_effect_name(eff_id):
            if eff_id in [0, -1, 4294967295]:
                return None
            return effects_json.get(str(eff_id), {}).get('name', f'Unknown ({eff_id})')

        def update_details(event=None):
            """Update details panel when selection changes"""
            selection = tree.selection()
            details_text.config(state='normal')
            details_text.delete('1.0', 'end')

            if not selection:
                details_text.insert('end', "Select a relic to view details")
                details_text.config(state='disabled')
                return

            ga_handle = tree.item(selection[0]).get('tags', [None])[0]
            if ga_handle is None:
                details_text.config(state='disabled')
                return

            relic_data = relic_data_map.get(int(ga_handle))
            if not relic_data:
                details_text.config(state='disabled')
                return

            # Display relic info
            details_text.insert('end', f"📦 {relic_data['name']}\n", 'title')
            details_text.insert('end', f"Color: {relic_data['color']}\n")
            details_text.insert('end', f"ID: {relic_data['real_id']}\n\n")

            # Effects (Blessings)
            details_text.insert('end', "✨ BLESSINGS:\n", 'section')
            effects = relic_data.get('effects', [])
            has_effects = False
            for i, eff in enumerate(effects):
                name = get_effect_name(eff)
                if name:
                    has_effects = True
                    details_text.insert('end', f"  {i+1}. {name}\n")
            if not has_effects:
                details_text.insert('end', "  (None)\n")

            details_text.insert('end', "\n")

            # Curses
            details_text.insert('end', "💀 CURSES:\n", 'section')
            curses = relic_data.get('curses', [])
            has_curses = False
            for i, curse in enumerate(curses):
                name = get_effect_name(curse)
                if name:
                    has_curses = True
                    details_text.insert('end', f"  {i+1}. {name}\n", 'curse')
            if not has_curses:
                details_text.insert('end', "  (None)\n")

            details_text.config(state='disabled')

        # Bind selection event
        tree.bind('<<TreeviewSelect>>', update_details)

        # Build list of available relics from ga_relic
        for relic in ga_relic:
            if len(relic) < 8:
                continue

            ga_handle = relic[0]
            item_id = relic[1]
            real_id = item_id - 2147483648 if item_id > 2147483648 else item_id
            item_info = items_json.get(str(real_id), {})

            name = item_info.get('name', f'ID:{real_id}')
            color = item_info.get('color', 'Unknown')

            # Store full relic data for details panel
            relic_data_map[ga_handle] = {
                'name': name,
                'color': color,
                'real_id': real_id,
                'effects': [relic[2], relic[3], relic[4]],
                'curses': [relic[5], relic[6], relic[7]]
            }

            # Filter by search - also search in effect names
            if search_lower:
                effect_names = []
                for eff in [relic[2], relic[3], relic[4]]:
                    eff_name = get_effect_name(eff)
                    if eff_name:
                        effect_names.append(eff_name.lower())
                searchable = f"{name.lower()} {' '.join(effect_names)}"
                if search_lower not in searchable:
                    continue

            # Filter by color (if not showing all)
            if not any_color and current_color and color != current_color:
                continue

            # Insert with GA handle as tag for lookup
            tree.insert('', 'end', values=(name, color), tags=(str(ga_handle),))

        # Auto-size columns after populating
        autosize_treeview_columns(tree)

    def show_vessel_context_menu(self, event, vessel_slot):
        """Show context menu for vessel relic slot"""
        def clear_relic(slot_index):
            hero_type = self.vessel_char_combo.current()+1
            vessel_id = loadout_handler.heroes[hero_type].vessels[vessel_slot]['vessel_id']
            loadout_handler.replace_vessel_relic(hero_type, vessel_id,
                                                 slot_index, 0)
            self.refresh_inventory_and_vessels()

        # Get vessel data
        tree = self.vessel_trees[vessel_slot]

        # Identify which row was clicked
        item = tree.identify_row(event.y)
        if not item:
            return

        # Select the clicked row
        tree.selection_set(item)

        # Get the slot index (1-based from display, convert to 0-based)
        values = tree.item(item, 'values')
        if not values:
            return

        slot_index = int(values[0]) - 1  # Convert from 1-based to 0-based
        relic_color = values[2]  # Color column
        relic_name = values[3]   # Name column

        # Check if slot has a relic (not empty)
        has_relic = relic_name != "(Empty)" and relic_name != "-"

        # Create context menu
        menu = tk.Menu(tree, tearoff=0)

        if has_relic:
            menu.add_command(
                label=f"Replace Relic ({relic_color})",
                command=lambda: self.open_replace_relic_dialog(vessel_slot, slot_index)
            )
            menu.add_command(
                label="📋 Copy Relic Effects",
                command=lambda: self.copy_vessel_relic_effects(vessel_slot, slot_index)
            )
            menu.add_command(
                label="Edit Relic",
                command=lambda: self.open_edit_relic_dialog(vessel_slot, slot_index)
            )
            menu.add_command(
                label="Clear Relic",
                command=lambda: clear_relic(slot_index)
            )
        else:
            # Empty slot - allow assigning a relic (we know slot color from vessel data)
            menu.add_command(
                label=f"Assign Relic ({relic_color})",
                command=lambda: self.open_replace_relic_dialog(vessel_slot, slot_index)
            )

        # Show the menu
        menu.tk_popup(event.x_root, event.y_root)

    def on_vessel_relic_double_click(self, event, vessel_slot):
        """Handle double-click on vessel relic - open replace/assign dialog"""
        tree = self.vessel_trees[vessel_slot]

        # Identify which row was clicked
        item = tree.identify_row(event.y)
        if not item:
            return

        # Get the slot index
        values = tree.item(item, 'values')
        if not values:
            return

        slot_index = int(values[0]) - 1
        # Allow double-click on any slot (empty or not) to open the dialog
        self.open_replace_relic_dialog(vessel_slot, slot_index)

    def open_replace_relic_dialog(self, vessel_slot, slot_index):
        """Open dialog to assign/replace a relic in a vessel slot"""
        char_name = self.vessel_char_var.get()
        loadout = get_character_loadout(char_name)

        vessel_info = loadout.get(vessel_slot, {})
        relics = vessel_info.get('relics', [])

        # Extend relics list if needed
        while len(relics) <= slot_index:
            relics.append((0, None))

        current_ga, current_relic = relics[slot_index]
        is_empty_slot = current_relic is None

        # Get slot color from vessel data (what color should go in this slot)
        # This properly handles White/Universal slots
        vessel_data_info = get_vessel_info(char_name, vessel_slot)
        vessel_colors = vessel_data_info.get('colors', None)
        slot_color = None
        if vessel_colors and slot_index < len(vessel_colors):
            slot_color = vessel_colors[slot_index]

        # Use slot color from vessel data, or fall back to equipped relic's color
        if slot_color:
            current_color = slot_color
        elif current_relic:
            current_color = current_relic.get('color', 'Unknown')
        else:
            current_color = 'Unknown'
        is_deep_slot = slot_index >= 3

        # Create dialog
        dialog = tk.Toplevel(self.root)
        dialog_title = "Assign Relic" if is_empty_slot else "Replace Relic"
        dialog.title(f"{dialog_title} - Slot {slot_index + 1}")
        dialog.geometry("700x500")
        dialog.minsize(500, 350)  # Set minimum size so buttons are always visible
        dialog.transient(self.root)
        dialog.grab_set()

        # Current relic info (or empty slot info)
        info_frame = ttk.LabelFrame(dialog, text="Current Slot" if is_empty_slot else "Current Relic")
        info_frame.pack(fill='x', padx=10, pady=5)

        info_left = ttk.Frame(info_frame)
        info_left.pack(side='left', fill='x', expand=True)

        if is_empty_slot:
            ttk.Label(info_left, text="(Empty Slot)").pack(anchor='w', padx=5)
            ttk.Label(info_left, text=f"Slot Color: {current_color}").pack(anchor='w', padx=5)
            ttk.Label(info_left, text=f"Slot Type: {'Deep' if is_deep_slot else 'Normal'}").pack(anchor='w', padx=5)
        else:
            ttk.Label(info_left, text=f"Name: {current_relic['name']}").pack(anchor='w', padx=5)
            ttk.Label(info_left, text=f"Color: {current_color}").pack(anchor='w', padx=5)
            ttk.Label(info_left, text=f"ID: {current_relic['real_id']}").pack(anchor='w', padx=5)
            ttk.Label(info_left, text=f"Slot Type: {'Deep' if is_deep_slot else 'Normal'}").pack(anchor='w', padx=5)

        # Edit button for current relic (only if slot has a relic)
        def edit_current_relic():
            if is_empty_slot:
                return
            # Create refresh callback that updates vessels
            def refresh_after_edit():
                self.refresh_inventory_and_vessels()
                # Refresh the dialog info if still open
                if dialog.winfo_exists():
                    # Reload current relic info
                    new_loadout = get_character_loadout(char_name)
                    new_vessel_info = new_loadout.get(vessel_slot, {})
                    new_relics = new_vessel_info.get('relics', [])
                    if slot_index < len(new_relics):
                        new_ga, new_relic = new_relics[slot_index]
                        if new_relic:
                            # Update labels
                            for widget in info_left.winfo_children():
                                widget.destroy()
                            ttk.Label(info_left, text=f"Name: {new_relic['name']}").pack(anchor='w', padx=5)
                            ttk.Label(info_left, text=f"Color: {new_relic.get('color', 'Unknown')}").pack(anchor='w', padx=5)
                            ttk.Label(info_left, text=f"ID: {new_relic['real_id']}").pack(anchor='w', padx=5)
                            ttk.Label(info_left, text=f"Slot Type: {'Deep' if is_deep_slot else 'Normal'}").pack(anchor='w', padx=5)

            def on_edit_dialog_close():
                # Restore grab when edit dialog closes
                if dialog.winfo_exists():
                    dialog.grab_set()

            # Release grab so edit dialog can receive input
            dialog.grab_release()

            # Open or reuse modify dialog
            if not self.modify_dialog or not self.modify_dialog.dialog.winfo_exists():
                self.modify_dialog = ModifyRelicDialog(self.root, current_ga, current_relic['real_id'], refresh_after_edit)
            else:
                self.modify_dialog.load_relic(current_ga, current_relic['real_id'])
                self.modify_dialog.callback = refresh_after_edit

            # Set up close protocol and bring to front
            self.modify_dialog.dialog.protocol("WM_DELETE_WINDOW", lambda: [self.modify_dialog.dialog.destroy(), on_edit_dialog_close()])
            self.modify_dialog.dialog.lift()
            self.modify_dialog.dialog.focus_force()

        info_right = ttk.Frame(info_frame)
        info_right.pack(side='right', padx=10, pady=5)
        edit_btn = ttk.Button(info_right, text="✏️ Edit Effects", command=edit_current_relic)
        edit_btn.pack()
        if is_empty_slot:
            edit_btn.config(state='disabled')

        # Options frame
        options_frame = ttk.Frame(dialog)
        options_frame.pack(fill='x', padx=10, pady=5)

        # For White/Universal slots, default to showing all colors
        is_universal_slot = (slot_color == 'White')
        any_color_var = tk.BooleanVar(value=is_universal_slot)
        search_var = tk.StringVar()
        equipped_by_var = tk.StringVar()

        # Checkbox to allow any color (pre-checked for White slots)
        any_color_cb = ttk.Checkbutton(
            options_frame,
            text="Show all colors" if is_universal_slot else "Show all colors (slot accepts any)",
            variable=any_color_var,
            command=lambda: self.refresh_replacement_list(relic_tree, current_color, any_color_var.get(), equipped_by_var.get(), is_deep_slot, search_var.get(), relics)
        )
        any_color_cb.pack(side='left', padx=5)

        # Add indicator if it's a universal slot
        if is_universal_slot:
            ttk.Label(options_frame, text="(Universal slot)", foreground='gray').pack(side='left', padx=5)
        
        # Equipped Option
        ttk.Label(options_frame, text="Equipped by").pack(side='left', padx=(20, 5))
        equipped_by_cb = ttk.Combobox(options_frame, values=['None', 'Other Characters', 'This Character', 'All'],
                                      state='readonly', width=20, textvariable=equipped_by_var)
        equipped_by_cb.pack(side='left', padx=5)
        equipped_by_cb.current(0)
        equipped_by_var.trace('w', lambda *args: self.refresh_replacement_list(
            relic_tree, current_color, any_color_var.get(), equipped_by_var.get(), is_deep_slot, search_var.get(), relics))
        
        # Search box
        options_frame_2 = ttk.Frame(dialog)
        options_frame_2.pack(fill='x', padx=10, pady=5)
        
        ttk.Label(options_frame_2, text="Search:").pack(side='left', padx=5)
        search_entry = ttk.Entry(options_frame_2, textvariable=search_var, width=30)
        search_entry.pack(side='left', padx=5, fill='x', expand=True)
        search_var.trace('w', lambda *args: self.refresh_replacement_list(
            relic_tree, current_color, any_color_var.get(), equipped_by_var.get(), is_deep_slot, search_var.get(), relics))
        
        # Buttons frame - pack FIRST with side='bottom' so it's always visible
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(side='bottom', fill='x', padx=10, pady=10)

        # Relic list - for universal slots, start with "All Colors"
        list_label = "All Colors" if is_universal_slot else current_color
        list_frame = ttk.LabelFrame(dialog, text=f"Available Relics ({list_label})")
        list_frame.pack(fill='both', expand=True, padx=10, pady=5)

        columns = ('Name', 'Color', 'ID', 'Effect 1', 'Effect 2', 'Effect 3')
        relic_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=15)

        for col in columns:
            relic_tree.heading(col, text=col)
        relic_tree.column('Name', width=180)
        relic_tree.column('Color', width=60)
        relic_tree.column('ID', width=70)
        relic_tree.column('Effect 1', width=120)
        relic_tree.column('Effect 2', width=120)
        relic_tree.column('Effect 3', width=120)

        vsb = ttk.Scrollbar(list_frame, orient="vertical", command=relic_tree.yview)
        relic_tree.configure(yscrollcommand=vsb.set)

        relic_tree.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')

        # Configure color tags
        for color_name, color_hex in RELIC_COLOR_HEX.items():
            if color_name:
                relic_tree.tag_configure(color_name, foreground=color_hex)

        # Store reference to update label
        self._replace_list_frame = list_frame

        # Populate initial list - use is_universal_slot to show all colors for white slots
        self.refresh_replacement_list(relic_tree, current_color, is_universal_slot, equipped_by_var.get(), is_deep_slot, "", relics)

        def do_replace():
            selection = relic_tree.selection()
            if not selection:
                messagebox.showwarning("Warning", "Please select a relic to replace with")
                return

            # Get selected relic GA
            item = selection[0]
            new_ga = int(relic_tree.item(item, 'text'))

            # Perform replacement
            success = self.replace_vessel_relic(char_name, vessel_slot, slot_index, new_ga)
            if success:
                dialog.destroy()
                self.refresh_vessels()
                messagebox.showinfo("Success", "Relic replaced successfully!")

        ttk.Button(btn_frame, text="Replace", command=do_replace).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Cancel", command=dialog.destroy).pack(side='left', padx=5)

    def refresh_replacement_list(self, tree, current_color, allow_any_color, equipped_by, is_deep_slot, search_term="", vessel_relics=None):
        """Refresh the replacement relic list based on filters"""
        if vessel_relics is None:
            vessel_relics = []

        # Clear existing items
        for item in tree.get_children():
            tree.delete(item)

        # Update frame label
        if hasattr(self, '_replace_list_frame'):
            color_text = "Any Color" if allow_any_color else current_color
            self._replace_list_frame.config(text=f"Available Relics ({color_text})")

        search_lower = search_term.lower()

        # Get relics from inventory that match criteria
        ga_hero_map = loadout_handler.relic_ga_hero_map
        for relic in ga_relic:
            ga = relic[0]
            real_id = relic[1] - 2147483648
            effects = [relic[2], relic[3], relic[4]]

            # Get item info
            item_data = items_json.get(str(real_id), {})
            item_name = item_data.get('name', f'Unknown ({real_id})')
            item_color = item_data.get('color', 'Unknown')

            # Filter by existing relics currently in the vessel
            if ga in [i[0] for i in vessel_relics]:
                continue

            # Filter by color unless "any color" is checked
            if not allow_any_color and item_color != current_color:
                continue

            # Filter by deep relic status if this is a deep slot
            relic_is_deep = 2000000 <= real_id <= 2019999
            if is_deep_slot and not relic_is_deep:
                continue  # Deep slots need deep relics
            if not is_deep_slot and relic_is_deep:
                continue  # Normal slots need normal relics
            hero_type = self.vessel_char_combo.current() + 1
            # Filter by Equipped state
            match equipped_by:
                case 'None':
                    if len(ga_hero_map.get(ga, [])) != 0:
                        continue
                case 'Other Characters':
                    if hero_type in ga_hero_map.get(ga, []) or not ga_hero_map.get(ga, []):
                        continue
                case 'This Character':
                    if hero_type not in ga_hero_map.get(ga, []) or not ga_hero_map.get(ga, []):
                        continue
                case 'All':
                    pass
                case _:
                    pass

            # Search filter
            if search_lower:
                # Search in name, ID, or effects
                effect_names = []
                for eff in effects:
                    if str(eff) in effects_json:
                        effect_names.append(effects_json[str(eff)].get('name', '').lower())

                searchable = f"{item_name.lower()} {real_id} {' '.join(effect_names)}"
                if search_lower not in searchable:
                    continue

            # Get effect names
            effect_displays = []
            for eff in effects:
                if eff == 0:
                    effect_displays.append("None")
                elif eff == 4294967295:
                    effect_displays.append("Empty")
                elif str(eff) in effects_json:
                    effect_displays.append(effects_json[str(eff)].get('name', 'Unknown')[:20])
                else:
                    effect_displays.append("Unknown")

            tree.insert('', 'end', text=str(ga), values=(
                item_name[:30],
                item_color,
                real_id,
                effect_displays[0],
                effect_displays[1],
                effect_displays[2]
            ), tags=(item_color,))

        # Auto-size columns after populating
        autosize_treeview_columns(tree)

    def replace_vessel_relic(self, char_name, vessel_slot, slot_index, new_ga):
        """Replace a relic in a vessel slot with a new one"""

        char_id = data_source.character_names.index(char_name) if char_name in data_source.character_names else -1
        hero_type = char_id + 1
        if char_id < 0:
            messagebox.showerror("Error", f"Unknown character: {char_name}")
            return False
        vessel_id = loadout_handler.get_vessel_id(hero_type, vessel_slot)
        try:
            loadout_handler.replace_vessel_relic(hero_type, vessel_id, slot_index, new_ga)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to replace relic: {e}")
            return False

        return True

    def open_edit_relic_dialog(self, vessel_slot, slot_index):
        """Open dialog to edit a relic's effects from the vessel page"""
        char_name = self.vessel_char_var.get()
        char_id = data_source.character_names.index(char_name) if char_name in data_source.character_names else -1
        hero_type = char_id + 1
        if char_id < 0:
            messagebox.showerror("Error", f"Unknown character: {char_name}")
            return

        # Get the GA handle for this slot
        vessel_id = loadout_handler.get_vessel_id(hero_type, vessel_slot)
        ga_handle = loadout_handler.get_relic_ga_handle(hero_type, vessel_id, slot_index)
        if ga_handle == 0:
            messagebox.showwarning("Warning", "Empty slot - no relic to edit")
            return

        # Find the relic in ga_relic to get its item_id
        for ga, id_val, e1, e2, e3, se1, se2, se3, offset, size in ga_relic:
            if ga == ga_handle:
                real_id = id_val - 2147483648

                # Open or update the modify dialog
                if not self.modify_dialog or not self.modify_dialog.dialog.winfo_exists():
                    self.modify_dialog = ModifyRelicDialog(self.root, ga_handle, real_id, self.refresh_inventory_and_vessels)
                else:
                    self.modify_dialog.load_relic(ga_handle, real_id)
                    self.modify_dialog.dialog.lift()
                return

        messagebox.showerror("Error", "Could not find relic data")

    def copy_vessel_relic_effects(self, vessel_slot, slot_index):
        """Copy effects from a relic in a vessel slot to clipboard"""
        char_name = self.vessel_char_var.get()
        char_id = data_source.character_names.index(char_name) if char_name in data_source.character_names else -1
        hero_type = char_id + 1
        if char_id < 0:
            messagebox.showerror("Error", f"Unknown character: {char_name}")
            return

        # Get the GA handle for this slot
        vessel_id = loadout_handler.get_vessel_id(hero_type, vessel_slot)
        ga_handle = loadout_handler.get_relic_ga_handle(hero_type, vessel_id, slot_index)
        if ga_handle == 0:
            messagebox.showwarning("Warning", "Empty slot - no relic to copy")
            return

        # Find the relic in ga_relic to get its effects
        for ga, id_val, e1, e2, e3, se1, se2, se3, offset, size in ga_relic:
            if ga == ga_handle:
                real_id = id_val - 2147483648
                effects = [e1, e2, e3, se1, se2, se3]
                item_name = items_json.get(str(real_id), {}).get("name", f"Unknown ({real_id})")
                self.clipboard_effects = (effects, real_id, item_name)

                effect_count = len([e for e in effects if e != 0])
                messagebox.showinfo("Copied", f"Copied effects from:\n{item_name}\n\nEffects: {effect_count} effect(s)")
                return

        messagebox.showerror("Error", "Could not find relic data")

    def save_loadout(self):
        """Save the current character's loadout to a JSON file"""
        if globals.data is None:
            messagebox.showwarning("Warning", "No character loaded")
            return

        char_name = self.vessel_char_var.get()
        loadout = get_character_loadout(char_name)

        # Convert to serializable format
        save_data = {
            'character': char_name,
            'vessels': {}
        }

        for vessel_slot, vessel_info in loadout.items():
            relics = vessel_info.get('relics', [])
            vessel_relics = []
            for ga, relic_info in relics:
                if relic_info:
                    vessel_relics.append({
                        'ga': ga,
                        'real_id': relic_info['real_id'],
                        'name': relic_info['name'],
                        'color': relic_info.get('color', 'Unknown'),
                        'effects': relic_info['effects'],
                        'curses': relic_info['curses']
                    })
                else:
                    vessel_relics.append(None)
            save_data['vessels'][str(vessel_slot)] = {
                'unlocked': vessel_info.get('unlocked', False),
                'relics': vessel_relics
            }

        # Ask for save location
        file_path = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json")],
            initialfile=f"{char_name}_loadout.json",
            title="Save Loadout"
        )

        if file_path:
            try:
                with open(file_path, 'w') as f:
                    json.dump(save_data, f, indent=2)
                messagebox.showinfo("Success", f"Loadout saved to {file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save loadout: {e}")

    def load_loadout(self):
        """Load a loadout from a JSON file and apply it"""
        if globals.data is None:
            messagebox.showwarning("Warning", "No character loaded")
            return

        file_path = filedialog.askopenfilename(
            filetypes=[("JSON files", "*.json")],
            title="Load Loadout"
        )

        if not file_path:
            return

        try:
            with open(file_path, 'r') as f:
                loadout_data = json.load(f)

            # Show what will be loaded
            char_name = loadout_data.get('character', 'Unknown')
            vessels = loadout_data.get('vessels', {})

            # Build summary
            summary_lines = [f"Loadout from: {char_name}", ""]
            for vessel_slot, relics in vessels.items():
                relic_names = [r['name'] if r else "(Empty)" for r in relics]
                summary_lines.append(f"Vessel {vessel_slot}: {', '.join(relic_names[:3])}...")

            summary = "\n".join(summary_lines[:10])

            messagebox.showinfo("Loadout Preview",
                f"{summary}\n\nNote: Applying loadouts will be available in a future update.\n"
                "Currently you can view and save loadouts.")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load loadout: {e}")

    def setup_inventory_tab(self):
        # Controls frame
        controls_frame = ttk.Frame(self.inventory_tab)
        controls_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Button(controls_frame, text="➕ Add Relic", style='Add.TButton',
                  command=self.add_relic_tk)  # .pack(side="left", padx=5)
        ttk.Button(controls_frame, text="🔄 Refresh Inventory", command=self.reload_inventory).pack(side='left', padx=5)
        ttk.Button(controls_frame, text="📤 Export to Excel", command=self.export_relics).pack(side='left', padx=5)
        ttk.Button(controls_frame, text="📥 Import from Excel", command=self.import_relics).pack(side='left', padx=5)
        ttk.Button(controls_frame, text="🗑️ Delete All Illegal", command=self.delete_all_illegal,
                  style='Danger.TButton')  # .pack(side='left', padx=5)
        ttk.Button(controls_frame, text="🗑️ Mass Delete Selected", command=self.mass_delete_relics,
                  style='Danger.TButton')  # .pack(side='left', padx=5)
        ttk.Button(controls_frame, text="🔧 Mass Fix", command=self.mass_fix_incorrect_ids).pack(side='left', padx=5)

        # ===========Language Combobox========================
        lang_frame = ttk.Frame(self.inventory_tab)
        lang_frame.pack(fill='x', padx=10, pady=5)

        ttk.Label(lang_frame, text="Language:").pack(side='left', padx=2)
        from source_data_handler import LANGUAGE_MAP
        lang_display_names = list(LANGUAGE_MAP.values())

        self.lang_combobox = ttk.Combobox(lang_frame,
                                          values=lang_display_names,
                                          state="readonly",
                                          width=15)
        self.lang_combobox.set(LANGUAGE_MAP.get(get_system_language()))
        self.lang_combobox.pack(side='left', padx=2)
        self.lang_combobox.bind("<<ComboboxSelected>>",
                                self.on_language_change)
        # ====================================================

        legend_frame = ttk.Frame(self.inventory_tab)
        legend_frame.pack(padx=10, fill='x')
        
        # Info label
        self.illegal_count_label = ttk.Label(
            legend_frame,
            text="",
            foreground='red',
            font=('Arial', 9, 'bold')
        )
        self.illegal_count_label.pack(side='left', padx=(0, 15))

        ttk.Label(legend_frame, text="Blue = Red + Orange", foreground="blue").pack(side='left', padx=5)
        ttk.Label(legend_frame, text="Red = Illegal", foreground="red").pack(side='left', padx=5)
        ttk.Label(legend_frame, text="Purple = Missing Curse", foreground="#800080").pack(side='left', padx=5)
        ttk.Label(legend_frame, text="Orange = Unique Relic (don't edit)", foreground="#FF8C00").pack(side='left', padx=5)
        ttk.Label(legend_frame, text="Teal = Strict Invalid", foreground="#008080").pack(side='left', padx=5)

        
        # Search frame - Row 1: Basic search and filters
        search_frame = ttk.Frame(self.inventory_tab)
        search_frame.pack(fill='x', padx=10, pady=5)

        ttk.Label(search_frame, text="🔍 Search:").pack(side='left', padx=5)
        self.search_var = tk.StringVar()
        self.search_var.trace('w', lambda *args: self.filter_relics())

        self.search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=25)
        self.search_entry.pack(side='left', padx=5)

        # Search field selector
        ttk.Label(search_frame, text="in:").pack(side='left', padx=2)
        self.search_field_var = tk.StringVar(value="All Fields")
        search_fields = ["All Fields", "Name", "ID", "Color", "Effect Name", "Effect ID", "Equipped By"]
        self.search_field_combo = ttk.Combobox(search_frame, textvariable=self.search_field_var,
                                                values=search_fields, state="readonly", width=12)
        self.search_field_combo.pack(side='left', padx=2)
        self.search_field_combo.bind("<<ComboboxSelected>>", lambda e: self.filter_relics())

        ttk.Button(search_frame, text="Clear", command=self.clear_search).pack(side='left', padx=5)

        # Character filter dropdown
        ttk.Label(search_frame, text="👤 Char:").pack(side='left', padx=(10, 2))
        self.char_filter_var = tk.StringVar(value="All")
        char_options = ["All"] + data_source.character_names
        self.char_filter_combo = ttk.Combobox(search_frame, textvariable=self.char_filter_var,
                                               values=char_options, state="readonly", width=10)
        self.char_filter_combo.pack(side='left', padx=2)
        self.char_filter_combo.bind("<<ComboboxSelected>>", lambda e: self.filter_relics())

        # Color filter
        ttk.Label(search_frame, text="🎨 Color:").pack(side='left', padx=(10, 2))
        self.color_filter_var = tk.StringVar(value="All")
        color_options = ["All", "Red", "Blue", "Yellow", "Green"]
        self.color_filter_combo = ttk.Combobox(search_frame, textvariable=self.color_filter_var,
                                                values=color_options, state="readonly", width=8)
        self.color_filter_combo.pack(side='left', padx=2)
        self.color_filter_combo.bind("<<ComboboxSelected>>", lambda e: self.filter_relics())

        # Status filter
        ttk.Label(search_frame, text="⚠️ Status:").pack(side='left', padx=(10, 2))
        self.status_filter_var = tk.StringVar(value="All")
        status_options = ["All", "Valid", "Illegal", "Curse Illegal", "Forbidden", "Strict Invalid", "Deep Only"]
        self.status_filter_combo = ttk.Combobox(search_frame, textvariable=self.status_filter_var,
                                                 values=status_options, state="readonly", width=12)
        self.status_filter_combo.pack(side='left', padx=2)
        self.status_filter_combo.bind("<<ComboboxSelected>>", lambda e: self.filter_relics())

        self.search_info_label = ttk.Label(search_frame, text="", foreground='gray')
        self.search_info_label.pack(side='left', padx=10)
        
        # Inventory display
        inv_frame = ttk.Frame(self.inventory_tab)
        inv_frame.pack(fill='both', expand=True, padx=10, pady=5)

        # Treeview for relics (with Equipped By and Deep columns)
        # Note: "Curse" columns show curse effects for deep relics
        # "#" column shows acquisition order (lower = older)
        columns = ('Item Name', 'Deep', 'Item ID', 'Color', 'Equipped By', 'Effect 1', 'Effect 2', 'Effect 3',
                   'Curse 1', 'Curse 2', 'Curse 3')

        self.tree = ttk.Treeview(inv_frame, columns=columns, show='tree headings', height=20)

        # Configure columns - # column shows acquisition order
        self.tree.heading('#0', text='#')
        self.tree.column('#0', width=50, minwidth=50, stretch=False)

        # Set column widths - more space for effect names
        col_widths = {
            'Item Name': 180,
            'Deep': 40,
            'Item ID': 80,
            'Color': 70,
            'Equipped By': 120,
            'Effect 1': 200,
            'Effect 2': 200,
            'Effect 3': 200,
            'Curse 1': 180,
            'Curse 2': 180,
            'Curse 3': 180
        }

        for col in columns:
            self.tree.heading(col, text=col, command=lambda c=col: self.sort_by_column(c))
            self.tree.column(col, width=col_widths.get(col, 150), minwidth=80)

        # Also allow sorting by # column (sorts by acquisition order)
        self.tree.heading('#0', text='#', command=lambda: self.sort_by_column('#'))

        # Track sort state
        self.sort_column = None
        self.sort_reverse = False

        # Scrollbars
        vsb = ttk.Scrollbar(inv_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(inv_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # Grid layout
        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        # Configure grid weights so scrollbars stay visible
        inv_frame.grid_rowconfigure(0, weight=1)
        inv_frame.grid_columnconfigure(0, weight=1)
        
        # Bind selection change
        self.tree.bind('<<TreeviewSelect>>', self.on_relic_select)

        # Bind double-click to open modify window
        self.tree.bind('<Double-1>', lambda e: self.modify_selected_relic())

        # Configure tree for extended selection (multiple items)
        self.tree.configure(selectmode='extended')

        # Context menu
        self.tree.bind("<Button-3>", self.show_context_menu)
        
        # Action buttons
        action_frame = ttk.Frame(self.inventory_tab)
        action_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Button(action_frame, text="Modify Selected", command=self.modify_selected_relic).pack(side='left', padx=5)
        ttk.Button(action_frame, text="Delete Selected", style='Danger.TButton',
                   command=self.delete_selected_relic)  # .pack(side='left', padx=5)
        
        # Selection controls
        selection_frame = ttk.Frame(action_frame)
        selection_frame.pack(side='left', padx=20)
        
        ttk.Button(selection_frame, text="Select All", command=self.select_all_relics).pack(side='left', padx=2)
        ttk.Button(selection_frame, text="Deselect All", command=self.deselect_all_relics).pack(side='left', padx=2)
        selection_count_text = tk.StringVar()
        selection_count_text.set("0 selected")
        self.selection_count_label = ttk.Label(selection_frame, textvariable=selection_count_text, foreground='blue', font=('Arial', 9, 'bold'))
        self.selection_count_label.pack(side='left', padx=10)
        def update_selection_count():
            selected_count = len(self.tree.selection())
            selection_count_text.set(f"{selected_count} selected")
        self.tree.bind('<<TreeviewSelect>>', lambda e: update_selection_count())
    
    
    def on_relic_select(self, event):
        """When a relic is selected and modify dialog is open, update the dialog"""
        if self.modify_dialog and self.modify_dialog.dialog.winfo_exists():
            selection = self.tree.selection()
            if selection:
                item = selection[0]
                tags = self.tree.item(item, 'tags')
                ga_handle = int(tags[0])
                item_id = int(tags[1])
                self.modify_dialog.load_relic(ga_handle, item_id)

    def update_vessel_tab_comboboxes(self):
        # Reload Vessel Character ComboBox Names
        vessel_char_combobox_idx = self.vessel_char_combo.current()
        self.vessel_char_combo['values'] = data_source.character_names
        self.vessel_char_var.set(data_source.character_names[vessel_char_combobox_idx])
        
    def update_inventory_comboboxes(self):
        # Reload Inventory Character Filter ComboBox Names
        filter_combo_idx = self.char_filter_combo.current()
        self.char_filter_combo['values'] = ["All"] + data_source.character_names
        self.char_filter_var.set(self.char_filter_combo['values'][filter_combo_idx])

    def on_language_change(self, event=None):
        selected_name = self.lang_combobox.get()
        from source_data_handler import LANGUAGE_MAP
        lang_code = next((code for code, name in LANGUAGE_MAP.items() if name == selected_name), "en_US")
        global data_source, items_json, effects_json
        if reload_language(lang_code):
            self.refresh_inventory_and_vessels()
        else:
            messagebox.showerror("Error", "Can't change language.")

    def open_file(self):
        global MODE, userdata_path
        
        file_path = filedialog.askopenfilename(
            title="Select Save File",
        )
        
        if not file_path:
            return
        
        file_name = os.path.basename(file_path)

        # Determine mode based on file content, not just filename
        # This allows custom save file names (e.g., from ModEngine 3 Manager)
        if file_name.lower() == 'memory.dat':
            MODE = 'PS4'
        elif file_path.lower().endswith('.sl2'):
            # Check if it's a valid SL2 file by looking for BND4 header
            try:
                with open(file_path, 'rb') as f:
                    header = f.read(4)
                if header == b'BND4':
                    MODE = 'PC'
                else:
                    messagebox.showerror("Error", "This .sl2 file does not have a valid BND4 header. It may be corrupted or not a valid Nightreign save file.")
                    return
            except Exception as e:
                messagebox.showerror("Error", f"Could not read file: {e}")
                return
        else:
            messagebox.showerror("Error", "Please select a valid save file:\n\n• PC: .sl2 file (e.g., NR0000.sl2 or custom named .sl2)\n• PS4: decrypted memory.dat file")
            return
        
        # Split files
        split_files(file_path, 'decrypted_output')

        # Load JSON data
        if not load_json_data():
            return
        self.update_inventory_comboboxes()
        self.update_vessel_tab_comboboxes()

        # Get character names
        name_to_path()

        # Save the opened file path to config
        config = load_config()
        config['last_file'] = file_path
        config['last_mode'] = MODE
        # Reset character index since we're opening a new file
        config['last_char_index'] = 0
        save_config(config)

        # Display character buttons
        self.display_character_buttons()
        
        
    def display_character_buttons(self):
        # Clear existing buttons
        for widget in self.char_button_frame.winfo_children():
            widget.destroy()
        
        # Create styles
        style = ttk.Style()
        style.configure("Char.TButton", font=('Arial', 10), padding=5)
        style.configure("Highlighted.TButton", font=('Arial', 10), padding=5, background="#AF4C4C", foreground="red")
        
        self.char_buttons = []
        
        columns = 4  # Number of buttons per row
        for idx, (name, path) in enumerate(char_name_list):
            row = idx // columns
            col = idx % columns
            
            # Button
            btn = ttk.Button(
                self.char_button_frame,
                text=f"{idx+1}. {name}",
                style="Char.TButton",
                command=lambda b_idx=idx, p=path, n=name: self.on_character_click(b_idx, p, n),
                width=20
            )
            btn.grid(row=row, column=col, padx=5, pady=5, sticky="nsew")
            self.char_buttons.append(btn)
        
        # Make columns expand evenly
        for col in range(columns):
            self.char_button_frame.grid_columnconfigure(col, weight=1)

    def on_character_click(self, idx, path, name):
        # Reset all buttons to normal style
        for b in self.char_buttons:
            b.configure(style="Char.TButton")

        # Highlight clicked button
        self.char_buttons[idx].configure(style="Highlighted.TButton")

        # Save selected character index to config
        self.last_char_index = idx
        config = load_config()
        config['last_char_index'] = idx
        save_config(config)

        # Load character
        self.load_character(path)

    def load_character(self, path):
        global userdata_path, steam_id, data_source, ga_relic, relic_checker, loadout_handler
        userdata_path = path

        try:
            with open(path, "rb") as f:
                globals.data = bytearray(f.read())  # Use bytearray for in-place modifications

            # Parse items
            gaprint(globals.data)

            # Parse Vessels and Presets
            loadout_handler = LoadoutHandler(data_source, ga_relic)
            loadout_handler.parse()

            # Initialize Relic Checker (set_illegal_relics will be called by reload_inventory)
            relic_checker = RelicChecker(ga_relic=ga_relic,
                                         data_source=data_source)
            # NOTE: Don't call set_illegal_relics() here - reload_inventory() will call it

            # Read stats
            read_murks_and_sigs(globals.data)

            steam_id = find_steam_id(globals.data)

            # Refresh all tabs (reload_inventory calls set_illegal_relics)
            self.reload_inventory()
            self.refresh_stats()
            self.refresh_vessels()

        except struct.error as e:
            messagebox.showerror("Error",
                f"Failed to load character: Data structure error\n\n"
                f"Details: {str(e)}\n\n"
                f"This may indicate:\n"
                f"• A corrupted save file\n"
                f"• Save file from a different game version\n"
                f"• Incompatible file format\n\n"
                f"Try deleting some relics in-game and saving again.")
        except IndexError as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Error",
                f"Failed to load character: Index error\n\n"
                f"Details: {str(e)}\n\n"
                f"This may indicate corrupted save data.\n"
                f"Check the console for detailed error location.")
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Error", f"Failed to load character: {str(e)}")
    
    def refresh_stats(self):
        if globals.data is None:
            return
        
        murks, sigs = read_murks_and_sigs(globals.data)
        self.murks_display.config(text=str(murks))
        self.sigs_display.config(text=str(sigs))
    
    def modify_murks(self):
        if globals.data is None:
            messagebox.showwarning("Warning", "No character loaded")
            return

        confrim = messagebox.askyesno("Confirm", "Modifying Murks would get you banned. Are you sure you want to proceed?")
        if not confrim:
            return
        
        new_value = simpledialog.askinteger("Modify Murks", 
                                           f"Current Murks: {current_murks}\n\nEnter new value (decimal):",
                                           initialvalue=current_murks)
        if new_value is not None:
            write_murks_and_sigs(new_value, current_sigs)
            self.refresh_stats()
            messagebox.showinfo("Success", "Murks updated successfully")
    
    def modify_sigs(self):
        if globals.data is None:
            messagebox.showwarning("Warning", "No character loaded")
            return
        
        
        new_value = simpledialog.askinteger("Modify Sigs", 
                                           f"Current Sigs: {current_sigs}\n\nEnter new value (decimal):",
                                           initialvalue=current_sigs)
        if new_value is not None:
            write_murks_and_sigs(current_murks, new_value)
            self.refresh_stats()
            messagebox.showinfo("Success", "Sigs updated successfully")
            
    def reparse(self):
        # Parse items - this updates ga_relic with current data
        gaprint(globals.data)

        # Re-parse vessel assignments
        # parse_vessel_assignments(data)
        loadout_handler.reload_ga_relics(ga_relic)
        loadout_handler.parse()
        
    def refresh_inventory_ui(self):
        global ga_relic, relic_checker
        self.update_inventory_comboboxes()
        # Clear treeview
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Check for illegal relics
        illegal_gas = check_illegal_relics()

        # Check for forbidden relics (unique relics that are technically invalid but allowed)
        forbidden_relics = get_forbidden_relics()

        # Count truly illegal relics (exclude forbidden/unique relics from count)
        # Forbidden relics are marked orange and shouldn't count as "illegal"
        truly_illegal_count = 0
        for ga, id, *_ in ga_relic:
            real_id = id - 2147483648
            if ga in illegal_gas and real_id not in forbidden_relics:
                truly_illegal_count += 1

        # Update illegal count label (only count non-forbidden illegal relics)
        if truly_illegal_count > 0:
            self.illegal_count_label.config(text=f"⚠️ {truly_illegal_count} Illegal Relic(s) Found")
        else:
            self.illegal_count_label.config(text="✓ All Relics Valid")
        
        # Store all relic data for filtering
        self.all_relics = []
        
        # Populate treeview
        for idx, (ga, id, e1, e2, e3, se1, se2, se3, offset, size) in enumerate(ga_relic):
            real_id = id - 2147483648
            
            # Get item name and color
            item_name = "Unknown"
            item_color = "Unknown"
            if str(real_id) in items_json:
                item_name = items_json[str(real_id)]["name"]
                item_color = items_json[str(real_id)].get("color", "Unknown")
            
            # Get effect names
            effects = [e1, e2, e3, se1, se2, se3]
            effect_names = []
            
            for eff in effects:
                if eff == 0:
                    effect_names.append("None")
                elif eff == 4294967295:
                    effect_names.append("Empty")
                elif str(eff) in effects_json:
                    effect_names.append(
                        "".join(effects_json[str(eff)]["name"].splitlines()))
                else:
                    effect_names.append(f"Unknown ({eff})")
            
            # Check if this relic is illegal or forbidden
            is_illegal = ga in illegal_gas
            is_forbidden = real_id in forbidden_relics
            is_curse_illegal = relic_checker and ga in relic_checker.curse_illegal_gas
            is_strict_invalid = relic_checker and ga in relic_checker.strict_invalid_gas

            # Get character assignment (which characters have this relic equipped)
            equipped_by_hero_type = loadout_handler.relic_ga_hero_map.get(ga, [])
            equipped_by = [data_source.character_names[h_t-1] for h_t in equipped_by_hero_type]
            equipped_by_str = ", ".join(equipped_by) if equipped_by else "-"

            # Check if this is a deep relic (ID range 2000000-2019999)
            is_deep = 2000000 <= real_id <= 2019999

            # Determine tag
            tag_list = [ga, real_id]
            if is_forbidden and is_illegal:
                tag_list.append('both')
            elif is_forbidden:
                tag_list.append('forbidden')
            elif is_curse_illegal:
                tag_list.append('curse_illegal')
            elif is_illegal:
                tag_list.append('illegal')
            elif is_strict_invalid:
                tag_list.append('strict_invalid')

            # Get acquisition order from inventory section (matches in-game sorting)
            # Lower number = acquired earlier (oldest)
            acquisition_order = ga_acquisition_order.get(ga, 999999)

            # Store relic data for filtering
            self.all_relics.append({
                'index': idx + 1,
                'ga': ga,
                'acquisition_index': acquisition_order,
                'item_name': item_name,
                'real_id': real_id,
                'item_color': item_color,
                'is_deep': is_deep,
                'equipped_by': equipped_by,
                'equipped_by_str': equipped_by_str,
                'effect_names': effect_names,
                'effect_ids': effects,  # Store raw effect IDs for searching
                'tag_list': tuple(tag_list),
                'is_forbidden': is_forbidden,
                'is_illegal': is_illegal,
                'is_curse_illegal': is_curse_illegal,
                'is_strict_invalid': is_strict_invalid,
                'both': is_forbidden and is_illegal
            })

        # Calculate acquisition rank (1, 2, 3...) based on acquisition_index
        # Sort by acquisition_index and assign ranks
        sorted_by_acq = sorted(self.all_relics, key=lambda r: r.get('acquisition_index', 999999))
        for rank, relic in enumerate(sorted_by_acq, start=1):
            relic['acquisition_rank'] = rank

        # Apply current filter (if any)
        self.filter_relics()
            
    def refresh_inventory_lightly(self):
        """Refresh the inventory UI without heavy loading"""
        global ga_relic, relic_checker

        if globals.data is None:
            messagebox.showwarning("Warning", "No character loaded")
            return
        self.reparse()
        self.refresh_inventory_ui()
        
    def refresh_inventory_and_vessels(self):
        self.refresh_inventory_lightly()
        self.refresh_vessels()
    
    def reload_inventory(self):
        """Reload inventory data from save file and refresh UI"""
        global ga_relic, relic_checker

        if globals.data is None:
            messagebox.showwarning("Warning", "No character loaded")
            return
        
        self.reparse()
        
        def heavy_loading():
            # Update relic checker with new ga_relic and recalculate illegal relics
            if relic_checker:
                relic_checker.ga_relic = ga_relic
                relic_checker.set_illegal_relics()
            
        self.run_task_async(heavy_loading, (), "Loading...",
                            callback=self.refresh_inventory_ui)
    
    def filter_relics(self):
        """Filter relics based on search term and all filter criteria"""
        if not hasattr(self, 'all_relics'):
            return

        # Clear treeview
        for item in self.tree.get_children():
            self.tree.delete(item)

        search_term = self.search_var.get().lower()
        search_field = self.search_field_var.get() if hasattr(self, 'search_field_var') else "All Fields"
        char_filter = self.char_filter_var.get() if hasattr(self, 'char_filter_var') else "All"
        color_filter = self.color_filter_var.get() if hasattr(self, 'color_filter_var') else "All"
        status_filter = self.status_filter_var.get() if hasattr(self, 'status_filter_var') else "All"

        # Filter relics
        filtered_relics = []
        for relic in self.all_relics:
            # Apply search filter based on selected field
            passes_search = True
            if search_term != '':
                if search_field == "All Fields":
                    # Search in all text fields
                    effect_names_str = " ".join(relic['effect_names']).lower()
                    effect_ids_str = " ".join(str(e) for e in relic.get('effect_ids', []))
                    passes_search = (search_term in relic['item_name'].lower() or
                                     search_term in str(relic['real_id']) or
                                     search_term in relic['item_color'].lower() or
                                     search_term in relic.get('equipped_by_str', '').lower() or
                                     search_term in effect_names_str or
                                     search_term in effect_ids_str)
                elif search_field == "Name":
                    passes_search = search_term in relic['item_name'].lower()
                elif search_field == "ID":
                    passes_search = search_term in str(relic['real_id'])
                elif search_field == "Color":
                    passes_search = search_term in relic['item_color'].lower()
                elif search_field == "Effect Name":
                    effect_names_str = " ".join(relic['effect_names']).lower()
                    passes_search = search_term in effect_names_str
                elif search_field == "Effect ID":
                    effect_ids_str = " ".join(str(e) for e in relic.get('effect_ids', []))
                    passes_search = search_term in effect_ids_str
                elif search_field == "Equipped By":
                    passes_search = search_term in relic.get('equipped_by_str', '').lower()

            # Apply character filter
            passes_char = True
            if char_filter != "All":
                equipped_by = relic.get('equipped_by', [])
                passes_char = char_filter in equipped_by

            # Apply color filter
            passes_color = True
            if color_filter != "All":
                passes_color = relic['item_color'] == color_filter

            # Apply status filter
            passes_status = True
            if status_filter != "All":
                if status_filter == "Valid":
                    passes_status = (not relic['is_illegal'] and not relic['is_forbidden']
                                     and not relic.get('is_curse_illegal', False)
                                     and not relic.get('is_strict_invalid', False))
                elif status_filter == "Illegal":
                    passes_status = relic['is_illegal']
                elif status_filter == "Curse Illegal":
                    passes_status = relic.get('is_curse_illegal', False)
                elif status_filter == "Forbidden":
                    passes_status = relic['is_forbidden']
                elif status_filter == "Strict Invalid":
                    passes_status = relic.get('is_strict_invalid', False)
                elif status_filter == "Deep Only":
                    passes_status = relic.get('is_deep', False)

            if passes_search and passes_char and passes_color and passes_status:
                filtered_relics.append(relic)

        # Configure tags once before populating
        self.tree.tag_configure('both', foreground='blue', font=('Arial', 9, 'bold'))
        self.tree.tag_configure('forbidden', foreground='#FF8C00', font=('Arial', 9, 'bold'))
        self.tree.tag_configure('curse_illegal', foreground='#9932CC', font=('Arial', 9, 'bold'))
        self.tree.tag_configure('illegal', foreground='red', font=('Arial', 9, 'bold'))
        self.tree.tag_configure('strict_invalid', foreground='#008080', font=('Arial', 9))  # Teal for strict invalid
        self.tree.tag_configure('deep', foreground='#9999BB')  # Subtle color for deep relics, no background

        # Populate treeview with filtered results
        for relic in filtered_relics:
            is_deep = relic.get('is_deep', False)
            deep_indicator = "🔮" if is_deep else ""  # Crystal ball emoji for deep relics

            # Build tags list - include 'deep' tag for deep relics
            tags = list(relic['tag_list'])
            if is_deep:
                tags.append('deep')

            # Use acquisition_rank for the # column (1 = oldest, 2 = second oldest, etc.)
            self.tree.insert('', 'end', text=str(relic.get('acquisition_rank', 0)),
                           values=(relic['item_name'], deep_indicator, relic['real_id'], relic['item_color'],
                                  relic.get('equipped_by_str', '-'),
                                  relic['effect_names'][0], relic['effect_names'][1], relic['effect_names'][2],
                                  relic['effect_names'][3], relic['effect_names'][4], relic['effect_names'][5]),
                           tags=tuple(tags))

        # Auto-size columns after populating
        autosize_treeview_columns(self.tree)

        # Update search info
        filter_active = search_term or char_filter != "All" or color_filter != "All" or status_filter != "All"
        if filter_active:
            self.search_info_label.config(text=f"Showing {len(filtered_relics)} of {len(self.all_relics)} relics")
        else:
            self.search_info_label.config(text="")

    def clear_search(self):
        """Clear the search box and all filters"""
        self.search_var.set("")
        if hasattr(self, 'search_field_var'):
            self.search_field_var.set("All Fields")
        if hasattr(self, 'char_filter_var'):
            self.char_filter_var.set("All")
        if hasattr(self, 'color_filter_var'):
            self.color_filter_var.set("All")
        if hasattr(self, 'status_filter_var'):
            self.status_filter_var.set("All")
        self.filter_relics()
        self.search_entry.focus()

    def sort_by_column(self, col):
        """Sort treeview by clicked column"""
        if not hasattr(self, 'all_relics') or not self.all_relics:
            return

        # Toggle sort direction if same column clicked
        if self.sort_column == col:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_column = col
            self.sort_reverse = False

        # Define sort key based on column
        def get_sort_key(relic):
            if col == '#':
                # Sort by acquisition rank (1 = oldest, 2 = second oldest, etc.)
                return relic.get('acquisition_rank', 999999)
            elif col == 'Item Name':
                return relic['item_name'].lower()
            elif col == 'Deep':
                return relic.get('is_deep', False)
            elif col == 'Item ID':
                return relic['real_id']
            elif col == 'Color':
                return relic['item_color'].lower() if relic['item_color'] else ''
            elif col == 'Equipped By':
                return relic.get('equipped_by_str', '').lower()
            elif col == 'Effect 1':
                return relic['effect_names'][0].lower()
            elif col == 'Effect 2':
                return relic['effect_names'][1].lower()
            elif col == 'Effect 3':
                return relic['effect_names'][2].lower()
            elif col == 'Curse 1':
                return relic['effect_names'][3].lower()
            elif col == 'Curse 2':
                return relic['effect_names'][4].lower()
            elif col == 'Curse 3':
                return relic['effect_names'][5].lower()
            return 0

        # Sort the data
        self.all_relics.sort(key=get_sort_key, reverse=self.sort_reverse)

        # Re-apply filter to refresh display
        self.filter_relics()

        # Update column header to show sort indicator
        columns = ('Item Name', 'Deep', 'Item ID', 'Color', 'Equipped By', 'Effect 1', 'Effect 2', 'Effect 3',
                   'Curse 1', 'Curse 2', 'Curse 3')
        for c in columns:
            indicator = ''
            if c == col:
                indicator = ' ▼' if self.sort_reverse else ' ▲'
            self.tree.heading(c, text=c + indicator)

        # Update # column header
        if col == '#':
            self.tree.heading('#0', text='#' + (' ▼' if self.sort_reverse else ' ▲'))
        else:
            self.tree.heading('#0', text='#')

    def show_context_menu(self, event):
        # Select item under cursor
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)

            menu = tk.Menu(self.root, tearoff=0)
            menu.add_command(label="Modify", command=self.modify_selected_relic)
            menu.add_command(label="Delete", command=self.delete_selected_relic)
            menu.add_separator()
            menu.add_command(label="📋 Copy Effects", command=self.copy_relic_effects)

            # Only enable paste if we have something in clipboard
            paste_label = "📋 Paste Effects"
            if self.clipboard_effects:
                paste_label += f" (from {self.clipboard_effects[2]})"
            menu.add_command(label=paste_label, command=self.paste_relic_effects,
                           state='normal' if self.clipboard_effects else 'disabled')
            menu.post(event.x_root, event.y_root)

    def copy_relic_effects(self):
        """Copy effects from selected relic to clipboard"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "No relic selected")
            return

        item = selection[0]
        tags = self.tree.item(item, 'tags')
        ga_handle = int(tags[0])
        item_id = int(tags[1])

        # Find the relic data
        for ga, id, e1, e2, e3, se1, se2, se3, offset, size in ga_relic:
            real_id = id - 2147483648
            if ga == ga_handle and real_id == item_id:
                effects = [e1, e2, e3, se1, se2, se3]
                item_name = items_json.get(str(real_id), {}).get("name", f"Unknown ({real_id})")
                self.clipboard_effects = (effects, real_id, item_name)
                messagebox.showinfo("Copied", f"Copied effects from:\n{item_name}\n\nEffects: {len([e for e in effects if e != 0])} effect(s)")
                return

        messagebox.showerror("Error", "Could not find relic data")

    def paste_relic_effects(self):
        """Paste copied effects to selected relic"""
        if not self.clipboard_effects:
            messagebox.showwarning("Warning", "No effects copied. Right-click a relic and select 'Copy Effects' first.")
            return

        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "No relic selected")
            return

        item = selection[0]
        tags = self.tree.item(item, 'tags')
        ga_handle = int(tags[0])
        item_id = int(tags[1])

        # Get target relic info
        target_name = items_json.get(str(item_id), {}).get("name", f"Unknown ({item_id})")
        source_effects, source_id, source_name = self.clipboard_effects

        # Build effect names for display
        effect_names = []
        for eff in source_effects:
            if eff != 0:
                eff_name = effects_json.get(str(eff), {}).get("name", f"Unknown ({eff})")
                effect_names.append(eff_name)

        # Confirm paste
        msg = f"Paste effects from:\n  {source_name}\n\nTo:\n  {target_name}\n\n"
        msg += f"Effects to paste ({len(effect_names)}):\n"
        for name in effect_names[:6]:
            msg += f"  • {name}\n"
        msg += "\nProceed?"

        if not messagebox.askyesno("Confirm Paste", msg):
            return

        # Check if this would make the relic illegal
        if relic_checker:
            would_be_illegal = relic_checker.check_invalidity(item_id, source_effects)
            if would_be_illegal:
                warn_msg = "⚠️ Warning: These effects may not be valid for this relic type.\n\n"
                warn_msg += "The relic may be flagged as illegal after pasting.\n\nContinue anyway?"
                if not messagebox.askyesno("Invalid Effects Warning", warn_msg, icon='warning'):
                    return

        # Apply the effects
        if modify_relic(ga_handle, item_id, source_effects):
            messagebox.showinfo("Success", f"Effects pasted to {target_name}")
            self.refresh_inventory_and_vessels()
        else:
            messagebox.showerror("Error", "Failed to paste effects")

    def modify_selected_relic(self):
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "No relic selected")
            return
        
        item = selection[0]
        tags = self.tree.item(item, 'tags')
        ga_handle = int(tags[0])
        item_id = int(tags[1])
        
        # Check if this is a forbidden relic
        if 'forbidden' in tags:
            result = messagebox.askyesno(
                "⚠️ Warning - Do Not Edit Relic",
                f"This relic (ID: {item_id}) is flagged as 'Do Not Edit'.\n\n"
                "Modifying this relic may cause as ban\n"
                "Are you sure you want to proceed?",
                icon='warning'
            )
            if not result:
                return
        
        # If dialog doesn't exist or was closed, create new one
        if not self.modify_dialog or not self.modify_dialog.dialog.winfo_exists():
            self.modify_dialog = ModifyRelicDialog(self.root, ga_handle, item_id, self.refresh_inventory_and_vessels)
        else:
            # Update existing dialog with new relic
            self.modify_dialog.load_relic(ga_handle, item_id)
            self.modify_dialog.dialog.lift()
    
    def delete_selected_relic(self):
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "No relic selected")
            return
        
        # Check if multiple items selected
        if len(selection) > 1:
            result = messagebox.askyesno("Confirm Delete", 
                                         f"Are you sure you want to delete {len(selection)} relics?")
        else:
            item = selection[0]
            tags = self.tree.item(item, 'tags')
            item_id = int(tags[1])
            result = messagebox.askyesno("Confirm Delete", 
                                         f"Are you sure you want to delete this relic (ID: {item_id})?")
        
        if result:
            deleted_count = 0
            failed_count = 0
            
            for item in selection:
                tags = self.tree.item(item, 'tags')
                ga_handle = int(tags[0])
                item_id = int(tags[1])
                
                if delete_relic(ga_handle, item_id):
                    deleted_count += 1
                else:
                    failed_count += 1
            
            if deleted_count > 0:
                messagebox.showinfo("Success", f"Deleted {deleted_count} relic(s) successfully" + 
                                  (f"\n{failed_count} failed" if failed_count > 0 else ""))
                self.refresh_inventory_lightly()
            else:
                messagebox.showerror("Error", "Failed to delete relics")
    
    def select_all_relics(self):
        """Select all relics in the tree"""
        all_items = self.tree.get_children()
        self.tree.selection_set(all_items)
    
    def deselect_all_relics(self):
        """Deselect all relics"""
        self.tree.selection_remove(self.tree.selection())
    
    def invert_selection(self):
        """Invert the current selection"""
        all_items = self.tree.get_children()
        currently_selected = set(self.tree.selection())
        
        # Select items that aren't currently selected
        new_selection = [item for item in all_items if item not in currently_selected]
        
        self.tree.selection_remove(self.tree.selection())
        self.tree.selection_set(new_selection)
    
    def mass_delete_relics(self):
        """Delete all currently selected relics"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "No relics selected. Use the tree selection to choose relics to delete.")
            return
        
        # Check for forbidden relics in selection
        forbidden_count = 0
        for item in selection:
            tags = self.tree.item(item, 'tags')
            if 'forbidden' in tags:
                forbidden_count += 1
        
        # Confirmation message
        confirm_msg = f"Are you sure you want to delete {len(selection)} selected relic(s)?"
        if forbidden_count > 0:
            confirm_msg += f"\n\n⚠️ WARNING: {forbidden_count} of these are 'Do Not Edit' relics!"
            confirm_msg += "\n\nDeleting these may cause issues!"
        
        result = messagebox.askyesno("Confirm Mass Delete", confirm_msg, icon='warning' if forbidden_count > 0 else 'question')
        
        if not result:
            return
        
        # Delete all selected relics
        deleted_count = 0
        failed_count = 0
        
        for item in selection:
            tags = self.tree.item(item, 'tags')
            ga_handle = int(tags[0])
            item_id = int(tags[1])
            
            if delete_relic(ga_handle, item_id):
                deleted_count += 1
            else:
                failed_count += 1
        
        # Show result
        if deleted_count > 0:
            message = f"Successfully deleted {deleted_count} relic(s)"
            if failed_count > 0:
                message += f"\n{failed_count} failed to delete"
            messagebox.showinfo("Mass Delete Complete", message)
            self.refresh_inventory_lightly()
        else:
            messagebox.showerror("Error", "Failed to delete any relics")

    def mass_fix_incorrect_ids(self):
        """Find and fix all problematic relics (illegal and strict invalid)"""
        global userdata_path

        if globals.data is None:
            messagebox.showwarning("Warning", "No character loaded")
            return

        if relic_checker is None:
            messagebox.showwarning("Warning", "Relic checker not initialized")
            return

        # Find all problematic relics that could be fixed
        fixable_illegal = []
        fixable_strict = []
        unfixable_relics = []

        for ga, id, e1, e2, e3, se1, se2, se3, offset, size in ga_relic:
            real_id = id - 2147483648
            effects = [e1, e2, e3, se1, se2, se3]

            # Skip unique relics
            if real_id in RelicChecker.UNIQUENESS_IDS:
                continue

            item_name = items_json.get(str(real_id), {}).get("name", f"Unknown ({real_id})")
            is_illegal = ga in relic_checker.illegal_gas
            is_strict_invalid = ga in relic_checker.strict_invalid_gas

            if not is_illegal and not is_strict_invalid:
                continue

            # For illegal relics: try to find strictly valid first, then fall back to valid
            if is_illegal:
                # First try strictly valid (best outcome)
                strict_order = relic_checker.get_strictly_valid_order(real_id, effects)
                if strict_order:
                    fixable_illegal.append((ga, id, real_id, real_id, item_name, f"{item_name} (reorder)", strict_order, False))
                    continue

                # Try finding a different ID that's strictly valid
                valid_id = self._find_strictly_valid_relic_id(real_id, effects)
                if valid_id and valid_id != real_id:
                    new_name = items_json.get(str(valid_id), {}).get("name", f"Unknown ({valid_id})")
                    strict_order = relic_checker.get_strictly_valid_order(valid_id, effects)
                    if strict_order and not relic_checker.check_invalidity(valid_id, strict_order):
                        fixable_illegal.append((ga, id, real_id, valid_id, item_name, new_name, strict_order, False))
                        continue

                # Fall back to just valid (not strictly valid)
                valid_id = self._find_valid_relic_id_for_effects(real_id, effects)
                if valid_id is not None:
                    new_name = items_json.get(str(valid_id), {}).get("name", f"Unknown ({valid_id})")
                    if valid_id == real_id:
                        fixable_illegal.append((ga, id, real_id, valid_id, item_name, f"{new_name} (reorder)", effects, True))
                    else:
                        fixable_illegal.append((ga, id, real_id, valid_id, item_name, new_name, effects, True))
                    continue

                unfixable_relics.append((real_id, item_name, "illegal", "No valid ID found"))

            # For strict invalid relics (not illegal)
            elif is_strict_invalid:
                # Try strictly valid permutation for current ID
                strict_order = relic_checker.get_strictly_valid_order(real_id, effects)
                if strict_order:
                    fixable_strict.append((ga, id, real_id, real_id, item_name, f"{item_name} (reorder)", strict_order, False))
                    continue

                # Try finding a different ID that's strictly valid
                valid_id = self._find_strictly_valid_relic_id(real_id, effects)
                if valid_id and valid_id != real_id:
                    new_name = items_json.get(str(valid_id), {}).get("name", f"Unknown ({valid_id})")
                    strict_order = relic_checker.get_strictly_valid_order(valid_id, effects)
                    if strict_order and not relic_checker.check_invalidity(valid_id, strict_order):
                        fixable_strict.append((ga, id, real_id, valid_id, item_name, new_name, strict_order, False))
                        continue

                unfixable_relics.append((real_id, item_name, "strict", "No valid permutation found"))

        fixable_relics = fixable_illegal + fixable_strict

        if not fixable_relics:
            msg = "No fixable relics found.\n\n"
            if unfixable_relics:
                msg += f"{len(unfixable_relics)} relic(s) cannot be auto-fixed:\n"
                for real_id, name, issue_type, reason in unfixable_relics[:5]:
                    msg += f"• {name} ({issue_type}): {reason}\n"
                if len(unfixable_relics) > 5:
                    msg += f"... and {len(unfixable_relics) - 5} more\n"
                msg += "\nThese may need manual effect changes."
            messagebox.showinfo("Mass Fix", msg)
            return

        # Show confirmation with details
        details = ""
        if fixable_illegal:
            details += f"Illegal relics: {len(fixable_illegal)}\n"
        if fixable_strict:
            details += f"Strict invalid relics: {len(fixable_strict)}\n"
        details += f"\nTotal: {len(fixable_relics)} relic(s) to fix:\n\n"

        for i, (ga, id, old_id, new_id, old_name, new_name, effects, is_fallback) in enumerate(fixable_relics[:10]):
            marker = " ⚠️" if is_fallback else ""
            if old_id == new_id:
                details += f"• {old_name} → reorder effects{marker}\n"
            else:
                details += f"• {old_name} → {new_name}{marker}\n"
        if len(fixable_relics) > 10:
            details += f"\n... and {len(fixable_relics) - 10} more"

        if any(r[7] for r in fixable_relics):  # Check if any fallback fixes
            details += "\n\n⚠️ = Fixed to valid but may still have 0% weight effects"

        details += "\n\nProceed with fixing these relics?"

        result = messagebox.askyesno("Confirm Mass Fix", details)
        if not result:
            return

        # Apply fixes
        fixed_count = 0
        failed_count = 0

        for ga, id, old_id, new_id, old_name, new_name, new_effects, is_fallback in fixable_relics:
            gaprint(globals.data)
            # Use sort_effects=True for fallback fixes (need sorting), False for strict fixes (already sorted)
            if modify_relic_by_ga(ga, new_effects, new_id, sort_effects=is_fallback):
                fixed_count += 1
            else:
                failed_count += 1

        # Reload data
        if userdata_path:
            with open(userdata_path, 'rb') as f:
                globals.data = bytearray(f.read())

        # Show result
        message = f"Fixed {fixed_count} relic(s)"
        if fixable_illegal:
            message += f"\n• {len([r for r in fixable_illegal if r[7] == False])} illegal → strictly valid"
            fallback_count = len([r for r in fixable_illegal if r[7] == True])
            if fallback_count:
                message += f"\n• {fallback_count} illegal → valid (may still be strict invalid)"
        if fixable_strict:
            message += f"\n• {len(fixable_strict)} strict invalid → strictly valid"
        if failed_count > 0:
            message += f"\n\n{failed_count} failed to fix"

        messagebox.showinfo("Mass Fix Complete", message)
        self.refresh_inventory_lightly()

    def add_relic_tk(self):
        if globals.data is None:
            messagebox.showwarning(
                "Warning", "No save file loaded. Please open a save file first."
            )
            return
        added_result, new_ga = add_relic()
        if added_result:
            messagebox.showinfo("Success", "Dummy relic added. Refreshing inventory.")
            self.refresh_inventory_and_vessels()
            # Find Added item by new_ga
            for item in self.tree.get_children():
                item_ga = int(self.tree.item(item, 'tags')[0])
                if item_ga == new_ga:
                    self.tree.selection_set(item)
                    self.tree.focus(item)
                    self.tree.see(item)
                    break
            self.modify_selected_relic()
        else:
            messagebox.showerror("Error", f"Failed to add relic: {e}")

    def _find_valid_relic_id_for_effects(self, current_id, effects):
        """Find a valid relic ID that can have the given effects (must be same color)"""
        # Get the current relic's color
        if current_id not in data_source.relic_table.index:
            return None
        current_color = data_source.relic_table.loc[current_id, "relicColor"]

        # Count how many effects need curses
        curses_needed = sum(1 for e in effects[:3]
                          if e not in [0, -1, 4294967295] and data_source.effect_needs_curse(e))

        # Get the range group of current ID
        id_range = relic_checker.find_id_range(current_id)
        if not id_range:
            return None

        group_name, (range_start, range_end) = id_range

        # Skip illegal range
        if group_name == "illegal":
            return None

        # First check if current ID is actually valid (effects match AND has enough curse slots)
        # Use allow_empty_curses=True because we're checking if primary effects fit the pools
        if relic_checker.has_valid_order(current_id, effects):
            # Also check it has enough curse slots for effects that need curses
            try:
                pools = data_source.get_relic_pools_seq(current_id)
                available_curse_slots = sum(1 for p in pools[3:] if p != -1)
                if available_curse_slots >= curses_needed:
                    return current_id  # Already valid, return same ID
            except KeyError:
                pass

        # Search within the same range for a valid ID with SAME color only
        for test_id in range(range_start, range_end + 1):
            if test_id not in data_source.relic_table.index:
                continue

            # Must be same color - never change color
            test_color = data_source.relic_table.loc[test_id, "relicColor"]
            if test_color != current_color:
                continue

            # Check if relic has enough curse slots for effects that need curses
            try:
                pools = data_source.get_relic_pools_seq(test_id)
                available_curse_slots = sum(1 for p in pools[3:] if p != -1)
                if available_curse_slots < curses_needed:
                    continue
            except KeyError:
                continue

            # Check if effects are valid for this ID (allow empty curses)
            if relic_checker.has_valid_order(test_id, effects):
                return test_id

        return None

    def _find_strictly_valid_relic_id(self, current_id, effects):
        """Find a relic ID where effects can be strictly valid (same color)"""
        if current_id not in data_source.relic_table.index:
            return None
        current_color = data_source.relic_table.loc[current_id, "relicColor"]

        # Get range
        id_range = relic_checker.find_id_range(current_id)
        if not id_range:
            return None

        group_name, (range_start, range_end) = id_range
        if group_name == "illegal":
            return None

        # Search for strictly valid ID
        for test_id in range(range_start, range_end + 1):
            if test_id == current_id:
                continue
            if test_id not in data_source.relic_table.index:
                continue

            test_color = data_source.relic_table.loc[test_id, "relicColor"]
            if test_color != current_color:
                continue

            # Check if effects can be strictly valid with this ID
            valid_order = relic_checker.get_strictly_valid_order(test_id, effects)
            if valid_order:
                return test_id

        return None

    def export_relics(self):
        """Export relics to Excel file"""
        if globals.data is None:
            messagebox.showwarning("Warning", "No character loaded")
            return
        
        filepath = filedialog.asksaveasfilename(
            title="Save Excel File",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )
        
        if not filepath:
            return
        
        success, message = export_relics_to_excel(filepath)
        
        if success:
            messagebox.showinfo("Success", message)
        else:
            messagebox.showerror("Error", message)
    
    def import_relics(self):
        """Import relics from Excel file"""
        if globals.data is None:
            messagebox.showwarning("Warning", "No character loaded")
            return
        
        filepath = filedialog.askopenfilename(
            title="Open Excel File",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )
        
        if not filepath:
            return
        
        # Confirm import
        result = messagebox.askyesno(
            "Confirm Import",
            "This will modify your current relics to match the imported file.\n\n"
            "• Relic IDs and effects will be replaced\n"
            "• Extra relics in the file will be ignored\n"
            "• Make sure you have a backup!\n\n"
            "Continue?"
        )
        
        if not result:
            return
        
        success, message = import_relics_from_excel(filepath)
        
        if success:
            messagebox.showinfo("Success", message)
            self.reload_inventory()
        else:
            messagebox.showerror("Error", message)
    
    def delete_all_illegal(self):
        """Delete all relics with illegal effects"""
        if globals.data is None:
            messagebox.showwarning("Warning", "No character loaded")
            return
        
        # Check for illegal relics first
        illegal_gas = check_illegal_relics()
        
        if not illegal_gas:
            messagebox.showinfo("Info", "No illegal relics found!")
            return
        
        # Confirm deletion
        result = messagebox.askyesno(
            "Confirm Deletion", 
            f"Found {len(illegal_gas)} illegal relic(s).\n\n"
            "This will permanently delete all relics with:\n"
            "• Effects in the illegal list\n"
            "• Duplicate effect IDs\n"
            "• Conflicting effect tiers\n\n"
            "Do you want to proceed?"
        )
        
        if not result:
            return
        
        # Delete all illegal relics
        count, message = delete_all_illegal_relics()
        
        if count > 0:
            messagebox.showinfo("Success", message)
            self.refresh_inventory_lightly()
        else:
            messagebox.showerror("Error", message)
    def import_save_tk(self):
        import_save()
        self.load_character(userdata_path)
    
    def save_changes(self):
        if globals.data and userdata_path:
            save_file()
            messagebox.showinfo("Success", "Changes saved to file")
        else:
            messagebox.showwarning("Warning", "No character loaded")


class ModifyRelicDialog:
    def __init__(self, parent, ga_handle, item_id, callback):
        self.callback = callback
        
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Modify Relic")
        self.dialog.geometry("700x600")
        self.dialog.transient(parent)

        self.safe_mode_var = tk.BooleanVar(value=True)
        
        self.setup_ui()
        self.load_relic(ga_handle, item_id)
    
    def load_relic(self, ga_handle, item_id):
        """Load relic data into the dialog"""
        self.ga_handle = ga_handle
        self.item_id = item_id

        # Get current effects
        self.current_effects = self.get_current_effects()

        # Update UI
        self.update_effects_display()

        # Update debug info (after UI is set up)
        self.dialog.after(100, self.update_debug_info)
    
    def get_current_effects(self):
        for ga, id, e1, e2, e3, se1, se2, se3, offset, size in ga_relic:
            real_id = id - 2147483648
            if ga == self.ga_handle and real_id == self.item_id:
                return [e1, e2, e3, se1, se2, se3]
        return [0, 0, 0, 0, 0, 0]
    
    def update_effects_display(self):
        """Update the effect entry fields with current values"""
        # Update item ID display (use StringVar - this triggers _on_item_id_change)
        self.current_item_label.config(text=f"{self.item_id}")
        self.item_id_var.set(str(self.item_id))

        for i, entry in enumerate(self.effect_entries):
            current_eff = self.current_effects[i]
            if current_eff == 0:
                entry.delete(0, tk.END)
                entry.insert(0, "0")
            else:
                entry.delete(0, tk.END)
                entry.insert(0, str(current_eff))

                # Also update the name display
                if str(current_eff) in effects_json:
                    name = effects_json[str(current_eff)]["name"]
                    self.effect_name_labels[i].config(text=name)
                else:
                    self.effect_name_labels[i].config(text="Unknown Effect")

        # Update curse indicators after loading
        self._update_curse_indicators()

    def _update_color_display(self):
        """Update the current color label"""
        color_map = {0: ("Red", "#ff6666"), 1: ("Blue", "#6666ff"),
                     2: ("Yellow", "#cccc00"), 3: ("Green", "#66cc66")}
        try:
            current_relic_id = int(self.item_id_entry.get())
            color_idx = data_source.relic_table.loc[current_relic_id, "relicColor"]
            color_name, color_hex = color_map.get(color_idx, ("Unknown", "gray"))
            self.current_color_label.config(text=color_name, foreground=color_hex)
        except (KeyError, ValueError):
            self.current_color_label.config(text="Unknown", foreground="gray")

    def _update_relic_structure_display(self):
        """Update the relic structure label showing effect/curse slots"""
        try:
            current_relic_id = int(self.item_id_entry.get())
            pools = data_source.get_relic_pools_seq(current_relic_id)

            # Count effect slots and curse slots
            effect_slots = sum(1 for p in pools[:3] if p != -1)
            curse_slots = sum(1 for p in pools[3:] if p != -1)

            # Format the display text
            effect_text = f"{effect_slots} effect slot" + ("s" if effect_slots != 1 else "")
            curse_text = f"{curse_slots} curse slot" + ("s" if curse_slots != 1 else "")

            structure_text = f"({effect_text}, {curse_text})"
            self.relic_structure_label.config(text=structure_text)
        except (KeyError, ValueError):
            self.relic_structure_label.config(text="(Unknown structure)")

    def _on_item_id_change(self):
        """Called when item ID entry changes - updates color and structure display"""
        self._update_color_display()
        self._update_relic_structure_display()
        self._update_relic_type_display()

    def _update_relic_type_display(self):
        """Update the relic type indicator (Original vs Scene/1.02)"""
        try:
            current_relic_id = int(self.item_id_entry.get())
            type_name, description, color_hex = data_source.get_relic_type_info(current_relic_id)
            self.relic_type_label.config(text=type_name, foreground=color_hex)
            self.relic_type_info_label.config(text=f"— {description}")
        except (KeyError, ValueError):
            self.relic_type_label.config(text="Unknown", foreground="gray")
            self.relic_type_info_label.config(text="")

    def _update_illegal_status_display(self, relic_id, effects):
        """Update the prominent illegal status display with human-readable reasons"""
        if not hasattr(self, 'status_label'):
            return

        try:
            pools = data_source.get_relic_pools_seq(relic_id)
        except KeyError:
            self.status_label.config(text="❌ ILLEGAL: Unknown Relic ID", foreground='red')
            self.illegal_reason_label.config(text=f"Relic ID {relic_id} does not exist in the game data.")
            return

        # Get invalid reason with index of first problematic effect
        invalid_reason, invalid_idx = relic_checker.check_invalidity(relic_id, effects, True) if relic_checker else (InvalidReason.VALIDATION_ERROR, -1)
        is_curse_illegal_flag = is_curse_invalid(invalid_reason)

        if invalid_reason == InvalidReason.NONE:
            # Check for strict invalid (valid but 0% weight effects)
            is_strict_invalid = relic_checker.is_strict_invalid(relic_id, effects, invalid_reason) if relic_checker else False
            if is_strict_invalid:
                strict_reason = relic_checker.get_strict_invalid_reason(relic_id, effects) if relic_checker else None
                self.status_label.config(text="⚠️ STRICT INVALID", foreground='#008080')  # Teal color
                reason_text = "This relic is technically valid but has effects with 0% drop weight in their assigned pools.\n"
                if strict_reason:
                    reason_text += f"\n{strict_reason}\n"
                reason_text += "\nThis may cause detection or unexpected behavior. Use 'Mass Fix' or 'Find Valid ID' to resolve."
                self.illegal_reason_label.config(text=reason_text)
            else:
                self.status_label.config(text="✅ VALID", foreground='green')
                self.illegal_reason_label.config(text="This relic configuration is legal and will work in-game.")
            return

        # Determine relic type for clearer messaging
        effect_pools = [p for p in pools[:3] if p != -1]
        curse_slot_count = sum(1 for p in pools[3:] if p != -1)
        is_deep_relic = any(p >= 2000000 for p in effect_pools)
        relic_type_desc = "Deep Relic" if is_deep_relic else "Normal Relic"

        # Check if fixable by reordering
        if invalid_reason == InvalidReason.EFFS_NOT_SORTED:
            reasons = [
                "• Effects are valid but in the WRONG ORDER",
                "• Use 'Find Valid ID' button or 'Mass Fix' to automatically reorder"
            ]
            self.status_label.config(text="❌ ILLEGAL (fixable)", foreground='#CC6600')
            self.illegal_reason_label.config(text="\n".join(reasons))
            return

        # Build human-readable message based on InvalidReason enum
        reason_messages = {
            InvalidReason.IN_ILLEGAL_RANGE: "• Relic ID is in an illegal/reserved range (20000-30035)",
            InvalidReason.INVALID_ITEM: f"• Relic ID {relic_id} is not a valid relic",
            InvalidReason.EFF_CONFLICT: "• Two effects have the same conflict ID and cannot be combined",
            InvalidReason.CURSE_CONFLICT: "• Two curses have the same conflict ID and cannot be combined",
            InvalidReason.CURSES_NOT_ENOUGH: f"• Not enough curses provided - some effects require curses on {relic_type_desc}s",
        }

        reasons = []

        # Add the main reason message
        if invalid_reason in reason_messages:
            reasons.append(reason_messages[invalid_reason])

        # Add specific details for effect/curse pool issues
        if invalid_reason == InvalidReason.EFF_NOT_IN_ROLLABLE_POOL:
            if invalid_idx >= 0 and invalid_idx < 3:
                eff = effects[invalid_idx]
                eff_name = effects_json.get(str(eff), {}).get("name", f"Unknown ({eff})")
                # Check if effect can exist on this relic type at all
                all_relic_pools = set(p for p in pools[:3] if p != -1)
                can_exist = any(eff in data_source.get_pool_rollable_effects(p) for p in all_relic_pools)
                if not can_exist:
                    reasons.append(f"• Effect '{eff_name}' has 0% drop weight on {relic_type_desc}s - cannot exist on this relic type")
                else:
                    reasons.append(f"• Effect '{eff_name}' is not valid in slot {invalid_idx + 1}")
            else:
                reasons.append(f"• An effect is not valid for this relic's pools")

        if invalid_reason == InvalidReason.EFF_MUST_EMPTY:
            if invalid_idx >= 0 and invalid_idx < 3:
                eff = effects[invalid_idx]
                eff_name = effects_json.get(str(eff), {}).get("name", f"Unknown ({eff})")
                reasons.append(f"• Effect slot {invalid_idx + 1} should be empty but has '{eff_name}'")
            else:
                reasons.append("• An effect slot should be empty but has a value")

        if invalid_reason == InvalidReason.CURSE_MUST_EMPTY:
            if invalid_idx >= 3:
                curse = effects[invalid_idx]
                curse_name = effects_json.get(str(curse), {}).get("name", f"Unknown ({curse})")
                slot_num = invalid_idx - 2  # Convert to 1-based curse slot
                reasons.append(f"• Curse slot {slot_num} should be empty but has '{curse_name}'")
            else:
                reasons.append("• A curse slot should be empty but has a value")

        if invalid_reason == InvalidReason.CURSE_REQUIRED_BY_EFFECT:
            if invalid_idx >= 3:
                eff_idx = invalid_idx - 3
                eff = effects[eff_idx]
                eff_name = effects_json.get(str(eff), {}).get("name", f"Unknown ({eff})")
                reasons.append(f"• Effect '{eff_name}' REQUIRES a curse but curse slot {eff_idx + 1} is empty")
            else:
                reasons.append("• An effect requires a curse but the corresponding curse slot is empty")

        if invalid_reason == InvalidReason.CURSE_NOT_IN_ROLLABLE_POOL:
            if invalid_idx >= 3:
                curse = effects[invalid_idx]
                curse_name = effects_json.get(str(curse), {}).get("name", f"Unknown ({curse})")
                slot_num = invalid_idx - 2
                reasons.append(f"• Curse '{curse_name}' is not valid for slot {slot_num}")
            else:
                reasons.append("• A curse is not valid for its assigned pool")

        # Fallback if no specific message was added
        if not reasons:
            reasons.append(f"• Invalid: {invalid_reason.name}")

        # Set the display
        self.status_label.config(text="❌ ILLEGAL", foreground='red')
        self.illegal_reason_label.config(text="\n".join(reasons))

    def update_debug_info(self):
        """Update debug info showing why relic is flagged"""
        if not hasattr(self, 'debug_text'):
            return

        self.debug_text.config(state='normal')
        self.debug_text.delete('1.0', tk.END)

        try:
            relic_id = int(self.item_id_entry.get())
            effects = []
            for entry in self.effect_entries:
                try:
                    effects.append(int(entry.get()))
                except ValueError:
                    effects.append(4294967295)

            # Update the human-readable status display
            self._update_illegal_status_display(relic_id, effects)

            lines = []
            lines.append(f"Relic ID: {relic_id}")
            lines.append(f"Effects: {effects[:3]}")
            lines.append(f"Curses:  {effects[3:]}")
            lines.append("")

            # Get relic pools
            try:
                pools = data_source.get_relic_pools_seq(relic_id)
                lines.append(f"Pools (eff1,eff2,eff3,curse1,curse2,curse3): {pools}")
            except KeyError:
                lines.append(f"Pools: Relic ID {relic_id} not found in table!")
                pools = None

            lines.append("")

            # Check validation status
            if relic_checker:
                invalid_reason, invalid_effect_index = relic_checker.check_invalidity(relic_id, effects, True)
                is_curse_illegal = is_curse_invalid(invalid_reason)
                _invalid_idx_msg = "" if invalid_effect_index == -1 else f" at effect index {invalid_effect_index + 1}"
                _invalid_curse_idx_msg = "" if invalid_effect_index == -1 else f" at curse index {invalid_effect_index - 2}"
                lines.append(f"invalid_reason(): {invalid_reason.name}{_invalid_idx_msg}")
                lines.append(f"is_curse_illegal(): {is_curse_illegal}{_invalid_curse_idx_msg}")
                lines.append("")

                # Show effect analysis
                lines.append("--- Effect Analysis ---")
                if pools:
                    effect_slot_count = sum(1 for p in pools[:3] if p != -1)
                    lines.append(f"Effect slots: {effect_slot_count} ({'single-effect' if effect_slot_count <= 1 else 'multi-effect'} relic)")

                    if effect_slot_count <= 1:
                        lines.append("Single-effect relics don't require curses")
                    else:
                        curse_required_count = 0
                        for i, eff in enumerate(effects[:3]):
                            if eff in [-1, 0, 4294967295]:
                                lines.append(f"Effect {i}: {eff} (empty)")
                            else:
                                needs_curse = data_source.effect_needs_curse(eff)
                                effect_pools = data_source.get_effect_pools(eff)
                                if needs_curse:
                                    curse_required_count += 1
                                lines.append(f"Effect {i}: {eff} -> needs_curse={needs_curse}, pools={effect_pools}")

                        curses_provided = sum(1 for c in effects[3:] if c not in [-1, 0, 4294967295])
                        lines.append("")
                        lines.append(f"Effects needing curses: {curse_required_count}, Curses provided: {curses_provided}")
                        if curse_required_count > curses_provided:
                            lines.append("⚠ NOT ENOUGH CURSES for effects that need them!")
                lines.append("")

                # Detailed check
                if pools:
                    effect_slot_count = sum(1 for p in pools[:3] if p != -1)
                    is_multi_effect = effect_slot_count > 1

                    lines.append("--- Sequence Check Details ---")
                    if not is_multi_effect:
                        lines.append("(Single-effect relic - curse checks skipped)")

                    possible_sequences = [[0, 1, 2], [0, 2, 1], [1, 0, 2],
                                          [1, 2, 0], [2, 0, 1], [2, 1, 0]]

                    for seq in possible_sequences:
                        cur_effects = [effects[i] for i in seq]
                        cur_curses = [effects[i + 3] for i in seq]

                        seq_valid = True
                        issues = []

                        for idx in range(3):
                            eff = cur_effects[idx]
                            curse = cur_curses[idx]
                            eff_pool = pools[idx]
                            curse_pool = pools[idx + 3]

                            # Check effect
                            if eff_pool == -1:
                                if eff != 4294967295:
                                    seq_valid = False
                                    issues.append(f"slot{idx}: pool=-1 but eff={eff}")
                            else:
                                pool_effs = data_source.get_pool_effects_strict(eff_pool)
                                if eff not in pool_effs:
                                    seq_valid = False
                                    issues.append(f"slot{idx}: eff {eff} not in pool {eff_pool}")

                            # Check curse (only for multi-effect relics)
                            if is_multi_effect:
                                eff_needs_curse = data_source.effect_needs_curse(eff)

                                if eff_needs_curse:
                                    # Effect requires a curse
                                    if curse_pool == -1:
                                        issues.append(f"curse{idx}: effect needs curse but slot has no curse_pool!")
                                        seq_valid = False
                                    elif curse in [-1, 0, 4294967295]:
                                        issues.append(f"curse{idx}: effect REQUIRES curse, but empty!")
                                        seq_valid = False
                                    else:
                                        pool_curses = data_source.get_pool_effects_strict(curse_pool)
                                        if curse not in pool_curses:
                                            issues.append(f"curse{idx}: {curse} not in pool {curse_pool}")
                                            seq_valid = False
                                elif curse_pool != -1:
                                    # Effect doesn't need curse but slot supports one (optional)
                                    if curse not in [-1, 0, 4294967295]:
                                        pool_curses = data_source.get_pool_effects_strict(curse_pool)
                                        if curse not in pool_curses:
                                            issues.append(f"curse{idx}: {curse} not in pool {curse_pool}")
                                            seq_valid = False
                                else:
                                    # curse_pool == -1: no curse allowed
                                    if curse not in [-1, 0, 4294967295]:
                                        issues.append(f"curse{idx}: slot doesn't support curse but curse={curse}")
                                        seq_valid = False

                        status = "✓ VALID" if seq_valid else "✗ invalid"
                        lines.append(f"Seq {seq}: {status}")
                        if issues:
                            for issue in issues:
                                lines.append(f"  - {issue}")

            self.debug_text.insert('1.0', '\n'.join(lines))
        except Exception as e:
            self.debug_text.insert('1.0', f"Error generating debug info: {str(e)}")

        self.debug_text.config(state='disabled')

    def setup_ui(self):
        # Title showing current relic
        self.title_label = ttk.Label(self.dialog, text="", font=('Arial', 14, 'bold'))
        self.title_label.pack(pady=10)

        # Main container with scrollbar
        main_frame = ttk.Frame(self.dialog)
        main_frame.pack(fill='both', expand=True, padx=10, pady=6)

        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Enable mouse wheel scrolling only when mouse is over this dialog
        def _on_mousewheel(event):
            if canvas.winfo_exists():
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")

        def _bind_mousewheel(event):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)

        def _unbind_mousewheel(event=None):
            canvas.unbind_all("<MouseWheel>")

        # Bind on enter, unbind on leave
        self.dialog.bind("<Enter>", _bind_mousewheel)
        self.dialog.bind("<Leave>", _unbind_mousewheel)
        # Also unbind when dialog is destroyed to prevent orphaned bindings
        self.dialog.bind("<Destroy>", _unbind_mousewheel)

        # Illegal Status Section - prominent display at top
        self.status_frame = ttk.LabelFrame(scrollable_frame, text="⚠️ Relic Status", padding=10)
        self.status_frame.pack(fill='x', pady=5)

        self.status_label = ttk.Label(self.status_frame, text="Checking...", font=('Arial', 11, 'bold'))
        self.status_label.pack(anchor='w')

        self.illegal_reason_label = ttk.Label(self.status_frame, text="", font=('Arial', 10),
                                               wraplength=650, foreground='#333333')
        self.illegal_reason_label.pack(anchor='w', pady=(5, 0))

        # Modifier Config Section
        modifier_frame = ttk.LabelFrame(scrollable_frame, text="Modifier Configuration", padding=10)
        modifier_frame.pack(fill='x', pady=5)
        self.safe_mode_cb = ttk.Checkbutton(modifier_frame,
                                            text="Safe Mode: Auto-filter legal effects",
                                            variable=self.safe_mode_var,
                                            onvalue=True, offvalue=False)
        self.safe_mode_cb.pack(anchor='w')
        
        # Item ID section (optional modification)
        item_frame = ttk.LabelFrame(scrollable_frame, text="Relic Item ID", padding=10)
        item_frame.pack(fill='x', pady=5)

        item_info_frame = ttk.Frame(item_frame)
        item_info_frame.pack(fill='x', anchor='w')

        ttk.Label(item_info_frame, text="Current Item ID:").pack(side='left')
        self.current_item_label = ttk.Label(item_info_frame, text="", font=('Arial', 10, 'bold'), foreground='blue')
        self.current_item_label.pack(side='left', padx=(5, 15))

        self.relic_structure_label = ttk.Label(item_info_frame, text="", font=('Arial', 9), foreground='#666666')
        self.relic_structure_label.pack(side='left')

        # Relic type indicator (Original vs Scene/1.02)
        relic_type_frame = ttk.Frame(item_frame)
        relic_type_frame.pack(fill='x', anchor='w', pady=(5, 0))

        self.relic_type_label = ttk.Label(relic_type_frame, text="", font=('Arial', 9, 'bold'))
        self.relic_type_label.pack(side='left')

        self.relic_type_info_label = ttk.Label(relic_type_frame, text="", font=('Arial', 8), foreground='#888888')
        self.relic_type_info_label.pack(side='left', padx=(10, 0))

        ttk.Label(item_frame, text="Enter new Item ID (decimal) or search:").pack(anchor='w', pady=(10, 0))

        item_entry_frame = ttk.Frame(item_frame)
        item_entry_frame.pack(fill='x', pady=5)

        self.item_id_var = tk.StringVar()
        self.item_id_var.trace('w', lambda *args: self._on_item_id_change())
        self.item_id_entry = ttk.Entry(item_entry_frame, width=15, textvariable=self.item_id_var)
        self.item_id_entry.pack(side='left', padx=5)

        ttk.Button(item_entry_frame, text="Search Items", command=self.search_items).pack(side='left', padx=5)
        ttk.Button(item_entry_frame, text="🔧 Find Valid ID", command=self.find_valid_relic_id).pack(side='left', padx=5)

        # Color change section
        color_frame = ttk.LabelFrame(scrollable_frame, text="Change Relic Color", padding=10)
        color_frame.pack(fill='x', pady=5)

        color_info_frame = ttk.Frame(color_frame)
        color_info_frame.pack(fill='x', pady=(0, 5))

        ttk.Label(color_info_frame, text="Current Color:").pack(side='left', padx=5)
        self.current_color_label = ttk.Label(color_info_frame, text="", font=('Arial', 10, 'bold'))
        self.current_color_label.pack(side='left', padx=5)

        color_btn_frame = ttk.Frame(color_frame)
        color_btn_frame.pack(fill='x', pady=5)

        ttk.Label(color_btn_frame, text="Change to:").pack(side='left', padx=5)

        # Color buttons with actual colors
        self.red_btn = tk.Button(color_btn_frame, text="🔴 Red", bg='#ffcccc', activebackground='#ff9999',
                                  command=lambda: self.change_relic_color(0))
        self.red_btn.pack(side='left', padx=3)

        self.blue_btn = tk.Button(color_btn_frame, text="🔵 Blue", bg='#ccccff', activebackground='#9999ff',
                                   command=lambda: self.change_relic_color(1))
        self.blue_btn.pack(side='left', padx=3)

        self.yellow_btn = tk.Button(color_btn_frame, text="🟡 Yellow", bg='#ffffcc', activebackground='#ffff99',
                                     command=lambda: self.change_relic_color(2))
        self.yellow_btn.pack(side='left', padx=3)

        self.green_btn = tk.Button(color_btn_frame, text="🟢 Green", bg='#ccffcc', activebackground='#99ff99',
                                    command=lambda: self.change_relic_color(3))
        self.green_btn.pack(side='left', padx=3)

        # Effect modification section
        effect_frame = ttk.LabelFrame(scrollable_frame, text="Modify Effects", padding=10)
        effect_frame.pack(fill='x', pady=5)

        self.effect_entries = []
        self.effect_name_labels = []
        self.slot_labels = []  # Store references to slot labels for updating curse indicators
        effect_labels = ['Effect 1', 'Effect 2', 'Effect 3',
                        'Curse 1', 'Curse 2', 'Curse 3']

        for i, label in enumerate(effect_labels):
            # Label with indicator for curse slots
            is_curse_slot = i >= 3
            slot_label = ttk.Label(effect_frame, text=f"{label}:", font=('Arial', 10, 'bold'))
            slot_label.grid(row=i*2, column=0, sticky='w', pady=(10, 2))
            self.slot_labels.append(slot_label)

            # Entry frame
            entry_frame = ttk.Frame(effect_frame)
            entry_frame.grid(row=i*2+1, column=0, sticky='ew', pady=(0, 5))

            # Manual entry
            entry = ttk.Entry(entry_frame, width=15)
            entry.pack(side='left', padx=5)
            entry.bind('<KeyRelease>', lambda e, idx=i: self.on_effect_change(idx))
            self.effect_entries.append(entry)

            # Search button
            ttk.Button(entry_frame, text="Search Effects",
                      command=lambda idx=i: self.search_effects(idx)).pack(side='left', padx=5)

            # Effect name display
            name_label = ttk.Label(entry_frame, text="", foreground='blue')
            name_label.pack(side='left', padx=5)
            self.effect_name_labels.append(name_label)
        
        effect_frame.grid_columnconfigure(0, weight=1)

        # Debug info section
        debug_frame = ttk.LabelFrame(scrollable_frame, text="⚙ Debug: Validation Info", padding=10)
        debug_frame.pack(fill='x', pady=5)

        self.debug_text = tk.Text(debug_frame, height=12, width=80, font=('Consolas', 9),
                                   state='disabled', bg='#f5f5f5')
        self.debug_text.pack(fill='x', padx=5, pady=5)

        ttk.Button(debug_frame, text="🔄 Refresh Debug Info",
                   command=self.update_debug_info).pack(anchor='w', padx=5)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Buttons at bottom
        button_frame = ttk.Frame(self.dialog)
        button_frame.pack(fill='x', padx=10, pady=10)

        ttk.Label(button_frame, text="Click different relics in inventory to switch",
                 foreground='gray').pack(side='left', padx=5)
        ttk.Button(button_frame, text="Apply Changes", command=self.apply_changes).pack(side='right', padx=5)
        ttk.Button(button_frame, text="Close", command=self.dialog.destroy).pack(side='right', padx=5)
        ttk.Button(button_frame, text="🔄 Auto Sort", command=self.auto_sort_effects).pack(side='right', padx=5)

    def auto_sort_effects(self):
        """Sort effects in the correct order (keeps curses with their corresponding effects)"""
        try:
            # Get current effects from entry fields
            current_effects = []
            for entry in self.effect_entries:
                try:
                    val = int(entry.get())
                    current_effects.append(val)
                except ValueError:
                    current_effects.append(4294967295)  # Empty

            if len(current_effects) != 6:
                messagebox.showerror("Error", "Invalid effects configuration")
                return

            # Use relic_checker to sort effects (keeps curses paired with their primary effects)
            if relic_checker:
                sorted_effects = relic_checker.sort_effects(current_effects)

                # Update entry fields with sorted values
                for i, entry in enumerate(self.effect_entries):
                    entry.delete(0, tk.END)
                    entry.insert(0, str(sorted_effects[i]))
                    self.on_effect_change(i)  # Update name labels

                self.update_debug_info()
                messagebox.showinfo("Auto Sort", "Effects sorted successfully!\nCurses remain paired with their corresponding effects.")
            else:
                messagebox.showerror("Error", "Relic checker not initialized")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to sort effects: {str(e)}")

    def on_effect_change(self, index):
        """When effect ID is manually entered, update the name display"""
        try:
            effect_id = int(self.effect_entries[index].get())
            if effect_id in [0, 4294967295, -1]:
                self.effect_name_labels[index].config(text="None")
            elif str(effect_id) in effects_json:
                name = effects_json[str(effect_id)]["name"]
                self.effect_name_labels[index].config(text=name)
            else:
                self.effect_name_labels[index].config(text="Unknown Effect")
        except ValueError:
            self.effect_name_labels[index].config(text="Invalid ID")

        # Auto-find valid relic ID in background (silent mode)
        # This handles cases where selected effect needs a curse
        self._auto_find_valid_relic_id()

        # Update curse indicators when effect or curse slots change
        self._update_curse_indicators()

    def _update_curse_indicators(self):
        """Update curse slot labels to show which ones NEED to be filled"""
        effect_labels_base = ['Effect 1', 'Effect 2', 'Effect 3',
                              'Curse 1', 'Curse 2', 'Curse 3']

        for i in range(3):
            effect_idx = i
            curse_idx = i + 3

            try:
                effect_id = int(self.effect_entries[effect_idx].get())
                curse_id = int(self.effect_entries[curse_idx].get())
            except ValueError:
                effect_id = 0
                curse_id = 0

            # Check if this effect needs a curse
            needs_curse = False
            has_curse = curse_id not in [0, -1, 4294967295]

            if effect_id not in [0, -1, 4294967295]:
                needs_curse = data_source.effect_needs_curse(effect_id)

            # Update the curse slot label
            base_label = effect_labels_base[curse_idx]
            if needs_curse and not has_curse:
                # Needs curse but doesn't have one - show warning
                self.slot_labels[curse_idx].config(
                    text=f"⚠️ {base_label} (REQUIRED):",
                    foreground='red'
                )
            elif needs_curse and has_curse:
                # Needs curse and has one - show satisfied
                self.slot_labels[curse_idx].config(
                    text=f"✓ {base_label}:",
                    foreground='green'
                )
            elif not needs_curse and has_curse:
                # Doesn't need curse but has one - ILLEGAL
                self.slot_labels[curse_idx].config(
                    text=f"⛔ {base_label} (ILLEGAL - remove curse):",
                    foreground='red'
                )
            else:
                # Doesn't need curse and doesn't have one - correct
                self.slot_labels[curse_idx].config(
                    text=f"{base_label} (not needed):",
                    foreground='gray'
                )
    
    def search_items(self):
        """Open search dialog for items"""
        _items = {}
        if self.safe_mode_var.get():
            _safe_range = data_source.get_safe_relic_ids() if relic_checker else []
            _df = data_source.relic_table.copy()
            _df = _df[_df.index.isin(_safe_range)]
            _items = data_source.cvrt_filtered_relic_origin_structure(_df)
        else:
            _items = items_json
        SearchDialog(self.dialog, self.item_id, "relics", _items, "Select Relic", self.on_item_selected)

    def find_valid_relic_ids(self, relic_id, effects, color):
        """Find all valid relics ID that matches the current effects configuration"""

        # Count how many effect slots are used (non-empty)
        effect_count = sum(1 for e in effects[:3] if e not in [0, -1, 4294967295])

        # Count how many curses are NEEDED based on the effects (not just present)
        # Effects that can only roll on 3-effect relics need curses
        curses_needed = sum(1 for e in effects[:3]
                            if e not in [0, -1, 4294967295] and data_source.effect_needs_curse(e))

        # Also count curses that are present (for validation)
        curses_present = sum(1 for e in effects[3:] if e not in [0, -1, 4294967295])

        # Check if current relic is a deep relic (ID range 2000000-2019999)
        is_current_deep = relic_checker.is_deep_relic(relic_id) if relic_checker else False

        # Search for valid relics with matching structure

        # Skip illegal range
        # Check if within valid range
        # Deep relics must stay deep, normal relics must stay normal
        # Check if this relic can accommodate our effects
        # Must have enough curse slots for effects that NEED curses
        # Must also have enough curse slots for curses that ARE present
        # Check color match (if we have a color preference)
        relic_table = data_source.get_filtered_relics_df(color, is_current_deep,
                                                         effect_count, curses_needed)
        relic_table.set_index("ID", inplace=True)
        valid_candidates = []

        for relic_id, row in relic_table.iterrows():

            # Get pool configuration for this relic
            pools = [
                row["attachEffectTableId_1"],
                row["attachEffectTableId_2"],
                row["attachEffectTableId_3"],
                row["attachEffectTableId_curse1"],
                row["attachEffectTableId_curse2"],
                row["attachEffectTableId_curse3"],
            ]

            # Check if effects are valid WITH rearrangement (like the game does)
            # Use require_curses_present=False so we can find relics where curses CAN be added
            invalid_reason = relic_checker.check_invalidity(relic_id, effects) if relic_checker else None
            if not is_curse_invalid(invalid_reason):
                if not relic_checker.is_strict_invalid(relic_id, effects, invalid_reason):
                    valid_candidates.append(relic_id)
        return valid_candidates

    def find_valid_relic_id(self):
        """Find a valid relic ID that matches the current effects configuration"""
        try:
            # Get current effects from the entry fields
            current_effects = []
            for entry in self.effect_entries:
                try:
                    val = int(entry.get())
                    current_effects.append(val)
                except ValueError:
                    current_effects.append(4294967295)  # Empty

            # Count how many effect slots are used (non-empty)
            effect_count = sum(1 for e in current_effects[:3] if e not in [0, -1, 4294967295])

            # Count how many curses are NEEDED based on the effects (not just present)
            # Effects that can only roll on 3-effect relics need curses
            curses_needed = sum(1 for e in current_effects[:3]
                                if e not in [0, -1, 4294967295] and data_source.effect_needs_curse(e))

            # Also count curses that are present (for validation)
            curses_present = sum(1 for e in current_effects[3:] if e not in [0, -1, 4294967295])

            # Get current relic's color and deep status
            current_relic_id = int(self.item_id_entry.get())
            try:
                current_color = data_source.relic_table.loc[current_relic_id, "relicColor"]
            except KeyError:
                current_color = None
            valid_candidates = self.find_valid_relic_ids(current_relic_id, current_effects, current_color)

            # Determine which positions NEED curse slots
            curse_positions_needed = []
            for i, eff in enumerate(current_effects[:3]):
                if eff not in [0, -1, 4294967295] and data_source.effect_needs_curse(eff):
                    curse_positions_needed.append(i + 1)  # 1-indexed for display

            if valid_candidates:
                # Pick the first valid candidate and update the entry
                best_match = valid_candidates[0]
                self.item_id_var.set(str(best_match))

                # Get the name for display
                relic_name = items_json.get(str(best_match), {}).get("name", "Unknown")
                relic_color = items_json.get(str(best_match), {}).get("color", "Unknown")

                # Get the new relic's curse slot configuration
                new_pools = data_source.get_relic_pools_seq(best_match)
                curse_slots_info = []
                for i in range(3):
                    if new_pools[i + 3] != -1:
                        curse_slots_info.append(f"Slot {i+1}: enabled")
                    else:
                        curse_slots_info.append(f"Slot {i+1}: disabled")

                messagebox.showinfo(
                    "Valid ID Found",
                    f"Found valid relic ID: {best_match}\n"
                    f"Name: {relic_name}\n"
                    f"Color: {relic_color}\n\n"
                    f"Curse slots: {', '.join(curse_slots_info)}"
                )
                self.update_debug_info()
            else:
                # Build detailed requirements message
                positions_str = ", ".join(str(p) for p in curse_positions_needed) if curse_positions_needed else "none"

                messagebox.showwarning(
                    "No Valid ID Found",
                    f"Could not find a valid relic ID that:\n"
                    f"- Has the same color\n"
                    f"- Has effects in pools at exact positions\n"
                    f"- Has curse slots at positions: {positions_str}\n\n"
                    f"Your effects that need curses are in slot(s): {positions_str}\n"
                    f"No relic of this color has curse slots in those exact positions.\n\n"
                    f"Options:\n"
                    f"- Try a different color\n"
                    f"- Rearrange effects so curse-needing ones are in slots 1-2"
                )

        except Exception as e:
            messagebox.showerror("Error", f"Failed to find valid ID: {str(e)}")

    def _auto_find_valid_relic_id(self):
        """Silently find and apply a valid relic ID when effects change.

        This runs in the background without showing dialogs, automatically
        updating the relic ID when needed (e.g., when an effect that needs
        a curse is selected).
        """
        try:
            # Get current effects from the entry fields
            current_effects = []
            for entry in self.effect_entries:
                try:
                    val = int(entry.get())
                    current_effects.append(val)
                except ValueError:
                    current_effects.append(4294967295)  # Empty

            # Get current relic ID and check if it's already valid
            current_relic_id = int(self.item_id_entry.get())

            # Check if current relic is already valid for these effects
            try:
                current_pools = data_source.get_relic_pools_seq(current_relic_id)
                invalid_reason = relic_checker.check_invalidity(current_relic_id, current_effects)
                if not invalid_reason and relic_checker.is_strict_invalid(current_relic_id, current_effects, invalid_reason):
                    # Current relic is fine, no change needed
                    return
            except (KeyError, IndexError):
                pass  # Current relic invalid, continue to find new one

            # Get current relic's color
            try:
                current_color = data_source.relic_table.loc[current_relic_id, "relicColor"]
            except KeyError:
                current_color = None

            
            valid_candidates = self.find_valid_relic_ids(current_relic_id, current_effects, current_color)

            if valid_candidates:
                # Silently apply the first valid candidate
                best_match = valid_candidates[0]
                self.item_id_var.set(str(best_match))
                self.update_debug_info()

        except Exception:
            # Silently fail - user can still use manual "Find Valid ID" button
            pass

    def _check_effects_valid_for_relic_exact(self, effects, pools):
        """Check if effects are valid for relic pools WITHOUT rearranging.

        This checks that:
        1. Each effect in position i is in pool i (or empty/pool is -1)
        2. Each curse in position i is in curse pool i (or empty/pool is -1)
        3. Effects that need curses have curse slots available in their position
        4. Curses present have valid curse pools in their position

        Returns True if the exact current arrangement works for this relic.
        """
        for idx in range(3):
            eff = effects[idx]
            curse = effects[idx + 3]
            effect_pool = pools[idx]
            curse_pool = pools[idx + 3]

            # Check effect placement
            if eff in [0, -1, 4294967295]:
                # Empty effect - OK
                pass
            elif effect_pool == -1:
                # Effect present but no pool available - invalid
                return False
            else:
                # Effect must be in its pool
                pool_effects = data_source.get_pool_effects_strict(effect_pool)
                if eff not in pool_effects:
                    return False

                # Check if this effect NEEDS a curse (deep-only effect)
                if data_source.effect_needs_curse(eff):
                    if curse_pool == -1:
                        # Effect needs curse but relic has no curse slot in this position
                        return False

            # Check curse placement
            if curse in [0, -1, 4294967295]:
                # Empty curse - OK
                pass
            elif curse_pool == -1:
                # Curse present but no pool available in this position - invalid
                return False
            else:
                # Curse must be in its pool
                pool_curses = data_source.get_pool_effects_strict(curse_pool)
                if curse not in pool_curses:
                    return False

        return True

    def _check_effects_valid_for_relic(self, relic_id, effects, pools, require_curses_present=False):
        """Check if the given effects are valid for the relic's effect pools.

        Args:
            relic_id: The relic ID to check
            effects: List of 6 effect IDs [e1, e2, e3, curse1, curse2, curse3]
            pools: List of 6 pool IDs for this relic
            require_curses_present: If True, effects that need curses MUST have curses.
                                   If False, just check that curse SLOTS exist.

        This checks:
        1. Each effect can be placed in at least one effect pool
        2. Each curse can be placed in at least one curse pool
        3. Effects that need curses have curse slots available
        4. Curses present have corresponding effect slots available
        """
        # Try all possible sequences to find one that works
        possible_sequences = [[0, 1, 2], [0, 2, 1], [1, 0, 2], [1, 2, 0],
                              [2, 0, 1], [2, 1, 0]]

        for seq in possible_sequences:
            sequence_valid = True
            reordered_effects = [effects[i] for i in seq]
            reordered_curses = [effects[i + 3] for i in seq]

            for idx in range(3):
                eff = reordered_effects[idx]
                curse = reordered_curses[idx]
                effect_pool = pools[idx]
                curse_pool = pools[idx + 3]

                # Check effect placement
                if eff in [0, -1, 4294967295]:
                    # Empty effect - OK
                    pass
                elif effect_pool == -1:
                    # Effect present but no pool available - invalid
                    sequence_valid = False
                    break
                else:
                    # Effect must be in its pool
                    pool_effects = data_source.get_pool_effects_strict(effect_pool)
                    if eff not in pool_effects:
                        sequence_valid = False
                        break

                    # Check if this effect NEEDS a curse (deep-only effect)
                    if data_source.effect_needs_curse(eff):
                        if curse_pool == -1:
                            # Effect needs curse but relic has no curse slot here
                            sequence_valid = False
                            break
                        # Only require curse to be present if flag is set
                        if require_curses_present and curse in [0, -1, 4294967295]:
                            # Effect needs curse but none provided
                            sequence_valid = False
                            break

                # Check curse placement
                if curse in [0, -1, 4294967295]:
                    # Empty curse - OK (unless effect needed one, checked above)
                    pass
                elif curse_pool == -1:
                    # Curse present but no pool available - invalid
                    sequence_valid = False
                    break
                else:
                    # Curse must be in its pool
                    pool_curses = data_source.get_pool_effects_strict(curse_pool)
                    if curse not in pool_curses:
                        sequence_valid = False
                        break

            if sequence_valid:
                return True

        return False

    def change_relic_color(self, target_color):
        """Change the relic to a different color while keeping effects legal"""
        color_names = {0: "Red", 1: "Blue", 2: "Yellow", 3: "Green"}
        target_color_name = color_names.get(target_color, "Unknown")

        # Guard: Check if relic is assigned to a character or preset
        assigned_to_hero_type = loadout_handler.relic_ga_hero_map.get(self.ga_handle, [])
        assigned_to = [data_source.character_names[h_t-1] for h_t in assigned_to_hero_type]
        if assigned_to:
            messagebox.showwarning(
                "Cannot Change Color",
                f"This relic is assigned to: {', '.join(assigned_to)}\n\n"
                f"Changing the color would break the assignment.\n"
                f"Please unequip the relic first before changing its color."
            )
            return

        try:
            # Get current effects from the entry fields
            current_effects = []
            for entry in self.effect_entries:
                try:
                    val = int(entry.get())
                    current_effects.append(val)
                except ValueError:
                    current_effects.append(4294967295)  # Empty

            # Count how many effect slots are used (non-empty)
            effect_count = sum(1 for e in current_effects[:3] if e not in [0, -1, 4294967295])

            curse_count = sum(1 for e in current_effects[3:] if e not in [0, -1, 4294967295])

            # Get current relic's color to check if already the target
            current_relic_id = int(self.item_id_entry.get())
            try:
                current_color = data_source.relic_table.loc[current_relic_id, "relicColor"]
                if current_color == target_color:
                    messagebox.showinfo("Info", f"Relic is already {target_color_name}!")
                    return
            except KeyError:
                pass

            valid_candidates = self.find_valid_relic_ids(current_relic_id, current_effects, target_color)

            if valid_candidates:
                # Pick the first valid candidate and update the entry
                best_match = valid_candidates[0]
                self.item_id_var.set(str(best_match))

                # Get the name for display
                relic_name = items_json.get(str(best_match), {}).get("name", "Unknown")

                messagebox.showinfo(
                    "Color Changed",
                    f"Changed to {target_color_name}!\n\n"
                    f"New Relic ID: {best_match}\n"
                    f"Name: {relic_name}\n\n"
                    f"Click 'Apply Changes' to save."
                )
                self.update_debug_info()
            else:
                messagebox.showwarning(
                    "Cannot Change Color",
                    f"Could not find a valid {target_color_name} relic that:\n"
                    f"- Supports {effect_count} effect slot(s)\n"
                    f"- Supports {curse_count} curse slot(s)\n"
                    f"- Has valid pools for all current effects\n\n"
                    f"Try removing some effects first, or the effects may not exist in {target_color_name} relics."
                )

        except Exception as e:
            messagebox.showerror("Error", f"Failed to change color: {str(e)}")

    def search_effects(self, effect_index):
        """Open search dialog for effects"""
        _items = {}
        is_curse_slot = effect_index >= 3

        if self.safe_mode_var.get():
            try:
                _cut_relic_id = int(self.item_id_entry.get())
            except ValueError:
                messagebox.showerror("Error", "Invalid relic ID in entry field")
                return

            try:
                _effects = [int(entry.get()) for entry in self.effect_entries]
                _pools = data_source.get_adjusted_pool_sequence(_cut_relic_id, _effects)
                _pool_id = _pools[effect_index]
            except (KeyError, IndexError, ValueError) as e:
                messagebox.showerror("Error", f"Could not get pool for relic {_cut_relic_id}: {e}")
                return

            _pool_effects = data_source.get_pool_rollable_effects(_pool_id)

            # For curse slots (index >= 3), if this specific pool is disabled,
            # use ALL available curse pools combined (game rearranges internally)
            if not _pool_effects and is_curse_slot:
                # Combine effects from all available curse pools
                all_curse_effects = set()
                for i in range(3):
                    curse_pool = _pools[3 + i]
                    if curse_pool != -1:
                        pool_effects = data_source.get_pool_rollable_effects(curse_pool)
                        all_curse_effects.update(pool_effects)
                _pool_effects = list(all_curse_effects)

            if not _pool_effects:
                # Slot is disabled and no alternatives available
                slot_type = "effect" if effect_index < 3 else "curse"
                slot_num = (effect_index % 3) + 1
                messagebox.showinfo(
                    "Slot Disabled",
                    f"This relic has no {slot_type} slots available.\n"
                    f"Try finding a different relic ID with 'Find Valid ID'."
                )
                return

            _effect_params_df = data_source.effect_params.copy()
            _effect_params_df = _effect_params_df[_effect_params_df.index.isin(_pool_effects)]
            match effect_index:
                case 1:
                    _effect_id_1 = int(self.effect_entries[0].get())
                    _conflic_id_1 = data_source.get_effect_conflict_id(_effect_id_1)
                    _effect_params_df = _effect_params_df[
                        (_effect_params_df["compatibilityId"] == -1) |
                        (_effect_params_df["compatibilityId"] != _conflic_id_1)
                    ]
                case 2:
                    _effect_id_1 = int(self.effect_entries[0].get())
                    _conflic_id_1 = data_source.get_effect_conflict_id(_effect_id_1)
                    _effect_id_2 = int(self.effect_entries[1].get())
                    _conflic_id_2 = data_source.get_effect_conflict_id(_effect_id_2)
                    _effect_params_df = _effect_params_df[
                        (_effect_params_df["compatibilityId"] == -1) |
                        ((_effect_params_df["compatibilityId"] != _conflic_id_1) &
                         (_effect_params_df["compatibilityId"] != _conflic_id_2))
                    ]
                case 4:
                    _effect_id_4 = int(self.effect_entries[3].get())
                    _conflic_id_4 = data_source.get_effect_conflict_id(_effect_id_4)
                    _effect_params_df = _effect_params_df[
                        (_effect_params_df["compatibilityId"] == -1) |
                        (_effect_params_df["compatibilityId"] != _conflic_id_4)
                    ]
                case 5:
                    _effect_id_4 = int(self.effect_entries[3].get())
                    _conflic_id_4 = data_source.get_effect_conflict_id(_effect_id_4)
                    _effect_id_5 = int(self.effect_entries[4].get())
                    _conflic_id_5 = data_source.get_effect_conflict_id(_effect_id_5)
                    _effect_params_df = _effect_params_df[
                        (_effect_params_df["compatibilityId"] == -1) |
                        ((_effect_params_df["compatibilityId"] != _conflic_id_4) &
                         (_effect_params_df["compatibilityId"] != _conflic_id_5))
                    ]
            _items = data_source.cvrt_filtered_effect_origin_structure(_effect_params_df)
        else:
            _items = effects_json

        # For curse slots, add "No Curse (Empty)" option at the top
        if is_curse_slot:
            _items = {"4294967295": {"name": "🚫 No Curse (Empty)"}, **_items}

        # Build dialog title with relic type info in safe mode
        if self.safe_mode_var.get():
            relic_type, _, _ = data_source.get_relic_type_info(_cut_relic_id)
            slot_type = "Curse" if is_curse_slot else "Effect"
            dialog_title = f"Select {slot_type} {(effect_index % 3) + 1} — {relic_type} Pools"
        else:
            dialog_title = f"Select Effect {effect_index + 1} — All Effects (Unsafe)"

        SearchDialog(self.dialog, self.item_id, "effects", _items, dialog_title,
                    lambda item_id: self.on_effect_selected(effect_index, item_id))
    
    def on_item_selected(self, item_id):
        """Callback when item is selected from search"""
        self.item_id_var.set(str(item_id))
    
    def on_effect_selected(self, effect_index, effect_id):
        """Callback when effect is selected from search"""
        self.effect_entries[effect_index].delete(0, tk.END)
        self.effect_entries[effect_index].insert(0, str(effect_id))
        self.on_effect_change(effect_index)
    
    def apply_changes(self):
        # Extract effect IDs from entries
        global relic_checker
        new_effects = []

        for entry in self.effect_entries:
            try:
                value = int(entry.get())
                new_effects.append(value)
            except ValueError:
                new_effects.append(0)
        new_effects = relic_checker.sort_effects(new_effects)

        # Check if item ID was changed
        new_item_id = None
        try:
            entered_id = int(self.item_id_entry.get())
            if entered_id != self.item_id:
                new_item_id = entered_id
        except ValueError:
            pass  # Keep original ID if invalid entry

        # Block deep/normal type mismatch - never allow changing between types
        original_is_deep = 2000000 <= self.item_id <= 2019999
        entered_is_deep = 2000000 <= entered_id <= 2019999
        if original_is_deep != entered_is_deep:
            messagebox.showerror(
                "Invalid Operation",
                f"Cannot change from a {'Deep' if original_is_deep else 'Normal'} relic "
                f"to a {'Deep' if entered_is_deep else 'Normal'} relic ID.\n\n"
                "Deep and Normal relics have different effect pools and are not interchangeable."
            )
            return

        invalid_reason = relic_checker.check_invalidity(entered_id, new_effects)
        is_curse_illegal = is_curse_invalid(invalid_reason)

        # Warn user if the relic will be invalid
        if invalid_reason:
            reason_text = {
                InvalidReason.IN_ILLEGAL_RANGE: "Relic ID is in the illegal range (20000-30035)",
                InvalidReason.INVALID_ITEM: "Relic ID is not in valid range",
                InvalidReason.EFF_NOT_IN_ROLLABLE_POOL: "One or more effects cannot roll on this relic",
                InvalidReason.EFF_MUST_EMPTY: "Effect slot must be empty for this relic",
                InvalidReason.EFF_CONFLICT: "Effects have conflicting IDs",
                InvalidReason.CURSE_NOT_IN_ROLLABLE_POOL: "One or more curses cannot roll on this relic",
                InvalidReason.CURSE_MUST_EMPTY: "Curse slot must be empty for this relic",
                InvalidReason.CURSE_REQUIRED_BY_EFFECT: "Effect requires a curse but none provided",
                InvalidReason.CURSE_CONFLICT: "Curses have conflicting IDs",
                InvalidReason.CURSES_NOT_ENOUGH: "Not enough curses for the effects",
                InvalidReason.EFFS_NOT_SORTED: "Effects are not properly sorted",
            }.get(invalid_reason, f"Unknown reason ({invalid_reason})")

            warning_msg = (
                f"This relic configuration will be INVALID:\n\n"
                f"• {reason_text}\n\n"
                "Invalid relics may cause issues in game.\n"
                "Do you want to save anyway?"
            )
            if not messagebox.askyesno("Invalid Relic Warning", warning_msg, icon='warning'):
                return
            
        # Apply modifications
        if modify_relic(self.ga_handle, self.item_id, new_effects, new_item_id):
            messagebox.showinfo("Success", "Relic modified successfully")
            self.callback()
            # Update current item_id if it was changed
            if new_item_id is not None:
                self.item_id = new_item_id
            # Reload the current relic to show updated values
            self.load_relic(self.ga_handle, self.item_id)
        else:
            messagebox.showerror("Error", "Failed to modify relic")


class SearchDialog:
    """Search dialog for JSON items"""
    def __init__(self, parent, item_id, search_type, json_data, title, callback):
        self.json_data = json_data
        self.callback = callback
        self.search_type = search_type
        self.item_id = item_id
        
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry("600x500")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        self.setup_ui()
    
    def setup_ui(self):
        # Search entry
        search_frame = ttk.Frame(self.dialog)
        search_frame.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(search_frame, text="Search:").pack(side='left', padx=5)
        self.search_var = tk.StringVar()
        self.search_var.trace('w', lambda *args: self.filter_results())
        
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var)
        search_entry.pack(side='left', fill='x', expand=True, padx=5)
        search_entry.focus()
        
        # Warp Result frame and Filter frame
        warp_frame = ttk.Frame(self.dialog)
        warp_frame.pack(fill='both',expand=True, padx=10, pady=5)
        
        # Results listbox
        results_frame = ttk.Frame(warp_frame)
        results_frame.pack(side='left', fill='both', expand=True, padx=10, pady=5)
        
        scrollbar = ttk.Scrollbar(results_frame)
        scrollbar.pack(side='right', fill='y')
        
        self.listbox = tk.Listbox(results_frame, yscrollcommand=scrollbar.set)
        self.listbox.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=self.listbox.yview)
        
        self.listbox.bind('<Double-Button-1>', self.on_select)
        
        if self.search_type == "relics":
            # Filter Frame
            filter_frame = ttk.Frame(warp_frame)
            filter_frame.pack(side='left', fill='both', padx=10, pady=5)
            color_row = ttk.Frame(filter_frame)
            color_row.pack(fill='x')
            self.lock_color_var = tk.BooleanVar(value=True)
            checkbox_lock_color = ttk.Checkbutton(color_row, variable=self.lock_color_var,
                                                onvalue=True, offvalue=False, text="Lock color:")
            cur_color = data_source.get_relic_color(self.item_id)
            
            self.color_var = tk.StringVar(value=cur_color)
            self.color_int_var = 0
            
            def color_map_to_int():
                self.color_int_var = COLOR_MAP.index(self.color_var.get())
                
            combobox_color = ttk.Combobox(color_row, textvariable=self.color_var, values=["Red", "Blue", "Yellow", "Green"],
                                          width=10, state="readonly")
            
            checkbox_lock_color.pack(side='left', padx=5)
            combobox_color.pack(side='left', padx=5)
            
            self.lock_color_var.trace('w', lambda *args: self.filter_results())
            self.color_var.trace('w', lambda *args: self.filter_results())
            self.color_var.trace_add('write', color_map_to_int)
            
            # Relic Type Row
            type_row = ttk.Frame(filter_frame)
            type_row.pack(fill='x', padx=5, pady=5)
            ttk.Label(type_row, text="Relic Type:").pack(side='left', padx=(19, 5))
            self.relic_type_var = tk.StringVar(value="All")
            combobox_type = ttk.Combobox(type_row, textvariable=self.relic_type_var, values=["All", "Deep", "Normal"],
                                         width=10, state="readonly")
            combobox_type.pack(side='left', padx=5)
            self.relic_type_var.trace('w', lambda *args: self.filter_results())
            
            # Structure filters
            ttk.Label(filter_frame, text="Effect Slots:").pack(anchor='w', pady=(10, 0))
            self.effect_slots_var = tk.StringVar(value="Any")
            ttk.Radiobutton(filter_frame, text="Any", variable=self.effect_slots_var, value="Any", command=self.filter_results).pack(anchor='w')
            ttk.Radiobutton(filter_frame, text="1 Effect", variable=self.effect_slots_var, value="1", command=self.filter_results).pack(anchor='w')
            ttk.Radiobutton(filter_frame, text="2 Effects", variable=self.effect_slots_var, value="2", command=self.filter_results).pack(anchor='w')
            ttk.Radiobutton(filter_frame, text="3 Effects", variable=self.effect_slots_var, value="3", command=self.filter_results).pack(anchor='w')

            ttk.Label(filter_frame, text="Curse Slots:").pack(anchor='w', pady=(10, 0))
            self.curse_slots_var = tk.StringVar(value="Any")
            ttk.Radiobutton(filter_frame, text="Any", variable=self.curse_slots_var, value="Any", command=self.filter_results).pack(anchor='w')
            ttk.Radiobutton(filter_frame, text="0 Curses", variable=self.curse_slots_var, value="0", command=self.filter_results).pack(anchor='w')
            ttk.Radiobutton(filter_frame, text="1 Curse", variable=self.curse_slots_var, value="1", command=self.filter_results).pack(anchor='w')
            ttk.Radiobutton(filter_frame, text="2 Curses", variable=self.curse_slots_var, value="2", command=self.filter_results).pack(anchor='w')
            ttk.Radiobutton(filter_frame, text="3 Curses", variable=self.curse_slots_var, value="3", command=self.filter_results).pack(anchor='w')
            
        # Populate initial results
        self.all_items = []
        for item_id, item_data in self.json_data.items():
            name = item_data.get('name', 'Unknown')
            item_str = ""
            if self.search_type == "effects":
                item_str = name
            elif self.search_type == "relics":
                relic_slot = data_source.get_relic_slot_count(int(item_id))
                item_str = f"{name} (effects: {relic_slot[0]}, curses:{relic_slot[1]})"
            self.all_items.append((item_id, item_str))
        
        self.all_items.sort(key=lambda x: int(x[0]) if x[0].isdigit() else 0)
        self.filter_results()
        
        # Buttons
        button_frame = ttk.Frame(self.dialog)
        button_frame.pack(fill='x', padx=10, pady=10)
        
        ttk.Button(button_frame, text="Select", command=self.on_select).pack(side='right', padx=5)
        ttk.Button(button_frame, text="Cancel", command=self.dialog.destroy).pack(side='right', padx=5)
    
    def filter_results(self):
        search_term = self.search_var.get().lower()
        
        self.listbox.delete(0, tk.END)
        
        for item_id, name in self.all_items:
            if self.search_type == "relics":
                # filter by color
                if self.lock_color_var.get() and data_source.get_relic_color(int(item_id)) != self.color_var.get():
                    continue
                eff_slots, curse_slots = data_source.get_relic_slot_count(int(item_id))
                # filter by relic type
                if self.relic_type_var.get() != "All":
                    if self.relic_type_var.get() == "Normal" and relic_checker.is_deep_relic(int(item_id)):
                        continue
                    if self.relic_type_var.get() == "Deep" and not relic_checker.is_deep_relic(int(item_id)):
                        continue
                # filter by effect slots
                if self.effect_slots_var.get() != "Any":
                    if self.effect_slots_var.get() != str(eff_slots):
                        continue
                # filter by curse slots
                if self.curse_slots_var.get() != "Any":
                    if self.curse_slots_var.get() != str(curse_slots):
                        continue
            
            if search_term in name.lower() or search_term in item_id:
                self.listbox.insert(tk.END, f"{name} (ID: {item_id})")
    
    def on_select(self, event=None):
        selection = self.listbox.curselection()
        if not selection:
            return
        
        selected_text = self.listbox.get(selection[0])
        item_id = selected_text.split("ID: ")[1].rstrip(")")
        
        self.callback(int(item_id))
        self.dialog.destroy()


def main():
    root = tk.Tk()
    app = SaveEditorGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()

